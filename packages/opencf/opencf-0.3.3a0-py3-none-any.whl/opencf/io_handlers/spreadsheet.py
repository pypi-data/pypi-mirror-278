"""
Spreadsheet I/O Handlers

This module provides classes for reading and writing spreadsheet files using the pandas library. It includes abstract base classes
and concrete implementations for converting between spreadsheet files and pandas DataFrame objects.
"""

from pathlib import Path
from typing import Any, Dict, List

import openpyxl
import pandas as pd
from opencf_core.io_handler import Reader, Writer
from openpyxl.utils import get_column_letter

from ..config.logging import logger


class XlsxToDictReader(Reader):
    """
    Reads content from an XLSX file and returns it as a list of dictionaries.

    Example:
        >>> reader = XlsxToDictReader()
        >>> content = reader.read(Path('input.xlsx'))
        >>> print(content)
        [{'name': 'John', 'age': 30}, {'name': 'Jane', 'age': 25}]
    """

    input_format = List[Dict[str, Any]]

    def _check_input_format(self, content: List[Dict[str, Any]]) -> bool:
        """
        Validates the input content to ensure it is a list of dictionaries.

        Args:
            content (List[Dict[str, Any]]): The content to validate.

        Returns:
            bool: True if the content is a list of dictionaries, False otherwise.
        """
        return isinstance(content, list) and all(
            isinstance(row, dict) for row in content
        )

    def _read_content(self, input_path: Path) -> List[Dict[str, Any]]:
        """
        Reads and parses the content from the XLSX file at the given path.

        Args:
            input_path (Path): The path to the XLSX file.

        Returns:
            List[Dict[str, Any]]: The parsed content as a list of dictionaries.
        """
        workbook = openpyxl.load_workbook(input_path)
        sheet = workbook.active
        if sheet is None:
            return []
        headers = [cell.value for cell in sheet[1]]
        rows = [
            {headers[i]: cell.value for i, cell in enumerate(row)}
            for row in sheet.iter_rows(min_row=2)
        ]
        return rows


class DictToXlsxWriter(Writer):
    """
    Writes content from a list of dictionaries to an XLSX file.

    Example:
        >>> writer = DictToXlsxWriter()
        >>> content = [{'name': 'John', 'age': 30}, {'name': 'Jane', 'age': 25}]
        >>> writer.write(Path('output.xlsx'), content)
    """

    output_format = List[Dict[str, Any]]

    def _check_output_format(self, content: List[Dict[str, Any]]) -> bool:
        """
        Validates the output content to ensure it is a list of dictionaries.

        Args:
            content (List[Dict[str, Any]]): The content to validate.

        Returns:
            bool: True if the content is a list of dictionaries, False otherwise.
        """
        return isinstance(content, list) and all(
            isinstance(row, dict) for row in content
        )

    def _write_content(self, output_path: Path, content: List[Dict[str, Any]]) -> None:
        """
        Writes the list of dictionaries content to an XLSX file at the given path.

        Args:
            output_path (Path): The path to the XLSX file.
            content (List[Dict[str, Any]]): The list of dictionaries content to write.
        """

        if len(content) == 0:
            logger.error("no input found")
            return

        workbook = openpyxl.Workbook()

        if workbook.active is None:
            logger.error("no sheet found in the current workbook")
            return

        sheet: openpyxl._WorksheetOrChartsheetLike = workbook.active

        headers = content[0].keys()
        sheet.append(dict(headers))

        for row in content:
            sheet.append(list(row.values()))

        for col_num, col_title in enumerate(headers, 1):
            col_letter = get_column_letter(col_num)
            sheet.column_dimensions[col_letter].width = max(len(col_title), 15)

        workbook.save(output_path)


class SpreadsheetToPandasReader(Reader):
    """
    Reads a spreadsheet file and returns a pandas DataFrame object.
    """

    # input_format = pd.DataFrame

    def _check_input_format(self, content: pd.DataFrame) -> bool:
        """
        Validates if the provided content is a pandas DataFrame object.

        Args:
            content (pd.DataFrame): The content to be validated.

        Returns:
            bool: True if the content is a pandas DataFrame object, False otherwise.
        """
        return isinstance(content, pd.DataFrame)

    def _read_content(self, input_path: Path) -> pd.DataFrame:
        """
        Reads and returns the content from the given input path as a pandas DataFrame object.

        Args:
            input_path (Path): The path to the input spreadsheet file.

        Returns:
            pd.DataFrame: The pandas DataFrame object read from the input file.
        """
        return pd.read_excel(input_path)


class PandasToSpreadsheetWriter(Writer):
    """
    Writes a pandas DataFrame object to a spreadsheet file.
    """

    # output_format = pd.DataFrame

    def _check_output_format(self, content: pd.DataFrame) -> bool:
        """
        Validates if the provided content is a pandas DataFrame object.

        Args:
            content (pd.DataFrame): The content to be validated.

        Returns:
            bool: True if the content is a pandas DataFrame object, False otherwise.
        """
        return isinstance(content, pd.DataFrame)

    def _write_content(self, output_path: Path, output_content: pd.DataFrame):
        """
        Writes the provided pandas DataFrame object to the given output path as a spreadsheet file.

        Args:
            output_path (Path): The path to the output spreadsheet file.
            output_content (pd.DataFrame): The pandas DataFrame object to be written to the output file.
        """
        output_content.to_excel(output_path, index=False)
