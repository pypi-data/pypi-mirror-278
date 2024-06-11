#!/usr/bin/env python3
# -*- coding: utf-8 -*-
""" A class to import Fedramp V4 and V5 POAMS """
import re
from pathlib import Path
from typing import List, Optional, Union

from openpyxl import load_workbook, Workbook
from openpyxl.utils import column_index_from_string
from rich.console import Console

from regscale.core.app.application import Application
from regscale.core.app.utils.app_utils import (
    check_file_path,
    convert_datetime_to_regscale_string,
    create_progress_object,
)
from regscale.core.utils.date import datetime_obj
from regscale.integrations.integration.issue import IntegrationIssue
from regscale.models.regscale_models.issue import Issue


class POAM(IntegrationIssue):
    """
    Custom Integration issue class
    """

    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        file_path: Union[str, None] = kwargs.get("file_path", None)
        if not file_path:
            raise ValueError("File path is required")
        console = Console()
        self.file_path = Path(file_path)
        app = Application()
        self.app = app
        self.module = kwargs.get("module", "securityplans")
        self.module_id = kwargs.get("module_id", 0)
        self.console = console
        self.poam_data: List[Issue] = []
        data = self.import_poam()
        self.data = data
        self.create_or_update_issues(issues=self.poam_data, parent_id=self.module_id, parent_module=self.module)

    def pull(self):
        """
        Pull inventory from an Integration platform into RegScale
        """
        # Implement the pull method here
        pass

    def file_type(self):
        """
        A method to return the file type
        """
        file_type = None
        if self.file_path:
            file_type = self.file_path.suffix
        return file_type

    @staticmethod
    def get_index_from_column_name(column_name: str) -> int:
        """
        A method to get the index from a column name

        :param str column_name: A column name
        :return: The index of the column
        :rtype: int
        """
        return column_index_from_string(column_name) - 1

    def gen_issue_from_row(self, row: tuple, status: str, category: str) -> Optional[Issue]:
        """
        A method to generate an issue from a row


        :param tuple row: A row
        :param str status: The status of the issue
        :param str category: The category of the issue
        :return: An issue or None
        :rtype: Optional[Issue]
        """

        risk_rating = (
            row[self.get_index_from_column_name("T")]
            if row[self.get_index_from_column_name("T")]
            else row[self.get_index_from_column_name("S")]
        )
        cve = row[29]
        if not cve:
            return None
        plugin_id = row[self.get_index_from_column_name("F")]
        date_created = row[self.get_index_from_column_name("K")]
        date_last_updated = datetime_obj(row[self.get_index_from_column_name("O")])
        due_date = row[self.get_index_from_column_name("L")]
        issue = Issue(
            date_created=date_created,
            date_last_updated=date_last_updated,
            due_date=row[self.get_index_from_column_name("AC")],
            title=row[self.get_index_from_column_name("A")],
            description=row[self.get_index_from_column_name("D")],
            status=status,
            severity=Issue.assign_severity(risk_rating),
            category=category,
            owner=row[7],
            source=row[8],
            source_id=row[9],
            assetIdentifier=row[self.get_index_from_column_name("G")],
            isPoam=True,
            issueOwnerId=self.app.config["userId"],
            securityPlanId=self.module_id if self.module == "securityplans" else 0,
            cve=cve,
            pluginId=str(plugin_id),
            autoApproved="No",
            dueDate=convert_datetime_to_regscale_string(due_date),
            parentId=self.module_id,
            parentModule=self.module,
            basisForAdjustment=row[23],
            dateCompleted=convert_datetime_to_regscale_string(date_last_updated) if status == "Closed" else None,
        )
        return issue

    def import_poam(self) -> Workbook:
        """
        A method to import the POAM data

        :return: The workbook
        :rtype: Workbook
        """
        workbook = load_workbook(filename=self.file_path)
        sheets = workbook.sheetnames
        pattern = "POA&M Items"
        status = "Closed"

        poam_sheets = [item for item in sheets if re.search(pattern, item)]
        with create_progress_object() as progress:
            parsing_progress = progress.add_task("[#f8b737]Parsing data from workbook...", total=len(poam_sheets))
            for sheet in poam_sheets:
                ws = workbook[sheet]
                parsing_poams = progress.add_task(
                    f"[#ef5d23]Parsing '{sheet}' sheet for POAMs...", total=ws.max_row - 7
                )
                category = ws["C3"].value
                for _, row in enumerate(ws.iter_rows(min_row=8, max_row=ws.max_row, values_only=True)):
                    if "open" in sheet.lower():
                        status = "Open"
                    try:
                        issue = self.gen_issue_from_row(row=row, status=status, category=category)
                        if issue:
                            self.poam_data.append(issue)
                    except TypeError:
                        continue
                    progress.update(parsing_poams, advance=1)
                progress.update(parsing_progress, advance=1)
        return workbook
