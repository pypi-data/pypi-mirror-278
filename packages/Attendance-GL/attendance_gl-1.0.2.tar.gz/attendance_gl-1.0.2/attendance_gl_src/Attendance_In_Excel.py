import os
from openpyxl import Workbook, load_workbook
from datetime import date
import subprocess
import configparser
from pathlib import Path
import logging
import logging.config

# Load logging configuration from the file
logging.config.fileConfig(Path(__file__).resolve().parent / 'resources' / 'logging.conf', disable_existing_loggers=False)
logger = logging.getLogger(__name__)


def retrieve_network_info():
    # Load logging configuration from the file
    logger.info("############################### Marking Attendance ###############################")

    # get the directory path from the config file
    config = configparser.ConfigParser()
    config_path = Path(__file__).resolve().parent / 'resources' / 'config.properties'
    logger.info(f'config path: {config_path}')
    config.read(config_path)
    directory = config['data']['dir.path']
    fileName = config['data']['attendance_excel_name']
    office_networks_list = config['data']['possible.networks']


    # Ensure the directory exists
    if not os.path.exists(directory):
        logger.info(f"Creating directory: {directory}")
        os.makedirs(directory)

    # Now the file path should be correct
    file_path = os.path.join(directory, fileName)

    # using the check_output() for having the network term retrieval 
    devices = subprocess.check_output(['netsh', 'wlan', 'show', 'network']) 

    # decode it to strings 
    try:
        devices = devices.decode('utf-8')
    except UnicodeDecodeError as e:
        logger.error(f"Error in decoding bytes: {e}")
        # Handle the error by ignoring or replacing problematic bytes
        devices = devices.decode('utf-8', errors='replace')
    devices = devices.replace("\r", "") 
    logger.info(f'Networks found: {devices.title()}')

    # if any of the given name is present then you are in office
    office_network_names = list(map(str.strip, office_networks_list.split(',')))
    office_network = any(name.lower() in devices.lower() for name in office_network_names)

    if not os.path.exists(file_path):
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws.append(['Date', 'Office'])
    else:
        wb = load_workbook(filename=file_path)
        ws = wb.active

    # Check if today's date is already present and if column B value is True
    for row in ws.iter_rows(min_row=2, values_only=True):
        date_from_row = row[0].date() #fetch the date from default coming datetime object
        if date_from_row == date.today():
            if row[1] == True:
                logger.info("Attendance already marked for today as Present!")
                return
            else:
                logger.info("Attendance was marked as Absent! Deleting the row to update the status.")
                ws.delete_rows(ws.max_row)
                break
        
    ws.append([date.today(), office_network])
    logger.info(f"Attendance marked for {date.today()} as {'Present' if office_network else 'Absent'}")
    try:
        wb.save(file_path)
    except Exception as e:
        logger.error("Error saving the file")
        logger.error(e)
