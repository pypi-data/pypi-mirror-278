import logging

import openpyxl
from llama_index.core import (
    StorageContext,
    SimpleDirectoryReader,
    VectorStoreIndex,
    Settings,
)
from llama_index.vector_stores.redis import RedisVectorStore
from redis import Redis
from redisvl.schema import IndexSchema

from turing_planet.llama_index.embeddings.sparkai import SparkAIEmbedding
from turing_planet.llama_index.llms.sparkai import SparkAI

logging.basicConfig(
    level=logging.ERROR,
    format="%(asctime)s - %(levelname)s - %(message)s",
)

# 使用自定义embedding和llm
endpoint = "172.31.164.103:9980"
Settings.llm = SparkAI(endpoint=endpoint)
Settings.embed_model = SparkAIEmbedding(endpoint=endpoint, domain="emb_xfyun")
Settings.chunk_size = 1024


def vector_store():
    print("init redis store begin ...")
    redis_index_schema = IndexSchema.from_dict(
        {
            # customize basic index specs
            "index": {
                "name": "llamaindex-0131-2048-emb_v1",
                "prefix": "llamaindex-0131-2048-emb_v1_vector",
                "key_separator": ":",
            },
            # customize fields that are indexed
            "fields": [
                {"name": "id", "type": "tag"},
                {"name": "doc_id", "type": "tag"},
                {"name": "text", "type": "text"},
                {"name": "vector", "type": "vector",
                 "attrs": {"dims": 2560, "algorithm": "flat", "distance_metric": "cosine", }},
            ],
        }
    )

    redis_client = Redis.from_url("redis://172.31.128.153:6379")
    return RedisVectorStore(schema=redis_index_schema, redis_client=redis_client, overwrite=True)


def build_index(vector_store):
    # 加载data目录下文件，

    docPath = "/Users/wujian/Downloads/test/合同"
    print("load data begin ...")

    documents = SimpleDirectoryReader(docPath).load_data(show_progress=True)
    # redis补丁，处理不了NoneType类型数据, file_type=None
    for document in documents:
        if document.metadata:
            for key, value in document.metadata.items():
                if value is None:
                    document.metadata[key] = ""

    storage_context = StorageContext.from_defaults(vector_store=vector_store)

    # 创建索引，这一步会触发embedding操作，特征默认存储在本机内存中
    print("build index begin ...")
    index = VectorStoreIndex.from_documents(
        documents=documents,
        storage_context=storage_context,
        show_progress=True,
    )
    print("build index end.")


def query(vector_store):
    # 直接从向量库中加载
    index = VectorStoreIndex.from_vector_store(
        vector_store=vector_store, show_progress=True
    )

    # 执行查询，询问文档里面的内容
    query_engine = index.as_query_engine(similarity_top_k=1)

    # 读取Excel文件
    workbook = openpyxl.load_workbook('/Users/wujian/Downloads/test/合同query.xlsx')

    # 选择工作表
    sheet = workbook.active

    # 从第二行开始读取数据
    row = 2
    while sheet.cell(row=row, column=2).value is not None:
        q = sheet.cell(row=row, column=2).value
        result = query_engine.query(f"{q} \n(全部使用中文回答)")
        # print("\n\n---\n\n")
        print(f"Q：{row}.{q}")
        r = result.response.replace("<ret>", "\n").replace("<end>", "")
        print(row, r)
        # 将第三列的文本写入第8列
        # sheet.cell(row=row, column=14).value = r
        row += 1

    # 保存修改后的Excel文件
    # workbook.save('/Users/lyhu/project/llm2/data/0123/qb_bak_2048.xlsx')


if __name__ == "__main__":
    redis_store = vector_store()
    build_index(redis_store)
    query(redis_store)
