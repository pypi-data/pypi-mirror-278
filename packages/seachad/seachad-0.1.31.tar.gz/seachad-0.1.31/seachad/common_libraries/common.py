# -*- coding: utf-8 -*-
"""
Created on Fri Apr  3 08:00:24 2020

common
    Contiene llamadas a gestión de directorios y de fechas

v1
    trabaja con CM_API_ROBOT_miniconfig_base_v1
 
@author: Fernando
"""
  


import os
import sys

import traceback
import pprint # imprime las listas y diccionarios de manera que se pueden leer bien por pantalla

# ? CACHE/SESSION INFORMATION 
# sube a versión 2.2.0

# ? FIN DE CACHE/SESSION INFORMATION

from seachad.common_libraries.cache_REDIS import *
from seachad.common_libraries.cache_REDIS import _CACHE_set, _CACHE_get, _CACHE_delete

MIN_INTEGRITY_1 = 20
MIN_INTEGRITY_2 = 40
MIN_INTEGRITY_3 = 60

def INTERNAL_check_integrity():
    import datetime
    _k = get_cache("MGM_check_integrity", encrypt_info = True)
    if _k == None:
        return False
    else:
        _ts = datetime.datetime.now()
        if _ts-_k > datetime.timedelta(days = MIN_INTEGRITY_1):
            return False
        return True
    pass


# import ar_decorator_functions

import seachad.common_libraries.resolve_fstring as rfs
import seachad.common_libraries.terminal_colors as tc

# ---- .
# ---- *          DECORATOR -----------------------------------------------------------------------------------------------
# ---- .

# @ar_decorator_functions.catch_exceptions
def _resolve_dynamic_fstring(
                 
                 clave = None, # clave que quiero resolver (si la clave es un diccionario, resolverá todos sus elementos y devolverá un diccionario)
                 # context = None, # contexto de queries (contiene toda la información lógica y el acceso a access con la información ya descargada de ENV y de ENCRYPT data)
                 list_of_dictionaries = None, # orden de diccionarios que trataremos para resolver la clave, desde el primero al último (se pasa por clave y la función los resuelve)
                 current_dictionary = None, # diccionario donde quiero que empiece a buscar
                 # context_query = None, # lista de diccionarios guardada en context y que se recupera por el logical_name
                 params = {},
                 param_delimiter = "_PARAM_", # identifico, dentro de la estructura de definición de API (API_ROBOT_Providers_Microsoft, por ejemplo) que algo es un parámetro del API
                 escupe = 0, # escupo por pantalla info de rastreo?
                 force_until_all_resolved = False, # por defecto dejo que se devuelvan parámetros sin resolver (cuidado con este flag, si queda algo sin resolver contínuamente, puede dar un reventado)
                 pending_resolution = None, # si el parámetro force_until_all_resolved está a True, en la salida tiene que mirar si le sigue quedando lo mismo por resolver que cuando entró en la función, en ese caso cesa de intentar resolver
                 ):
    """ 
    llama a resolve_dynamic_fstring y si lo que se resuelve es un formato de fecha, se encarga de resolverlo antes de retornar
    """

    res, lparams = rfs._resolve_dynamic_fstring(
                clave = clave, # clave que quiero resolver (si la clave es un diccionario, resolverá todos sus elementos y devolverá un diccionario)
                # context = None, # contexto de queries (contiene toda la información lógica y el acceso a access con la información ya descargada de ENV y de ENCRYPT data)
                list_of_dictionaries = list_of_dictionaries, # orden de diccionarios que trataremos para resolver la clave, desde el primero al último (se pasa por clave y la función los resuelve)
                current_dictionary = current_dictionary, # diccionario donde quiero que empiece a buscar
                # context_query = None, # lista de diccionarios guardada en context y que se recupera por el logical_name
                params = params,
                param_delimiter = param_delimiter, # identifico, dentro de la estructura de definición de API (API_ROBOT_Providers_Microsoft, por ejemplo) que algo es un parámetro del API
                escupe = escupe, # escupo por pantalla info de rastreo?
                force_until_all_resolved = force_until_all_resolved, # por defecto dejo que se devuelvan parámetros sin resolver (cuidado con este flag, si queda algo sin resolver contínuamente, puede dar un reventado)
                pending_resolution = pending_resolution, # si el parámetro force_until_all_resolved está a True, en la salida tiene que mirar si le sigue quedando lo mismo por resolver que cuando entró en la función, en ese caso cesa de intentar resolver
                )

    """ resolución de fecha? """
    resultado = False
    if not res == None:
        resultado = dealWithDates(res)
    if not resultado == False: 
        res = resultado

    return res, lparams                 


from retrying import retry # permite retries de casi cualquier tipo de función (usado para ficheros, lo probaré en api_rest), source: https://pypi.org/project/retrying/
import filelock # bloqueos de ficheros source: https://pypi.org/project/filelock/

# ---- .
# ---- colores en pantalla
# ---- .

# =============================================================================
# ---- IMPRIMIR EN COLORES EN LA PANTALLA
# =============================================================================
# RED   = "\033[1;31m"  
# BLUE  = "\033[1;34m"
# CYAN  = "\033[1;36m"
# GREEN = "\033[1;32m"
# RESET = "\033[0;0m"
# BOLD    = "\033[;1m"
# REVERSE = "\033[;7m"

# Terminal color definitions

import os
import platform

# # permite ANSI Colorize en Windows y Powershell
# # source: https://gist.github.com/RDCH106/6562cc7136b30a5c59628501d87906f7
# if os.name == 'nt' and platform.release() == '10' and platform.version() >= '10.0.14393':
#     # Fix ANSI color in Windows 10 version 10.0.14393 (Windows Anniversary Update)
#     import ctypes
#     kernel32 = ctypes.windll.kernel32
#     kernel32.SetConsoleMode(kernel32.GetStdHandle(-11), 7)


ESC = '\033'


class fg:
    BLACK   = ESC+'[30m'
    RED     = ESC+'[31m'
    GREEN   = ESC+'[32m'
    YELLOW  = ESC+'[33m'
    BLUE    = ESC+'[34m'
    MAGENTA = ESC+'[35m'
    CYAN    = ESC+'[36m'
    WHITE   = ESC+'[37m'
    RESET   = ESC+'[39m'

class bg:
    BLACK   = ESC+'[40m'
    RED     = ESC+'[41m'
    GREEN   = ESC+'[42m'
    YELLOW  = ESC+'[43m'
    BLUE    = ESC+'[44m'
    MAGENTA = ESC+'[45m'
    CYAN    = ESC+'[46m'
    WHITE   = ESC+'[47m'
    RESET   = ESC+'[49m'

class style:
    BRIGHT    = ESC+'[1m'
    DIM       = ESC+'[2m'
    NORMAL    = ESC+'[22m'
    RESET     = ESC+'[0m'


# ACELERADORES - colores de salida
B = fg.BLUE + style.BRIGHT    
G = fg.GREEN + style.BRIGHT    
R = fg.RED + style.BRIGHT    
Y = fg.YELLOW + style.BRIGHT 
C = fg.CYAN + style.BRIGHT 
W = fg.WHITE + style.BRIGHT

# sin BRIGHT
b = fg.BLUE
g = fg.GREEN
r = fg.RED
y = fg.YELLOW
c = fg.CYAN
w = fg.WHITE

# fondos
# bb = bg.BLUE
# bg = bg.GREEN
# br = bg.RED
# by = bg.YELLOW
# bc = bg.CYAN
# bw = bg.WHITE


def print_in_color(col=style.RESET):
    tc. print_in_color(col)
    # try:
    #     sys.stdout.write(col)
    # except:
    #     sys.stdout.write(fg.RESET+bg.RESET+style.RESET)

# =============================================================================
# Se utilizaría en lugar de _print para localizarlo rápidamente en todo el proyecto y poder quitárselo de encima después
# =============================================================================
def __trace(message, color = fg.CYAN+style.BRIGHT, func = None, byscreen = True):
    _print(message, color, func, byscreen)
        
# =============================================================================
# Imprime con colores!!!!
# =============================================================================
def _print(message, color = fg.CYAN+style.BRIGHT, func = None, byscreen = True):

    tc._print(message, color, func, byscreen)    

    # if tracing_print == False:
    #     return
    
    # # if isinstance(message, dict):
    # #     message = pretty_dict(message) 
    
    # if byscreen: # puede que sólo quiera ejecutar la función y no pintar en la pantalla   
    #     m_color = color
    #     print_in_color(m_color)
    #     if not isinstance(message, (str, int)):
    #         pprint.pprint(message)
    #     else:
    #         _print(message)
    #     print_in_color()
    
    # # por si queremos grabar en disco, por ejemplo (habría que pasarle como parámetro _print("mensaje", Y, lambda: funcion(param1, param2, ...)))
    # if func:
    #     func()        
        
        
# ---- .
# ---- helpers --------------------------------------------------------------


# una sessiontag tiene que tener
# SESSION_NAME nombre de la sesison
# TIMESTAMP
# USER
        
        
def get_session_tag(session_name = "not_informed", session_user = "not_informed", sep = "@"):
    from datetime import datetime
    return f'{session_name}{sep}{session_user}{sep}{datetime.now().strftime("%Y-%m-%d_%H-%M-%S-%f")}'   

def _input(prompt = "continue? Y/N: ", exit_code = -1):
    """ hace un input de Y/N, todo lo que no sea N o n, continúa
    devuelve la elección 
    se puede personalizar el prompt del input"""
    a = input(prompt+": continue? Y/N: ")
    if a == "N" or a == "n" or a == "Q" or a == "q":
        sys.exit(exit_code)
    else:
        return a
    
def exit(exit_code):
    # input(exit_code = exit_code)    
    pass

# =============================================================================
# Devuelve un json bien impreso
# =============================================================================
def pretty_dict(message):
    return json.dumps(message, indent=4)


import hashlib
# =============================================================================
# Crea un identificador de sesión
# una sesión se compone de un context y un timestamp en la forma context+sep+timestamp
# context -> compañía|'usuario'
# timestamp -> timestamp en la forma "%Y-%m-%d_%H-%M-%S-%f"
# =============================================================================
def _create_session_tag(context = "not informed|'user'", sep = "@"):
    from datetime import datetime
    timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S-%f")
    retorno = f"{context}{sep}{timestamp}"
    return retorno, _create_hash_code(retorno)

def _create_hash_code(cadena = None):
    if not cadena == None:
        return hashlib.sha224(cadena.encode('utf-8')).hexdigest()
    return None

def _decode_hash_code(hascode = None):
    if not hascode== None:
        pass

# =============================================================================
# Devuelve el contexto de la session tag
# =============================================================================
def _get_context_from_session_tag(session_tag = None):
    pass

def _get_timestamp_from_session_tag(session_tag = None):
    pass




def test_print():
    # utilizando el HELPER
    _print("Hola mundo") # imprimirá por defecto en CYAN con BRIGHT (así sabemos que se está usando esta función por defecto)
    _print("Hola mundo", fg.BLUE+bg.RED+style.BRIGHT) # imprimirá en AZUL sobre ROJO en BRIGHT
    _print("Hola mundo") # vuelve a imprimir en CYAN con BRIGHT (así sabemos que se está usando esta función por defecto)
    
    # utilizando las primitivas
    print_in_color(fg.BLUE+bg.RED+style.BRIGHT)
    _print("Show me your color")    
    print_in_color(fg.WHITE)    
    _print( "All following prints will be WHITE ...")
    _print( "Esto sigue siendo WHITE")
    print_in_color()    
    _print( "Ahora normal")

# ----------------------------------------------
# CONFIGURACION INICIAL
# ----------------------------------------------
# access = {
#         #"API_ROBOT_MONGODB_HOST":"",

#         #"API_ROBOT_MONGODB_DB":"",
#         #"API_ROBOT_MONGODB_COLLECTION":"",

#         #"API_ROBOT_MONGODB_USER":"",
#         #"API_ROBOT_MONGODB_PASSWORD":"",
#         }

import builtins
try:
    access = builtins.access
except:
    access = {}


import seachad.common_libraries.IPL

# access = IPL.getAccess(access)
# ----------------------------------------------
# ----------------------------------------------

from dotenv import load_dotenv
if access.get("dot_env_path", None) is not None:
    load_dotenv(access["dot_env_path"]) # cargo el .env de la aplicación que llama


import seachad.common_libraries.ar_decorator_functions

# import ar_logging as arlg
tracing = True
tracing_print = True

# =============================================================================
# TIMESTAMPS
# =============================================================================

# formato por defecto para timestamp
str_timestamp = "%Y-%m-%d_%H:%M:%S.%f"

# devuelve en formato string el time de ahora mismo
def get_str_timestamp():
    from datetime import datetime
    timenow = datetime.now().strftime(str_timestamp)
    return timenow

# recrea y devuelve un objeto datetime enviándole una cadena que concida con el formato especificado
def get_obj_timestamp(string, formato = str_timestamp):
    from datetime import datetime
    timenow = datetime.now().strptime(string, formato)
    return timenow    


# =============================================================================
# guardo en formato "readable" el fichero json
# =============================================================================
import json
from pathlib import Path # esto debería servir para MACOS y para WINDOWS
from inspect import getargvalues, stack, getframeinfo, currentframe
# import LoggingAndTracking as lg # de momento tiene que estar en el directorio de trabajo (hay que ver como configurar para que lo coja de Github o de dónde corresponda)
import re
from datetime import datetime
import os
import sys
import copy # para hacer copias de diccionarios

# ---- .
# ---- * Resolución de fstrings dinámica --------------------------------------


# esta sección va a poder cogerse directamente desde IPL porque lo tiene allí cargado


# # crea f_strings dinámicas para resolver claves de diccionarios (del diccionario que recibe)
# # la orden de búsqueda es en el diccionario actual, en el diccionario padre o en access
# @ar_decorator_functions.catch_exceptions 
def make_magic(current_dictionary = None,# diccionario desde el que he obtenido el valor
               target_dictionary = None, # diccionario en el que buscar
               key = None, # clave sobre la que he operado en el diccionario original
               magic = None,
               param_delimiter = "_PARAM_"): # clave que buscamos en el diccionario target
    res = -1
    
    # tenemos que evitar que una clave se esté dirigiendo a sí misma, por ejemplo
    # "client_id" : "{client_id}"

    l_params = {}
    

    if "{" in magic and "}" in magic: # si está entre {} hay que resolver la magia -> tiene que haber dos (cuidado si algún valor viene con cosas raras como UUID, secrets o algo que pueda traer estos caracteres porque podría dar un error, aunque no creo que se resuelva)     
    # en el caso de que magic sea una lista de campos entre {} (es decir, que lleve varios campos), los sacamos todos y limpiamos
        # 20200912 -> arreglado el control de si la clave se encuentra como parámetro o dentro de la lista de parametros cuando se busca dentro del mismo diccionario
        # si la clave aparece en la lista, entonces no es resoluble en este nivel    
        import re
        start_delimiter = "{"
        end_delimiter = "}"
        regex_patron = r"{start_delimiter}[\w\d+-:]+{end_delimiter}".format(
            start_delimiter = start_delimiter,
            end_delimiter = end_delimiter) # buscamos secuencias de palabras dígitos, signos + y menos y : entre @
        # quitar
        # key = "scope_base"
        # magic = "{scope_base}/{authority_uri}/{scope_end}"
        # fin quitar
        cadena = magic
        comandos = re.findall(regex_patron, cadena)
        is_in = False
        # limpiamos la lista
        for comando in comandos:
            comando = comando.replace(start_delimiter, "")
            comando = comando.replace(end_delimiter, "")
            if key == comando:
                is_in = True # está dentro, no podemos resolver...
                break
        
        # if current_dictionary == target_dictionary and key in magic: # si estoy buscando en el diccionario actual y la clave se corresponde con lo que busco entre {} (quiero evitar que se crea que lo ha encontrado cuando, por ejemplo: "scope_base" : "{scope_base}")
        if current_dictionary == target_dictionary and is_in: # si estoy buscando en el diccionario actual y la clave se corresponde con lo que busco entre {} (quiero evitar que se crea que lo ha encontrado cuando, por ejemplo: "scope_base" : "{scope_base}")
            magic = None
            res = -1                        
        else:
            lmagic = magic
            magic = magic.replace("{", "{target_dictionary[\"")
            magic = magic.replace("}", "\"]}")
            # _print(magic)

            # intento recoger la lista de parámetros que puede venir            
            _,_,l_params = get_fstring_parameters(cadena = lmagic, 
                                                param_delimiter = param_delimiter,          
                                                start_delimiter = "{", 
                                                end_delimiter = "}", 
                                                identificador = "@", 
                                                limpiar_delimiters = False) 
            
            
            string = "f" + "'" + magic + "'"  # intento resolver un fstring con target_dictionary y la clave que me ha llegado
            try:     
                res = 0
                magic = eval(string)
                # si lo que obtengo contiene los delimiters, quiere decir que no es un valor, es algo a buscar en otro diccionario
                if "{" in magic and "}" in magic:
                    res = 2
                # _print(magic, CYAN)    
            except: # notificamos que no se ha resuelto dentro del traget_dictionary
                res = -1
#                 cambio el 23/10/2020 - quiero que magic contenga el valor de string, aunque no haya encontrado nada...
                magic = None
#                magic = string
                # _print(f"La clave {string} no existe", RED)
                pass                
    else: # como no está entre {} se supone que es un valor
        res = -1 # no hay cambio
        magic = current_dictionary[key]

    return res, magic, l_params

# =============================================================================
# # obtiene una lista de elementos contenidos entre los separadores (probablemente{}), se pueden limpiar los separadores (delimiters)
# # devuelve una lista conteniendo el elemento, y el identificador+elemento y, adicionalmente, la cadena original sustituyendo los elementos por identificador+elemento para posteriormente 
# # poder cambiar los valores    
# =============================================================================
# @ar_decorator_functions.catch_exceptions
def get_fstring_parameters(cadena = "{authority_host_uri}/_PARAM_{tenant_id}/_PARAM_{token_postfix}", # cadena que nos llega con elementos que deberían ser tratados como f-string
                           param_delimiter = "_PARAM_", # delimitador de parámetro para cargar la lista interna de parámetros de este API_call
                           start_delimiter = "{", # delimitador de comienzo de parámetro
                           end_delimiter = "}", # delimitador de fin de parámetro
                           identificador = "@", # en la cadena temporal de ubicación original del parámetro, qué voy a usar para identificarlo - produciría algo así @{authority_host_uri}/@{tenant_id}/@{token_postfix}
                           limpiar_delimiters = False, # cuando devuelvo los parámetros les quito los delimitadores? ({ y }, por ejemplo).
                           escupe = False,
                           ):
    """ devuelve una lista de parámetros (entendido como algo entre {}) en la forma:
        [parámetro, @+parametro]
        [@parametro]
        [_PARAM_parametro, _PARAM_@parametro]
        
        si se le manda limpiar_delimiters a True, quitará {} y _PARAM_
    """
    import re
    
    dict_params = {} # identificamos y devolvemos todo lo que sean _PARAM_, conteniendo su identificador y su valor
    
    cadena_cambiada = cadena # nos quedamos con la cadena original, donde sustituiremos lo que encontramos por algún tipo de identificador
    parameter = []
    lista_parameters = []

    if isinstance(cadena, str): # ojo, solo podemos operar con cadenas!
        _params = param_delimiter # obligamos a una búsqueda de _PARAM_
    
    
        """ _PARAM_ """
        start_delimiter_acumulado = f"{_params}{start_delimiter}"
        # buscamos primero con _PARAMS_, obtengo la información, y luego quito _PARAMS_
        patron_comando_params = r"{start_delimiter}[\w\d+-:]+{end_delimiter}".format(
                start_delimiter = start_delimiter_acumulado,
                end_delimiter = end_delimiter) # buscamos secuencias de palabras dígitos, signos + y menos y : entre @    
    
        regex_patron_params = patron_comando_params
        comandos = re.findall(regex_patron_params, cadena)
        if len(comandos)>0: # no es un comando
            for i in comandos:
                parameter = []
                if limpiar_delimiters:
                    i = i.replace(start_delimiter_acumulado,"")
                    i = i.replace(end_delimiter,"")
                else:
                    # construimos el identificador en la misma manera que se espera 
                    s_from = f"{_params}"
                    s_target = f"{_params}{identificador}"
                    i_con_identificador = i
                    i_con_identificador = i_con_identificador.replace(s_from, s_target)
                    # ahora le quito _params
                    i = i.replace(_params, "")
                    i_con_identificador = i_con_identificador.replace(_params, "")
                dict_params[i] = i_con_identificador # de momento lo dejamos vacío?
    
        """ LISTA DE COMANDOS """
        # obtenemos la lista de comandos    
        patron_comando = r"{start_delimiter}[\w\d+-:]+{end_delimiter}".format(
                start_delimiter = start_delimiter,
                end_delimiter = end_delimiter) # buscamos secuencias de palabras dígitos, signos + y menos y : entre @
    
        regex_patron = patron_comando
        comandos = re.findall(regex_patron, cadena)
        if len(comandos)>0: # no es un comando
            for i in comandos:
                parameter = []
                if limpiar_delimiters:
                    i = i.replace(start_delimiter,"")
                    i = i.replace(end_delimiter,"")
                parameter.append(i)
                parameter.append(identificador+i)
                lista_parameters.append(parameter)
                cadena_cambiada = re.sub(i,parameter[1],cadena_cambiada)
                # dict_params[i] = cadena_cambiada
    
    return lista_parameters, cadena_cambiada, dict_params

def clean_string(s, f, to = ""):
    """ cambia, en la cadena s
    los caracteres que aparecen en f (se entiende como una lista de caracteres)
    y lo cambia por los caracteres en to (por cada aparición de un caracter en f, lo cambia por la secuencia en to)
    """
    res = ""
    for c in f:
        res = s.replace(c, to)
        s = res

    return res

def update_list_of_params(l_params, params):
    """ coge la lista de parámetros y lo mete en la lista params, aún sin resolver """

    for k,v in l_params.items(): # estamos cambiando un parámetro?
        # l_params[i[0]] = valor # guardamos su valor correspondiente
        # limpiamos la clave de params porque contendrá { y }
        valor = v.replace("@","")
        s_key = clean_string(k, "{}")
        params[s_key] = valor # lo guardo en el diccionario global    
        
    return params
    pass

# =============================================================================
# Busqueda con profundidad en listas de diccionarios
# =============================================================================
def find_key_in_list_of_dictionaries(clave = None, list_of_dictionaries = None, current_dictionary = None):
    """
    busca recursiva de una clave en una lista de diccionarios

    Parameters
    ----------
    clave : TYPE, optional
        DESCRIPTION. The default is None.
    list_of_dictionaries : TYPE, optional
        DESCRIPTION. The default is None.
    current_dictionary : TYPE, optional
        DESCRIPTION. The default is None.

    Returns
    -------
    value : TYPE
        DESCRIPTION.
    m_current_dictionary : TYPE
        DESCRIPTION.

    """
    # si no me ha llegado una lista de diccionarios no puedo resolver nada    
    if len(list_of_dictionaries) > 0: # rastreo por los diferentes diccionarios intentando resolver la clave, pero para ello necesito que me haya llegado una lista de diccionarios...
        if not current_dictionary == None: # si no me indican por dónde comenzar, comienzo por el primer diccionario
            m_current_dictionary = current_dictionary            
        else:
            m_current_dictionary = list_of_dictionaries[0]
            
        if clave in m_current_dictionary: # está lo que busco en el diccionario?
            value = m_current_dictionary[clave]
            # if escupe:    
            #     _print(f"valor encontrado = {value}", Y)            
        else: # si no está empiezao a buscar el diccionario dónde se contiene la clave que busco
            value = None
            # primero busco si la clave está en el diccionario en que me dicen que busque
            if not clave in m_current_dictionary: # si no está, tengo que buscar en todos los diccionarios
                for ld in list_of_dictionaries:
                    
                    # 20210519 - si el valor que tengo es otro diccionario o una lista, tengo que seguir buscando...
                    if isinstance(ld, (dict, list)):
                        value, m_current_dictionary = find_key_in_list_of_dictionaries(clave, ld, current_dictionary)
                        break # ya tengo el diccionario, dejo de buscar porque esta clave puede aparecer en otro diccionario superior y siempre tiene más prioridad el primer diccionario por orden
                    else:        
                        if clave in ld: # encontrada en este diccionario!
                            value = ld[clave] # consigo el valor
                            m_current_dictionary = ld # y apunto al diccionario en el que está
                            break # ya tengo el diccionario, dejo de buscar porque esta clave puede aparecer en otro diccionario superior y siempre tiene más prioridad el primer diccionario por orden
            else: # como está, cojo su valor                    
                value = m_current_dictionary[clave]    
    
    return value, m_current_dictionary


# =============================================================================
# Convierte una estructura profunda de diccionarios en una lista simple con diccionarios clave, valor
# así se prepara para resolve_f_string que sólo busca en primer nivel en diccioniarios
# =============================================================================
def convert_deep_dictionary_in_list_of_dictionaries(d, l = []):
    
    if not isinstance(d, (list, dict, str)):
        raise Exception(f"get_key_in_dictionary: {d} no es un diccionario ni una lista ni un string")
        return None
    
    if isinstance(d, dict):
        for lk,lv in d.items():
            if isinstance(lv,list):
                # for i in range(len(lv)): # cada i es un diccionario o un string
                #     if isinstance(lv[i], dict): # sólo lo mando a resolver si es un diccionario
                #         list_of_dictionaries = convert_deep_dictionary_in_list_of_dictionaries(lv[i], list_of_dictionaries)
                #         continue
                #     if isinstance(lv[i], str):
                #         ld = {}
                #         ld[lv[i]] = lv[i] # creo un diccionario con clave y valor iguales (transformo la lista en diccionario)
                #         list_of_dictionaries.append(ld.copy())                        
                #         continue
                l = convert_deep_dictionary_in_list_of_dictionaries(lv, l)
                continue
                    
                        
            if isinstance(lv,dict): # si tengo un diccionario como valor, también busco
                l = convert_deep_dictionary_in_list_of_dictionaries(lv, l)
                continue
       
            if isinstance(lv, str): 
                ld = {}
                ld[lk] = lv
                l.append(ld.copy())
                continue

            ld = {}
            ld[lk] = lv
            l.append(ld.copy())
                
    if isinstance(d, list):
        for i in range(len(d)): # cada i es un diccionario, string o lista
            if isinstance(d[i], dict): # sólo lo mando a resolver si es un diccionario
                l = convert_deep_dictionary_in_list_of_dictionaries(d[i], l)

            if isinstance(d[i], str):
                ld = {}
                ld[d[i]] = d[i] # lo convierto en clave, valor
                l.append(ld.copy())

            if isinstance(d[i],list):
                l = convert_deep_dictionary_in_list_of_dictionaries(d[i], l)



    return l    
    
    pass



# =============================================================================
# resuelve la API_CALL que le enviemos desde los diccionarios de descripción
# siempre empieza con SCOPE, de ahí resuelve SECURITY (un scope puede obligar a una resolución de TOKEN de seguridad distinta, o un API puede estar en STAGGER, etc...)    
# =============================================================================
# @ar_decorator_functions.catch_exceptions    
# esta función existía hasta el 17/02/2022 - si todo funciona bien hay que quitarla definitivamente porque se usará la última versión que existe en rfs    
def _resolve_dynamic_fstring_(
                 
                 clave = None, # clave que quiero resolver (si la clave es un diccionario, resolverá todos sus elementos y devolverá un diccionario)
                 # context = None, # contexto de queries (contiene toda la información lógica y el acceso a access con la información ya descargada de ENV y de ENCRYPT data)
                 list_of_dictionaries = None, # orden de diccionarios que trataremos para resolver la clave, desde el primero al último (se pasa por clave y la función los resuelve)
                 current_dictionary = None, # diccionario donde quiero que empiece a buscar
                 # context_query = None, # lista de diccionarios guardada en context y que se recupera por el logical_name
                 params = {},
                 param_delimiter = "_PARAM_", # identifico, dentro de la estructura de definición de API (API_ROBOT_Providers_Microsoft, por ejemplo) que algo es un parámetro del API
                 escupe = 0, # escupo por pantalla info de rastreo?
                 force_until_all_resolved = False, # por defecto dejo que se devuelvan parámetros sin resolver (cuidado con este flag, si queda algo sin resolver contínuamente, puede dar un reventado)
                 pending_resolution = None, # si el parámetro force_until_all_resolved está a True, en la salida tiene que mirar si le sigue quedando lo mismo por resolver que cuando entró en la función, en ese caso cesa de intentar resolver
                 # first_time = True, # es la primera vez que entra en la función?
                 ):
    """ objetivo, devolver contenido de una clave, buscando en una serie de diccionarios (query contexts) incluyendo la búsqueda en el área access (diccionario global)
    y resolver fstrings dinámicas (que pueden estar escritas en el propio diccionario - mediante la función magic)
    por diseño siempre busca la clave en el primer nivel (no hace búsquedas profundas) del diccionario que recibe
    """
#    logger = lg.arguments(currentframe(), verbose = g_lg_verbose) # si lleva información sensible obligo a que no la escriba (SALVO CUANDO ESTE EN GOD_MODE)
    # params = params # contiene un diccionario en el que se meterán los parámetros que detectemos (siempre empiezan por _PARAM_{})
    
    resultado = None
    res = 0
    valor = 0
    resultado_dict = {} # por si hemos de devolver un diccionario
    
    # 20210202
    # si no es una lista de diccionarios, lo convierto en una lista
    if not isinstance(list_of_dictionaries, list):
        list_of_dictionaries = [list_of_dictionaries]
    
    # # 20210615
    # if first_time:
    #     lld = list_of_dictionaries.copy()
    #     # aplano la lista con diccionarios que me llega (pueden ser diccionarios con cierta profundidad)
    #     # así permito que esta función trabaje hasta cualquier profundidad...
    #     list_of_dictionaries = convert_deep_dictionary_in_list_of_dictionaries(lld)
    #     first_time = False
    
    # la información para obtener el token de seguridad depende del PROVIDER y del SCOPE al que queramos llamar, por eso se tiene que obtener desde esta función
    cadena_cambiada = None
    """ si no llegan diccionarios creo una lista por defecto """
    # 26/11/2020
    # if list_of_dictionaries == None: # como este es el orden de búsqueda por defecto lo meto sin preguntar a nadie, aunque se pouede modificar pasándole otro orden de búsqueda
    #     query_context = ar_api_query_context(context_name = context.logical_name)        
    #     list_of_dictionaries = []
    #     d = query_context.create_query("Security")
    #     list_of_dictionaries.append(query_context.resolve_query(d))
    #     d = query_context.create_query("Scope")
    #     list_of_dictionaries.append(query_context.resolve_query(d)) 
    # 26/11/2020    
    # if not context == None:    
    #     d = context.getAccess()
    #     if not d == None:
    #         list_of_dictionaries.append(d) # siempre acabo buscando en el diccionario global en memoria
    #     d = context.get_configuration_machine() # !!!FIX - me está llegando un dic_context... y este no tiene información del context de configuration_machine, que sólo lo tiene ar_api_context
    #     if not d == None:
    #         list_of_dictionaries.append(d) # le añado la información confidencial del cliente


    # si no me ha llegado una lista de diccionarios no puedo resolver nada    
    if len(list_of_dictionaries) > 0: # rastreo por los diferentes diccionarios intentando resolver la clave, pero para ello necesito que me haya llegado una lista de diccionarios...
        if not current_dictionary == None: # si no me indican por dónde comenzar, comienzo por el primer diccionario
            m_current_dictionary = current_dictionary            
        else:
            m_current_dictionary = list_of_dictionaries[0]

        # if escupe:    
        #     _print(f"diccionario a trabajar = {m_current_dictionary}", Y)

            
        if clave in m_current_dictionary: # está lo que busco en el diccionario?
            value = m_current_dictionary[clave]
            # if escupe:    
            #     _print(f"valor encontrado = {value}", Y)            
        else: # si no está empiezao a buscar el diccionario dónde se contiene la clave que busco
        
            found, value, m_current_dictionary = get_key_dictionary_in_dictionary(m_current_dictionary, clave)
        
            # value = None
            # # primero busco si la clave está en el diccionario en que me dicen que busque
            # if not clave in m_current_dictionary: # si no está, tengo que buscar en todos los diccionarios
            #     for ld in list_of_dictionaries:
            #         if clave in ld: # encontrada en este diccionario!
            #             value = ld[clave] # consigo el valor
            #             m_current_dictionary = ld # y apunto al diccionario en el que está
            #             break # ya tengo el diccionario, dejo de buscar porque esta clave puede aparecer en otro diccionario superior y siempre tiene más prioridad el primer diccionario por orden
            # else: # como está, cojo su valor                    
            #     value = m_current_dictionary[clave]
        
        
        if value == None:
            return None, None # no se ha encontrado nada en los diccionarios, con lo que devolvemos precisamente que no hemos encontrado nada
        
        if not isinstance(value, dict): # he encontrado un valor, ya que no es un diccionario, veo qué contiene (valor en sí mismo, lista de parámetros a resolver,...)
            
            # si hay una lista de parámetros a resolver por magic, los obtengo y dejo marcada la cadena para luego hacer una sencilla sustitución
            lista_parameters, cadena_cambiada, l_params = get_fstring_parameters(cadena = value, 
                                                                       param_delimiter = param_delimiter,          
                                                                       start_delimiter = "{", 
                                                                       end_delimiter = "}", 
                                                                       identificador = "@", 
                                                                       limpiar_delimiters = False)   
            # 13/01/2021
            if len(l_params): # si tenemos parámetros
                params = update_list_of_params(l_params, params)
            
            
            # if escupe:    
            #     _print(f"lista_parameters = {lista_parameters} cadena_cambiada = {cadena_cambiada}", B)
                
            if len(lista_parameters)==0: # no hay nada que cambiar
                resultado = cadena_cambiada
                
            for i in lista_parameters: # tengo lista de parámetros? - i contiene el parámetro a buscar y la identificación que habrá que cambiar por ese valor en la cadena original ('{scope_uri}', '@{scope_uri}')
                # i = ['{resellerid}', '@{resellerid}']
                # if escupe:
                #     _print(f"-- lista_parameters = {lista_parameters}, cadena_cambiada = {cadena_cambiada}", B+style.BRIGHT)
                key = clave             
                magic_target = i[0] # lo que tengo que buscar siempre es el primer elemento del par i (lo devuelto en la lista de parámetros)

                for dic in list_of_dictionaries: # rastreo en todos los diccionarios, la lista de parámetros
                    # dic = list_of_dictionaries[0]
                    current_dictionary = m_current_dictionary
                    target_dictionary = dic
        
                    # busco dentro del diccionario en que me encuentro
                    res, valor, l = make_magic(current_dictionary = current_dictionary, target_dictionary = target_dictionary, key = key, magic = magic_target, param_delimiter = param_delimiter)
                    # 13/01/2021
                    if len(l): # si tenemos parámetros
                        params = update_list_of_params(l, params)                    
                    
                    """ Problema, cuando resuelve una cadena, deja de procesar parámetros """
                    
                    if res == 0: # encontrado
                        cadena_cambiada = cadena_cambiada.replace(i[1], valor)
                        # 26/11/2020                        
                        if i[0] in l_params: # estamos cambiando un parámetro?
                            l_params[i[0]] = valor # guardamos su valor correspondiente
                            # limpiamos la clave de params porque contendrá { y }
                            s_key = clean_string(i[0], "{}")
                            params[s_key] = valor # lo guardo en el diccionario global
                            
                        resultado = cadena_cambiada
                        # if escupe:
                        #     _print(f"res = {res}, valor = {valor}", fg.CYAN+style.BRIGHT)

                        break
                    if res == 2: # encontrado pero tenemos un campo todavía a resolver {})
                        #cadena_cambiada = cadena_cambiada.replace(i[1], valor)
                        resultado = cadena_cambiada
                        # if escupe:
                        #     _print(f"res = {res}, valor = {valor}", fg.CYAN+style.BRIGHT)

                        magic_target = valor
                        i[0] = magic_target # puede ocurrir el caso de que el parámetro no se resuelva... (por ejemplo {params} u otros parámetros que no estén en ninguna lista de diccionarios)

            # (23/10/2020) por si nos quedan elementos que no hemos podido cambiar
            for i in lista_parameters:
                cadena_cambiada = cadena_cambiada.replace(i[1], i[0]) # quitamos el identificador y lo dehamos como {parametro} para que vuelva a intentarlo después si procede (puede que no tengamos información suficiente ahora)
                # 26/11/2020
                if i[0] in l_params: # estamos cambiando un parámetro?
                    l_params[i[0]] = i[0] # guardamos su valor correspondiente
                    s_key = clean_string(i[0], "{}")                    
                    params[s_key] = i[0] # lo guardo en el diccionario global
                    
                    
            # if escupe:
            #     _print(f"-- res {res}, valor = {valor}", B+style.BRIGHT)    
            #     _print(f"-- lista_parameters = {lista_parameters}, cadena_cambiada = {cadena_cambiada}", B+style.BRIGHT)                    
            # if escupe:    
            #     _print(f"cadena_cambiada = {cadena_cambiada}", Y+style.BRIGHT)
            # cuando ha terminado hago un último cambio por si han quedado parámetros sin cambiar
            for i in lista_parameters:
                cadena_cambiada = cadena_cambiada.replace(i[1], i[0]) 
                # 26/11/2020 antes de devolver el control, quitamos todos los identificadores de params                              
                cadena_cambiada = cadena_cambiada.replace(param_delimiter, "")                
                resultado = cadena_cambiada
                # 13/01/2021
                if i[0] in l_params: # estamos cambiando un parámetro?
                    l_params[i[0]] = i[0] # guardamos su valor correspondiente
                    s_key = clean_string(i[0], "{}")                    
                    params[s_key] = i[0] # lo guardo en el diccionario global                  
                
                
                
        if isinstance(value, dict): # si es un diccionario, itero y resuelvo todas sus claves
            for k,v in value.items():
                clave = k
                # _print(f"Clave a trabajar - [{clave}]", fg.RED+style.BRIGHT)

                # ojo, no permitimos diccionarios de más nivel en esta versión (eixigiría una función recursiva) --- hay que cambiarlo
                if not isinstance(v,dict):                    
                # si hay una lista de parámetros a resolver por magic, los obtengo y dejo marcada la cadena para luego hacer una sencilla sustitución
                    lista_parameters, cadena_cambiada, l_params = get_fstring_parameters(cadena = v, 
                                                                               param_delimiter = param_delimiter,          
                                                                               start_delimiter = "{", 
                                                                               end_delimiter = "}", 
                                                                               identificador = "@", 
                                                                             limpiar_delimiters = False)   
                    # 13/01/2021
                    if len(l_params): # si tenemos parámetros
                        params = update_list_of_params(l_params, params)                    

                    if len(lista_parameters)>0:
                        for i in lista_parameters: 
                            # if escupe:
                            #     _print(f"lista_parameters = {lista_parameters}, cadena_cambiada = {cadena_cambiada}", Y)
                            magic_target = i[0]                                 
                            key = clave   
                            current_dictionary = m_current_dictionary                                
                            for dic in list_of_dictionaries:  

                                # if escupe:
                                #     if LOGICAL_NAME in dic:
                                #         _print(f"del dicionario [{dic[LOGICAL_NAME]}]", R)
                                #     _print(dic, fg.RED+style.BRIGHT)                                    
                                target_dictionary = dic
       
                                # busco en el diccionario en que me encuentro
                                res, valor, l = make_magic(current_dictionary = current_dictionary, target_dictionary = target_dictionary, key = key, magic = magic_target, param_delimiter = param_delimiter)
                                if len(l_params): # si tenemos parámetros
                                    params = update_list_of_params(l_params, params)                                
                                # si me devuelve algo con los identificadores quiere decir que no lo ha podido resolver en el diccionario actual (está apuntando a una clave que vuelve a devolver algo que está en otro diccionario superior)
                                # ahora cambio en la cadena_cambiada los valores obtenidos
                                # if escupe:
                                #     _print(f"key = {key}, i[0] = {i[0]}, res = {res}, valor = {valor}", Y)

                                if res == 0: # encontrado
                                    cadena_cambiada = cadena_cambiada.replace(i[1], valor)
                                    resultado_dict[k] = cadena_cambiada
                                    # 13/01/2021
                                    if i[0] in l_params: # estamos cambiando un parámetro?
                                        l_params[i[0]] = valor # guardamos su valor correspondiente
                                        # limpiamos la clave de params porque contendrá { y }
                                        s_key = clean_string(i[0], "{}")
                                        params[s_key] = valor # lo guardo en el diccionario global
                                    
                                    # if escupe:
                                    #     _print(f"res = {res}, valor = {valor}", C)
                                    break
                                if res == 2: # encontrado pero tenemos un campo todavía a resolver {})
                                    #cadena_cambiada = cadena_cambiada.replace(i[1], valor)
                                    resultado_dict[k] = cadena_cambiada
                                    # if escupe:
                                    #     _print(f"res = {res}, valor = {valor}", C)
                                    magic_target = valor
                                    i[0] = magic_target # añadido 23/10/2020 - por si hay parámetros que no se resuelven en la lista de diccionarios (el caso de {params}) - devolvemos lo que haya podido resolver
                        # por si nos quedan elementos que no hemos podido cambiar
                        for i in lista_parameters:
                            cadena_cambiada = cadena_cambiada.replace(i[1], i[0]) 
                            # 26/11/2020 antes de devolver el control, quitamos todos los identificadores de params                              
                            cadena_cambiada = cadena_cambiada.replace(param_delimiter, "")
                            resultado_dict[k] = cadena_cambiada
                            # 13/01/2021
                            if i[0] in l_params: # estamos cambiando un parámetro?
                                l_params[i[0]] = i[0] # guardamos su valor correspondiente
                                s_key = clean_string(i[0], "{}")                    
                                params[s_key] = i[0] # lo guardo en el diccionario global                            

                    else: # no hay nada que cambiar porque no tenemos lista_parameters
                        resultado_dict[k] = cadena_cambiada                             
            resultado = resultado_dict
        
    """ 
    29012021
    nuevo para intentar resolver a mayores profundidades, cuando llega aquí puede que se haya encontrado con un parámetro
    que se resolvería en otra pasada 
    """    
    if force_until_all_resolved: # si me piden que resuelva todo, es porque aún quedan parámetros por resolver
        if resultado == pending_resolution: 
            # esto quiere decir que no ha podido resolver nada más (puede que haya información que se resolverá después, cuando se carguen parámetros adicionales)
            # así que nos vamos
            return resultado, params
        else:
            # intentamos resolver nuevamente con una profundidad adicional y desde el principio
            # creo una clave temporal
            clave = str(IPL.timestamp())
            # copio lo que no hemos podido resolver
            valor = resultado
            # lo meto en la lista de diccionario para que intente resolver la clave con la información que ya tiene
            list_of_dictionaries.insert(0,{clave : valor})
            # y veo si se resuelve
            resultado, params = _resolve_dynamic_fstring(
                     clave = clave, # clave que quiero resolver (si la clave es un diccionario, resolverá todos sus elementos y devolverá un diccionario)
                     # context = None, # contexto de queries (contiene toda la información lógica y el acceso a access con la información ya descargada de ENV y de ENCRYPT data)
                     params = params, # arrastro los parámetros para no perderlos
                     list_of_dictionaries = list_of_dictionaries, # orden de diccionarios que trataremos para resolver la clave, desde el primero al último (se pasa por clave y la función los resuelve)
                     current_dictionary = None, # diccionario donde quiero que empiece a buscar
                     # context_query = None, # lista de diccionarios guardada en context y que se recupera por el logical_name
                     param_delimiter = param_delimiter, # identifico, dentro de la estructura de definición de API (API_ROBOT_Providers_Microsoft, por ejemplo) que algo es un parámetro del API
                     escupe = tracing, # escupo por pantalla info de rastreo?
                     force_until_all_resolved = True, # por defecto dejo que se devuelvan parámetros sin resolver (cuidado con este flag, si queda algo sin resolver contínuamente, puede dar un reventado)
                     pending_resolution = resultado, # si el parámetro force_until_all_resolved está a True, en la salida tiene que mirar si le sigue quedando lo mismo por resolver que cuando entró en la función, en ese caso cesa de intentar resolver
                     # first_time = False, # no es la primera vez que entra en la función, ...
                     )  
        
        
    return resultado, params

def fix_dictionary_as_param(dictionary, swap = ("'",'"') ):
    """
    convierte un dictionary que está dentro de un string, pero esperando que las claves
    aparezcan entre dobles comillas (que es lo que espera json)

    Parameters
    ----------
    dictionary : TYPE
        DESCRIPTION.

    swap : TYPE tuple
        DESCRIPTION. will change first to second

    Returns
    -------
    None.

    """
    import re
    p = re.compile('(?<!\\\\)\'')
    d = p.sub('\"', dictionary)
    import json
    dic = json.loads(d)
    return dic
    
    
def desenrosca_elemento(p, p_flatten = []):
    """
    Description
    -----------

    Esta función es responsable de "desenroscar" elementos lista o diccionario
    devolviendo un flatten dictionary
    Si la lista contiene diccionarios los analiza y busca si cada clave, recursivamente, es una lista o un diccionario
    Si es una lista mira sus diccionarios internos
    Si es un diccionario va creando pares k,v hasta que un v es una lista o un diccionario, dónde vuelve a ejecutarse recursivamente



    """

    if isinstance(p, list): # desenrosco diccionarios en una lista
        for elemento in p:
            p_flatten = desenrosca_elemento(elemento, p_flatten)

    if isinstance(p, dict): # desenrosco diccionarios
        elemento = p
        for k,v in elemento.items():
            p_flatten.append({k : v})
            p_flatten = desenrosca_elemento(v, p_flatten)

    return p_flatten    

# =============================================================================
# Obtiene la información de una clave en una lista de diccionarios, resolviendo como fstring
# devuelve el resultado de la búsqueda, y todo lo que sean parámetros, a medida que ha ido buscando
# =============================================================================
def getFromDictionaries(
                 clave = None, # clave que quiero resolver (si la clave es un diccionario, resolverá todos sus elementos y devolverá un diccionario)
                 list_of_dictionaries = None, # orden de diccionarios que trataremos para resolver la clave, desde el primero al último (se pasa por clave y la función los resuelve)
                 current_dictionary = None, # diccionario donde quiero que empiece a buscar
                 params = {},
                 param_delimiter = "_PARAM_", # identifico, dentro de la estructura de definición de API (API_ROBOT_Providers_Microsoft, por ejemplo) que algo es un parámetro del API
                 escupe = 0, # escupo por pantalla info de rastreo?
                 force_until_all_resolved = True, # por defecto dejo que se devuelvan parámetros sin resolver (cuidado con este flag, si queda algo sin resolver contínuamente, puede dar un reventado)
                 pending_resolution = None, # si el parámetro force_until_all_resolved está a True, en la salida tiene que mirar si le sigue quedando lo mismo por resolver que cuando entró en la función, en ese caso cesa de intentar resolver
                 # first_time = True,
        
        ):
    
    """
    objetivo, devolver contenido de una clave, buscando en una serie de diccionarios (query contexts) incluyendo la búsqueda en el área access (diccionario global)
    y resolver fstrings dinámicas (que pueden estar escritas en el propio diccionario - mediante la función magic)

    Parameters
    ----------
    clave : TYPE, optional
        DESCRIPTION. The default is None. clave que quiero resolver (si la clave es un diccionario, resolverá todos sus elementos y devolverá un diccionario)                
    list_of_dictionaries : TYPE, optional
        DESCRIPTION. The default is None. orden de diccionarios que trataremos para resolver la clave
    current_dictionary : TYPE, optional
        DESCRIPTION. The default is None. se usa para controlar qué diccionario se está tratando cuando se hace recursividad en la función, diccionario donde quiero que empiece a buscar
     params : TYPE, optional
        DESCRIPTION. The default is {}. construye la lista de parámetros que va encontrando (todo lo que no comience por "AR_")
    param_delimiter : TYPE, optional
        DESCRIPTION. The default is "_PARAM_". delimitador usado para parámetros, identifico dentro de la estructura de definición de API (API_ROBOT_Providers_Microsoft, por ejemplo) que algo es un parámetro del API                 escupe : TYPE, optional
    escupe : TYPE, optional 
        DESCRIPTION. The default is 0. escupo por pantalla info de rastreo?
    force_until_all_resolved : TYPE, optional
        DESCRIPTION. The default is False. por defecto dejo que se devuelvan parámetros sin resolver (cuidado con este flag, si queda algo sin resolver contínuamente, puede dar un reventado)
    pending_resolution : TYPE, optional
        DESCRIPTION. The default is None. si el parámetro force_until_all_resolved está a True : TYPE

    Returns
    -------
    TYPE
        DESCRIPTION.

    """    
    
    # if not isinstance(list_of_dictionaries, list):
    #     list_of_dictionaries = [list_of_dictionaries] # siempre se envía una lista
    
    # # hacemos un wrapper porque realmente quiero que funcione con esta signature de función
    # # _resolve_dynamic_fstring se queda como signature interna
    # return  _resolve_dynamic_fstring(
    #              clave = clave, # clave que quiero resolver (si la clave es un diccionario, resolverá todos sus elementos y devolverá un diccionario)
    #              list_of_dictionaries = list_of_dictionaries, # orden de diccionarios que trataremos para resolver la clave, desde el primero al último (se pasa por clave y la función los resuelve)
    #              current_dictionary = current_dictionary, # diccionario donde quiero que empiece a buscar
    #              params = params,
    #              param_delimiter = param_delimiter, # identifico, dentro de la estructura de definición de API (API_ROBOT_Providers_Microsoft, por ejemplo) que algo es un parámetro del API
    #              escupe = escupe, # escupo por pantalla info de rastreo?
    #              force_until_all_resolved = force_until_all_resolved, # por defecto dejo que se devuelvan parámetros sin resolver (cuidado con este flag, si queda algo sin resolver contínuamente, puede dar un reventado)
    #              pending_resolution = pending_resolution, # si el parámetro force_until_all_resolved está a True, en la salida tiene que mirar si le sigue quedando lo mismo por resolver que cuando entró en la función, en ese caso cesa de intentar resolver
    #              # first_time = first_time,
    #              )


    """ creo una lista de diccionarios
    contendrá solamente diccionarios
    si uno de los elementos es un diccionario que, a su vez, contiene un diccionario se mete el diccionario original 
    y luego, como diccionarios, todos sus elementos

    """    

    """ CAMBIO 17/03/2022 para permitir busquedas en cualquier profundidad """
    list_to_look_for = []
    # if not isinstance(list_of_dictionaries, list):
    #     if isinstance(list_of_dictionaries, dict):
    #         # añado todas sus claves como diccionarios
    #         for k,v in list_of_dictionaries.items():
    #             if not isinstance(v, dict):
    #                 list_to_look_for.append({k:v})
    #             if isinstance(v, dict):
    #                 list_to_look_for.append({k:v})
    #                 list_to_look_for.append(v)
    # else:
    #     list_to_look_for = list_of_dictionaries                    
    
    list_to_look_for = desenrosca_elemento(list_of_dictionaries, p_flatten = [])
    """ FIN CAMBIO 17/03/2022 para permitir busquedas en cualquier profundidad """


    # hacemos un wrapper porque realmente quiero que funcione con esta signature de función
    # _resolve_dynamic_fstring se queda como signature interna
    return  _resolve_dynamic_fstring(
                 clave = clave, # clave que quiero resolver (si la clave es un diccionario, resolverá todos sus elementos y devolverá un diccionario)
                 list_of_dictionaries = list_to_look_for, # orden de diccionarios que trataremos para resolver la clave, desde el primero al último (se pasa por clave y la función los resuelve)
                 current_dictionary = current_dictionary, # diccionario donde quiero que empiece a buscar
                 params = params,
                 param_delimiter = param_delimiter, # identifico, dentro de la estructura de definición de API (API_ROBOT_Providers_Microsoft, por ejemplo) que algo es un parámetro del API
                 escupe = escupe, # escupo por pantalla info de rastreo?
                 force_until_all_resolved = force_until_all_resolved, # por defecto dejo que se devuelvan parámetros sin resolver (cuidado con este flag, si queda algo sin resolver contínuamente, puede dar un reventado)
                 pending_resolution = pending_resolution, # si el parámetro force_until_all_resolved está a True, en la salida tiene que mirar si le sigue quedando lo mismo por resolver que cuando entró en la función, en ese caso cesa de intentar resolver
                 # first_time = first_time,
                 )


                
    """ esto no me vale porque no devolvía PARAMS y es necesario """
    # clave_a_buscar = "{" + clave + "}"
    # resultado = rfs.compute_fstring(fstring = clave_a_buscar,
    #                      list_of_dictionaries = list_of_dictionaries)

    # if resultado[1] == []:
    #     # como no ha encontrado nada devuelvo None
    #     return None

    # return resultado[0] # contiene la solución 




# ---- .
# ---- * STRINGS --------------------------------------------------------------

def  get_elements_in_string(patron = "bd59a770-df9d-4947-a787-b3be57b4b844/{subscriptionId}_Office365_ScoreProfiles", # cadena a buscar
                                       start_delimiter = "{",
                                       end_delimiter = "}",
                                       ): 
    """
    recupera todos los elementos en un string que estén entre los delimitadares

    Parameters
    ----------
    patron : TYPE, optional
        DESCRIPTION. The default is "bd59a770-df9d-4947-a787-b3be57b4b844/{subscriptionId}_Office365_ScoreProfiles".
    # cadena a buscar                                       start_delimiter : TYPE, optional
        DESCRIPTION. The default is "{".
    end_delimiter : TYPE, optional
        DESCRIPTION. The default is "}".
     : TYPE
        DESCRIPTION.

    Returns
    -------
    None.

    """            

    import re

    patron_comando = r"{start_delimiter}[\w\d+-:]+{end_delimiter}".format(
        start_delimiter = start_delimiter,
        end_delimiter = end_delimiter) # buscamos secuencias de palabras dígitos, signos + y menos y : entre @
    
    regex_patron = patron_comando
    comandos = re.findall(regex_patron, patron)
    return comandos

def sustitute_elements_in_string(patron = "bd59a770-df9d-4947-a787-b3be57b4b844/{subscriptionId}_Office365_ScoreProfiles", # cadena a buscar
                                       start_delimiter = "{",
                                       end_delimiter = "}",
                                       prefix = "", # a añadir antes a lo no encontrado
                                       substitute_by = None, # sustituimos por una cadena concreta? o por lo que se encuentra entre los delimitadores
                                       suffix = "", # añadimos al final a lo no encontrado
                                       ):
    """
    sustituye, en lo que recibe en patron, todo lo que esté entre los delimitadores
    lo cambia por lo que está entre los delimitadores, decorato con prefix al principio y suffix al final
    salvo que la cadena substitute_by venga con al diferente a None, en ese caso lo sustituye por lo que pongamos en la cadena

    Parameters
    ----------
    patron : TYPE, optional
        DESCRIPTION. The default is "bd59a770-df9d-4947-a787-b3be57b4b844/{subscriptionId}_Office365_ScoreProfiles".
    # cadena a buscar                                       start_delimiter : TYPE, optional
        DESCRIPTION. The default is "{".
    end_delimiter : TYPE, optional
        DESCRIPTION. The default is "}".
    prefix : TYPE, optional
        DESCRIPTION. The default is "".
    # a añadir antes a lo no encontrado                                       substitute_by : TYPE, optional
        DESCRIPTION. The default is None.
    # sustituimos por una cadena concreta? o por lo que se encuentra entre los delimitadores                                       suffix : TYPE, optional
        DESCRIPTION. The default is "".
    # añadimos al final a lo no encontrado : TYPE
        DESCRIPTION.

    Returns
    -------
    patron : TYPE str
        DESCRIPTION. cadena con las sustituciones, o cadena original si no ha habido sustituciones

    """

    import re

    patron_comando = r"{start_delimiter}[\w\d+-:]+{end_delimiter}".format(
        start_delimiter = start_delimiter,
        end_delimiter = end_delimiter) # buscamos secuencias de palabras dígitos, signos + y menos y : entre @
    
    regex_patron = patron_comando
    comandos = re.findall(regex_patron, patron)
    if len(comandos)==0: # no es un comando
        _print("no encontrado")
    else:
        lista_cambiar = []
        for el in comandos:
            resultado = el.replace(start_delimiter, "")
            resultado = resultado.replace(end_delimiter, "")  
            if substitute_by == None:
                lista_cambiar.append(resultado)
            else:
                lista_cambiar.append(substitute_by)
                
        # ahora cambio en patron todo lo que necesito
    for i in range(0, len(comandos)):
        patron = patron.replace(comandos[i], f"{prefix}{lista_cambiar[i]}{suffix}" )
    
    return patron
        


def test_sustitute_elements_in_string():

    patron = "bd59a770-df9d-4947-a787-b3be57b4b844/{subscriptionId}_Office365_ScoreProfiles"
    start_delimiter = "{"
    end_delimiter = "}"
    prefix = ""
    substitute_by = None
    suffix = ""    


    _print(sustitute_elements_in_string(patron = patron, # cadena a buscar
                                           start_delimiter = start_delimiter,
                                           end_delimiter = end_delimiter,
                                           prefix = prefix, # a añadir antes a lo no encontrado
                                           substitute_by = substitute_by, # sustituimos por una cadena concreta? o por lo que se encuentra entre los delimitadores
                                           suffix = suffix, # añadimos al final a lo no encontrado
                                           ))

    patron = "bd59a770-df9d-4947-a787-b3be57b4b844/{subscriptionId}_{kk}_Office365_ScoreProfiles"
    start_delimiter = "{"
    end_delimiter = "}"
    prefix = "[NOT_"
    substitute_by = None
    suffix = "_FOUND] "    


    _print(sustitute_elements_in_string(patron = patron, # cadena a buscar
                                           start_delimiter = start_delimiter,
                                           end_delimiter = end_delimiter,
                                           prefix = prefix, # a añadir antes a lo no encontrado
                                           substitute_by = substitute_by, # sustituimos por una cadena concreta? o por lo que se encuentra entre los delimitadores
                                           suffix = suffix, # añadimos al final a lo no encontrado
                                           ))



# =============================================================================
# Hace un parsing dependiendo de lo que devuelva la llamada al API
# de momento el parsing es JSON o CSV    
# =============================================================================
def parseRes( res, parser ):
    return

def replace_list_of_values(cadena, lista):
    """
    reemplaza, en una cadena, todo lo que aparezca entre {} por el contenido 
    que viene en un list de diccionarios
    los diccionarios vendrán en la forma {"name":name, "value":value}
    todo lo que aparezca como name, se cambiará por value
    y se devolverá la cadena cambiada

    Parameters
    ----------
    cadena : TYPE
        DESCRIPTION.
    lista : TYPE
        DESCRIPTION.

    Returns
    -------
    None.

    """
    cadena_a_cambiar = cadena
    for d in lista:
        lfrom = d["name"]
        lto = d["value"]
        cadena_a_cambiar = cadena_a_cambiar.replace("{"+lfrom+"}", lto)
        
    return cadena_a_cambiar

# ---- .
# ---- CARGA LIBRERIAS



import importlib
# @ar_decorator_functions.catch_exceptions
def custom_import_library(namelibrary = "IPL", library_name = "c:/PATH/library.py"):
    """[summary]

    Args:
        namelibrary (str, optional): [nombre de librería a cargar]. Defaults to "IPL".
        library_name (str, optional): [path y filename completo de la librería a cargar, conteniendo .py]. Defaults to "c:/PATH/library.py".

    Raises:
        Exception: [description]

    Returns:
        [module]: [modulo cargado]
    """
    spec = importlib.util.spec_from_file_location(namelibrary, library_name)
    l = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(l)
    # si todo va bien devolvemos el módulo cargado para ser ejecutado, si da un error, lo captura catch_exceptions
    return l


# ---- .
# ---- * PERSISTENCIA DE OBJETOS  -------------------------------------------------------------------------------------------




# source: https://pymotw.com/2/json/index.html#module-json
    
import json

# =============================================================================
# Devuelve la información para que un objeto persista (dormir)
# por defecto lo devuelve como un JSON, si se quiere string habrá de ponerse to_dict a False
# =============================================================================
def sleep(obj, to_dict = True):
    """
    crea una imagen, en memoria, del objeto que recibe, 
    esta imagen puede posteriormente guardarse dónde se le diga

    Parameters
    ----------
    obj : TYPE
        DESCRIPTION.
    to_dict : TYPE, optional
        DESCRIPTION. The default is True.

    Returns
    -------
    generated_dict : TYPE
        DESCRIPTION.

    """

    generated_dict = json.dumps(obj, default=convert_to_builtin_type)
    if to_dict:
        generated_dict = json.loads(generated_dict)
    return generated_dict

def awake(encoded_object):
    """
    wrapper de resurrect

    Parameters
    ----------
    encoded_object : TYPE
        DESCRIPTION.

    Returns
    -------
    None.

    """
    return resurrect(encoded_object)


# =============================================================================
# Con la información de un objeto se le permite recrearlo (resurrecitar)
# =============================================================================
def resurrect(encoded_object):
    """ desde memoria es capaz de reiniciar el objeto al que se le hizo sleep """
    encoded_object = json.dumps(encoded_object)
    myobj_instance = json.loads(encoded_object, object_hook=dict_to_object)
    return myobj_instance
   
import types

def test_sleep_and_resurrect():
    """ testeo la capacidad de volcar y recuperar un objeto de memoria a disco """
    
    """ prueba content """
    import ar_tasks

    original_dict = [{'id': '/subscriptions/3be226db-7330-491e-8bc6-eca10000d669', 'authorizationSource': 'RoleBased', 'managedByTenants': [], 'subscriptionId': '3be226db-7330-491e-8bc6-eca10000d669', 'tenantId': 'bd59a770-df9d-4947-a787-b3be57b4b844', 'displayName': 'Microsoft Azure', 'state': 'Enabled', 'subscriptionPolicies': {'locationPlacementId': 'Public_2014-09-01', 'quotaId': 'CSP_2015-05-01', 'spendingLimit': 'Off'}}]

    # creo un content
    c = ar_tasks._content(
                          _from = "_prueba",
                          _name_content = "t.name", 
                          _task_produces = "_volcado",
                          _type = ar_tasks.C_FINAL,
                          _t = "t.get_logical_name()", # le mando quién ha construido el content (le mando el nombre, no la referencia)
                          # _dict = {"decorated_name":decorated_name},
                          _dict = {"original_dict":original_dict},                          
                          # _dict = original_dict,
                          _manager = "nombre de la instancia del manager, no la referencia", # puedo eliminar información de memoria si un contet sabe persistir...
                          )       
    a = sleep(c)


# =============================================================================
# # construyo un hook para los tipos que no se pueden serializar
# =============================================================================
def convert_to_builtin_type(obj):
    import json
    import datetime

    # si da problemas, activar para saber por dónde va y resolver en qué clave está dando problemas    
    _print( f'default({repr(obj)} )')
    
    # Convert objects to a dictionary of their representation
    # cuidado que datetime es un objeto escrito en C y no contiene la info de Python
    try:
        d = {}
        # trabajamos las fechas para evitar que sea un objeto en C
        if isinstance(obj, datetime.date):
            string = obj.strftime(str_timestamp)
            d = dict(year=obj.year, month=obj.month, day=obj.day)
            return d
        # si es una función no lo devuelvo, habría que volver a cargarlo
        if isinstance(obj, types.FunctionType):
            return None
        
        d = { '__class__':obj.__class__.__name__, 
            '__module__':obj.__class__.__module__,
            }
            
        d.update(obj.__dict__)
    except Exception as e:
        _print(f'convert_to_builtin_type: Exception reason {e}')
        return None
    return d
        
        
# =============================================================================
# # creo cómo instanciar dinámicamente un objeto en base a su info serializada
# para poder hacer que un objeto resurrezca necesita recibir en el __init__ todos los parámetros que tenga en self    
# =============================================================================
def dict_to_object(d):
    import json    
    import pprint
    if '__class__' in d:
        class_name = d.pop('__class__')
        module_name = d.pop('__module__')
        module = __import__(module_name)
        # _print( f"MODULE: {module}")
        class_ = getattr(module, class_name)
        # _print( f"CLASS: {class_}")
        
        args = {}
        for key, value in d.items():
            args[key] = value
        # args = dict( (key.encode('ascii'), value) for key, value in d.items())
        # args = dict( (key.encode("utf8"), value) for key, value in d.items())
        # _print( f"INSTANCE class_: {class_}")
        # _print( f"INSTANCE ARGS: {args}")
        inst = class_(**args)
    else:
        inst = d
    return inst

# ---- .
# ---- * EMPAQUETAR Y DESEMPAQUETAR BINARIO PARA ENVIAR COMO LINEA DE COMANDOS ---------------------------------------------


# =============================================================================
# Conversión de CSV a JSON. La primera línea del CSV se entiende que tiene los headers
# =============================================================================
def CSVtoJSON(csvdata):
    
    import csv, json
    
    # 1) guardamos el CSV como un fichero temporal


    import csv
    
    rawdata = csvdata
    myreader = csv.reader(rawdata.splitlines())
    # for row in myreader:
    #     _print(row)
    
    JSON = []
    # lineas = csv.split("\n")
    
    # # quitamos los retornos de carro
    # for n in range(n, len(lineas)):
    #     lineas[n] = lineas[n].replace('\r', '')
    # # quitamos las comillas
    # for n in range(n, len(lineas)):
    #     lineas[n] = lineas[n].replace('"', "'")
        
    count = 0
    headers = []
    for row in myreader:
        fields = []
        d = {} # el JSON sólo tendrá una lista de diccionarios, con cada KEY mapeada con su correspondiente valor
        if count == 0:
            for field in row:
                headers.append(field)
            # for x in range(0, len(headers)): # limpio el caracter \r que me trae posiblemente cada línea
            #     headers[x] = headers[x].replace('\r', '')
            count = 1
        else:
            # creamos una lista de campos con los contenidos
            for field in row:
                if field == "False":
                    field = False
                if field == "True":
                    field = True
                fields.append(field)
            # metemos la lista de campos en el diccionario    
            for n in range(0,len(fields)):
                d[headers[n]] = fields[n]
            # añadimos el diccionario al JSON        
            JSON.append(d)
            # limpiamos la lista de campos
            fields.clear()
   
        
    return JSON    
    
# =============================================================================
# Empaqueta un string en una cadena binaria
# =============================================================================
def binaryPack(cad):
    import base64
    import json
    json_to_pass_B = str.encode(cad)
    json_to_pass_B = base64.b64encode(json_to_pass_B)    
    return json_to_pass_B
# =============================================================================
# Desempaqueta una cadena binaria y devuelve un string    
# =============================================================================
def binaryUnpack(cad):    
    import base64
    # desempaquetado binario de la cadena enviada
    json_passed = cad[1:]
    json_passed = base64.b64decode(json_passed)
    return json_passed

# =============================================================================
# Función sencillita que escribe un string en un fichero
# =============================================================================
def escribeDatos(cadena, filename):
    # cadena = get_str_timestamp()
    
    writeRawFile(cadena, filename, "w")


# =============================================================================
# Empaqueta un JSON y lo envía al script que se le pide
# Devuelve el retorno del script
# se puede ejecutar esperando a la finalización del comando (sync = True, por defecto)
# o en modo asíncrono (no espera la ejecución del comando sync = False) esto último no está probado    
# =============================================================================

import subprocess
def JsonToProgram(script, json_param, sync = True, 
                  preparam = "", # si hay algún parámetro a pasar al programa que recibe los parámetros ("--bjson=") por ejemplo
                  cmd_param = "/C", # con /C la consola se cierra automáticamente, con /K el input se queda en la consola que se puede haber abierto
                  title_of_window = "", # título que se quiere pasar a la ventana
                  ):
    """
    Empaqueta un JSON y lo envía al script que se le pide
    Devuelve el retorno del script
    se puede ejecutar esperando a la finalización del comando (sync = True, por defecto)
    o en modo asíncrono (no espera la ejecución del comando sync = False) esto último no está probado    

    Parameters
    ----------
    script : TYPE
        DESCRIPTION. Nombre del script Python al que queremos llamar.
    json_param : TYPE dict - Toda la información que se quiere enviar al script
        DESCRIPTION. 
    sync : TYPE, optional
        DESCRIPTION. The default is True. Ejecución síncrona (True) o asíncrona (False)
    preparam : TYPE, optional    
        DESCRIPTION. si el programa destino hace parsing de parámetros, se puede incluir (por ejemplo, --bjson= [json en binario])
    cmd_param : TYPE, optional
        DESCRIPTION. The default is "/C". Se mantienen las ventanas abiertas (/K) o se cierran tras la ejecución (/C)?
    title_of_window : TYPE, optional
        DESCRIPTION. The default is "". Lo que aparecerá en el título de la ventana
    Returns
    -------
    resultado : TYPE
        DESCRIPTION. Proceso que se ha ejecutado

    """
    import json
    resultado = False
    
    # ? lo pasamos a cadena
    json_to_pass = json.dumps(json_param)
    # ? lo empaquetamos en binario
    json_to_pass_B = binaryPack(json_to_pass)
    # ? ahora lo enviamos al programa que nos piden y retorno el resulto (0 es que todo ha ido bien)
    cadena_de_ejecucion = f"start \"{title_of_window}\" cmd {cmd_param} python \"{script}\" {preparam}\"{json_to_pass_B}\""
    
    if sync == True:    
        import os
        resultado = os.system(cadena_de_ejecucion)    
    if sync == False:
        process = subprocess.Popen(cadena_de_ejecucion,
                            creationflags=subprocess.CREATE_NEW_CONSOLE,
                            # start_new_session=True,
                            shell=True)
        resultado = process    
    return resultado 

# =============================================================================
# Ejecuta un comando Windows síncrono o asíncrono
# =============================================================================
def execute_Windows_command(comando, sync = True):
    cadena_de_ejecucion = comando
    
    resultado = True
    
    if sync == True:    
        import os
        resultado = os.system(cadena_de_ejecucion)   
        
    """ ejecutamos un proceso en desatendido """    
    if sync == False:
        
        subprocess.Popen(cadena_de_ejecucion,
                            creationflags=subprocess.CREATE_NEW_CONSOLE,
                            shell=True)    
    return resultado
    


# =============================================================================
# Recibe una cadena binaria y devuelve un JSON
# =============================================================================
def BinaryToJson(json_passed_B):
    """
    convierte un json en binario al json original. Sirve para pasar JSON entre scripts e iniciar procesos desatendidos.

    Parameters
    ----------
    json_passed_B : TYPE
        DESCRIPTION.

    Returns
    -------
    json_passed : TYPE
        DESCRIPTION.

    """
    import json
    # desempaqueta la cadena
   
    json_passed = binaryUnpack(json_passed_B)    
    json_passed = json.loads(json_passed)    

    return json_passed
# ---- .
# ---- * ENCRIPTACION -----------------------------------------------------------------------------------------------

def encryption_get_method_to_encrypt():
    """
    (c) Seachad - FGV
    
    Description:
    ---------------------- 
        encryption_get_method_to_encrypt
        Obtiene el método de encriptación y desencriptación
    
    Extended Description:
    ---------------------
        _extended_summary_
    
    Returns:
    --------
            _type_: _description_
    """
    # from common import ENCRYPTION_METHOD_FERNET, ENCRYPTION_METHOD_AES, ENCRYPTION_METHOD_CHACHA20, ENCRYPTION_METHOD_OBFUSCATE_BASE64    
    encryption_method = os.getenv("CM_ENCRYPTION_METHOD_TO_ENCRYPT", "FERNET")    
    METHOD_TO_ENCRYPT = get_encryption_method_by_name(encryption_method)
    # if METHOD_TO_ENCRYPT.lower() == "fernet":
    #     METHOD_TO_ENCRYPT = ENCRYPTION_METHOD_FERNET
    # elif METHOD_TO_ENCRYPT.lower() == "aes":
    #     METHOD_TO_ENCRYPT = ENCRYPTION_METHOD_AES
    # elif METHOD_TO_ENCRYPT.lower() == "chacha20":
    #     METHOD_TO_ENCRYPT = ENCRYPTION_METHOD_CHACHA20
    # elif METHOD_TO_ENCRYPT.lower() == "obfuscate_base64":
    #     METHOD_TO_ENCRYPT = ENCRYPTION_METHOD_OBFUSCATE_BASE64    
        
    # manda el encryption temporal
    temporal_method = _CACHE_get("TMP_encryption_method", singleton = True)
    if not temporal_method == None:
        if isinstance(temporal_method, str):
            temporal_method = get_encryption_method_by_name(temporal_method)    
        METHOD_TO_ENCRYPT = temporal_method
        
    return METHOD_TO_ENCRYPT

def encryption_clean_temporal_information(encryption_temporal_keys_prefix = "TMP_encryption_"):
    """
    (c) Seachad - FGV
    
    Description:
    ---------------------- 
        encryption_clean_temporal_information
        borramos todas las claves temporales de encryption
    
    Extended Description:
    ---------------------
        _extended_summary_
    """
    _CACHE_delete(encryption_temporal_keys_prefix, singleton = True)
    
    
    pass

def encryption_set_temporal_support(tablename = "", status = False):
    """
    (c) Seachad - FGV
    
    Description:
    ---------------------- 
        set_temporal_encryption_support
        podemos activar o deactivar temporalmente la encryptación para una tabla o para todo el mundo
    
    Extended Description:
    ---------------------
        _extended_summary_
    """

    _CACHE_set(f"TMP_encryption_{tablename}", status, singleton = True)
    
    
def encryption_get_temporal_support(tablename = ""):
    """
    (c) Seachad - FGV
    
    Description:
    ---------------------- 
        get_temporal_encryption_support
        Devuelve si encryption está habilitado para una tabla o para todo el mundo
    
    Extended Description:
    ---------------------
        _extended_summary_
    
    Args:
    -----
            - tablename (_type_, optional): _description_. Defaults to None.
    """
    status = None
    status_tabla = False
    status_global = False
    status_tabla = _CACHE_get(f"TMP_encryption_{tablename}", singleton = True) # con esto tenemos el de tabla
    status_global = _CACHE_get(f"TMP_encryption_", singleton = True)
    if status_tabla == None: # no hay estado para tabla, tiraremos del global
        status = status_global
    else:
        status = status_tabla    
    
    if status == None:
        status = True # devolvemos True siempre porque el que manda es el status de encryption support que está en el environment, este es más restrictivo, solamente    
    
    return status

def encryption_set_temporal_encryption_method(enryption_method):
    """
    (c) Seachad - FGV
    
    Description:
    ---------------------- 
        encryption_set_temporal_encryption_method
        Permite cambiar el método de encriptación
    
    Extended Description:
    ---------------------
        _extended_summary_
        
    """
    _CACHE_set("TMP_encryption_method", enryption_method, singleton = True)
    
    pass

def encryption_get_std_method():
    import os
    return os.getenv("CM_ENCRYPTION_METHOD", "FERNET")

def encryption_get_temporal_encryption_method(encryption_method = None):
    """
    (c) Seachad - FGV
    
    Description:
    ---------------------- 
        get_temporal_encryption_method
        Si hay un método de encriptación temporal, manda sobre todos los demás
    
    Extended Description:
    ---------------------
        _extended_summary_
    
    Args:
    -----
            - encryption_method (_type_): _description_
    """
    # si hay un método temporal, manda
    tmp_encryption_method = _CACHE_get("TMP_encryption_method", singleton = True)
    if isinstance(tmp_encryption_method, str):
        # ? hay que convertir el método a su valor numérico
        tmp_encryption_method = get_encryption_method_by_name(tmp_encryption_method)
    if tmp_encryption_method != None:
        encryption_method = tmp_encryption_method
    
    return encryption_method


def get_include_encryption_prefix():
    """
    (c) Seachad - FGV
    
    Description:
    ---------------------- 
        get_include_encryption_prefix
        Devuelve si incluir o no el prefijo de encriptación
        Devuelve lo que viene en env, pero si hay un cambio temporal, lo aplica
    
    Extended Description:
    ---------------------
        _extended_summary_
    
    Returns:
    --------
            _type_: _description_
    """
    retorno = encryption_get_include_encryption_prefix()
    # tmp_include_prefix_code = encryption_get_temporal_include_encryption_prefix()
    # if tmp_include_prefix_code != retorno:
    #     retorno = tmp_include_prefix_code
    return retorno    

def encryption_get_include_encryption_prefix():
    """
    (c) Seachad - FGV
    
    Description:
    ---------------------- 
        encryption_get_include_encryption_prefix
        Devuelve si hay que operar con el encryption_prefix
    
    Extended Description:
    ---------------------
        _extended_summary_
    
    Args:
    -----
            - encryption_method (_type_): _description_
    """
    # si hay un método temporal, manda
    # include_encryption_prefix = os.getenv("CM_ENCRYPTION_INCLUDE_ENCRYPTION_PREFIX", "False")
    # if include_encryption_prefix.lower() == "true":
    #     return True
    return False

def encryption_get_temporal_include_encryption_prefix():
    """
    (c) Seachad - FGV
    
    Description:
    ---------------------- 
        get_temporal_encryption_method
        Si hay un método de encriptación temporal, manda sobre todos los demás
    
    Extended Description:
    ---------------------
        _extended_summary_
    
    Args:
    -----
            - encryption_method (_type_): _description_
    """
    # si hay un método temporal, manda
    tmp_include_encryption_prefix = _CACHE_get("TMP_encryption_include_encryption_prefix", singleton = True)
    return tmp_include_encryption_prefix

def encryption_set_temporal_include_encryption_prefix(include_encryption_prefix):
    """
    (c) Seachad - FGV
    
    Description:
    ---------------------- 
        encryption_set_temporal_include_encryption_prefix
        Si hay un método de encriptación temporal, manda sobre todos los demás
    
    Extended Description:
    ---------------------
        _extended_summary_
    
    Args:
    -----
            - encryption_method (_type_): _description_
    """
    # si hay un método temporal, manda
    _CACHE_set("TMP_encryption_include_encryption_prefix", include_encryption_prefix, singleton = True)
    return False

# =============================================================================
# Guardamos toda la información sensible (READ ONLY)
# =============================================================================

import cryptography
from cryptography.fernet import Fernet, InvalidToken

import base64
import os
from cryptography.hazmat.backends import default_backend
from cryptography.hazmat.primitives import hashes
from cryptography.hazmat.primitives.kdf.pbkdf2 import PBKDF2HMAC


# key = b""

# =============================================================================
# Inicializa una nueva clave de encriptacion
# =============================================================================
def initializeEncryptionKey():
    # # global key
    # key = Fernet.generate_key()
    # return key
    import os
    CM_Encrypt_Password = "CM_API_ROBOT_Encrypt_Password_v1"
    key = CM_Encrypt_Password
    # key = os.environ[key] 
    key = get_user_env(key)
    return key

# =============================================================================
# Genero una clave en base a un password
# =============================================================================


import hashlib

def _get_fixed_length_bytes_from_string(input_string, length=32):
    """
    Generates a consistent sequence of bytes of a fixed length based on the given input string.

    Args:
        input_string (str): The input string to be hashed.
        length (int): The desired length of the output byte sequence. Default is 32.

    Returns:
        bytes: A consistent sequence of bytes of the specified length derived from the input string.
    """
    # Create a SHA-256 hash of the input string
    hash_object = hashlib.sha256(input_string.encode())
    # Get the full hash as bytes
    full_hash = hash_object.digest()
    
    # If the required length is less than or equal to the hash length, truncate the hash
    if length <= len(full_hash):
        return full_hash[:length]
    else:
        # If the required length is greater than the hash length, repeat the hash bytes and truncate
        repeat_count = (length // len(full_hash)) + 1
        repeated_hash = (full_hash * repeat_count)[:length]
        return repeated_hash

import base64
ENCRYPTION_METHOD_FERNET = 1 # most secure but slower
ENCRYPTION_METHOD_AES = 2 # secure but faster
ENCRYPTION_METHOD_CHACHA20 = 3 # secure but faster
ENCRYPTION_METHOD_OBFUSCATE_BASE64 = 4 # not secure but esay

MAP_ENCRYPTION_CODE_TO_STR = {
    ENCRYPTION_METHOD_FERNET : "fernet",
    ENCRYPTION_METHOD_AES : "aes",
    ENCRYPTION_METHOD_CHACHA20 : "chacha20",
    ENCRYPTION_METHOD_OBFUSCATE_BASE64 : "obfuscate_base64"
}

MAP_ENCRYPTION_METHOD_STR_TO_CODE = {
    "fernet": ENCRYPTION_METHOD_FERNET,
    "aes": ENCRYPTION_METHOD_AES,
    "chacha20": ENCRYPTION_METHOD_CHACHA20,
    "obfuscate_base64": ENCRYPTION_METHOD_OBFUSCATE_BASE64
}

def get_encryption_method_by_code(encryption_method):
    name = None
    # name = MAP_ENCRYPTION_CODE_TO_STR.get(encryption_method, None)
    if encryption_method == ENCRYPTION_METHOD_FERNET:
        name = "FERNET"
    if encryption_method == ENCRYPTION_METHOD_AES:
        name = "AES"
    if encryption_method == ENCRYPTION_METHOD_CHACHA20:
        name = "CHACHA20"
    if encryption_method == ENCRYPTION_METHOD_OBFUSCATE_BASE64:
        name = "OBFUSCATE_BASE64"
    return name

def get_encryption_method_by_name(name):
    code = None
    # code = MAP_ENCRYPTION_METHOD_STR_TO_CODE.get(name.lower(), None)   
    if name.lower() == "fernet":
        code = ENCRYPTION_METHOD_FERNET
    if name.lower() == "aes":
        code = ENCRYPTION_METHOD_AES
    if name.lower() == "chacha20":
        code = ENCRYPTION_METHOD_CHACHA20
    if name.lower() == "obfuscate_base64":
        code = ENCRYPTION_METHOD_OBFUSCATE_BASE64
    return code

def generateEncryptionKey(password = None, encryption_method = None):
    """
    (c) Seachad
    
    Description:
    ---------------------- 
    generateEncryptionKey
    
    Dependiendo del método elegido, devuelve una clave para los métodos de encriptación soportados
    Los métodos de obfuscation siempre devuelven None, porque no necesitan clave
    
    Sabe qué método de encriptación se está utilizando
    
    Extended Description:
    ---------------------
    _extended_summary_
    
    
    Args:
    -----
        - password (_type_, optional): _description_. Defaults to None.
        - encryption_method (_type_, optional): _description_. Defaults to ENCRYPTION_METHOD_FERNET.
    
    Returns:
    --------
        _type_: _description_
    """

    tc._print(f"Generando encryption key con encryption_method {encryption_method}", tc.W)
    if encryption_method == None:
        encryption_method = encryption_get_method_to_encrypt()
    else:
        if isinstance(encryption_method, str):
            encryption_method = get_encryption_method_by_name(encryption_method)
        # si hay un método temporal, manda
        # encryption_method = encryption_get_temporal_encryption_method(encryption_method)
        
    key = password

    import base64
    if encryption_method == ENCRYPTION_METHOD_FERNET:
        if password == None: # si no mandamos password la coge del PATH del sistema
            tc._print(f"recogiendo el password de encriptación del sistema!!!!!", tc.W)
            password = initializeEncryptionKey()
            
        password_provided = password
        password = password_provided.encode()
        
        salt = b"\xb9\x1f|}'S\xa1\x96\xeb\x154\x04\x88\xf3\xdf\x05"
        
        kdf = PBKDF2HMAC(algorithm=hashes.SHA256(),
                        length=32,
                        salt=salt,
                        iterations=100000,
                        backend=default_backend())
        
        key = base64.urlsafe_b64encode(kdf.derive(password))
        _CACHE_set("ENC_CACHED_GENERATED_KEY_fernet", key, singleton = True)
        
    if encryption_method == ENCRYPTION_METHOD_AES:
        from Crypto.Cipher import AES
        from Crypto.Random import get_random_bytes
        import base64        
        
        # key = get_random_bytes(16)  # AES key must be either 16, 24, or 32 bytes long
        key = _get_fixed_length_bytes_from_string(password, 16)
        _CACHE_set("ENC_CACHED_GENERATED_KEY_aes", key, singleton = True)
        pass
    if encryption_method == ENCRYPTION_METHOD_CHACHA20:
        import os
        # key = os.urandom(32)  # ChaCha20 key must be 32 bytes long
        key = _get_fixed_length_bytes_from_string(password, 32)
        _CACHE_set("ENC_CACHED_GENERATED_KEY_chacha20", key, singleton = True)
        pass
    if encryption_method == ENCRYPTION_METHOD_OBFUSCATE_BASE64:
        key = None
        _CACHE_set("ENC_CACHED_GENERATED_KEY_obfuscated_base64", key, singleton = True)
        pass    
    
    _CACHE_set("ENC_CACHED_GENERATED_KEY", key, singleton = True)    
    
    return key    

def get_cached_fernet_key():
    return _CACHE_get("ENC_CACHED_GENERATED_KEY_fernet", singleton = True)

def get_cached_aes_key():
    return _CACHE_get("ENC_CACHED_GENERATED_KEY_aes", singleton = True)

def get_cached_chacha20_key():
    return _CACHE_get("ENC_CACHED_GENERATED_KEY_chacha20", singleton = True)

def get_cached_obfuscated_base64_key():
    return _CACHE_get("ENC_CACHED_GENERATED_KEY_obfuscated_base64", singleton = True)    

MAP_ENCRYPTION_CODE_TO_KEY = {
    ENCRYPTION_METHOD_FERNET : get_cached_fernet_key,
    ENCRYPTION_METHOD_AES : get_cached_aes_key,
    ENCRYPTION_METHOD_CHACHA20 : get_cached_chacha20_key,
    ENCRYPTION_METHOD_OBFUSCATE_BASE64 : get_cached_obfuscated_base64_key
}
    
def get_cached_key_for_encryption_method(encryption_method = None):
    """
    (c) Seachad
    
    Description:
    ---------------------- 
    get_cached_key_for_encryption_method
    
    Devuele la clave que se usará para el método de encriptación concreto
    
    Extended Description:
    ---------------------
    _extended_summary_
    
    
    Args:
    -----
        - encryption_method (puede ser str o int, optional): Método concreto. Defaults to None. Si es None, devuelve la clave para el método por defecto del sistema o el escogido en el momento, termporalmente
    
    Returns:
    --------
        _type_: _description_
    """
    if isinstance(encryption_method, str):
        encryption_method = get_encryption_method_by_name(encryption_method)
    if encryption_method == None:
        encryption_method = encryption_get_method_to_encrypt() # nos coge el método que está elegido por defecto o seleccionado temporalmente
    # el método temporal manda sobre todos los demás    
    # temporal_method = _CACHE_get("TMP_encryption_method", singleton = True)
    # if temporal_method != None:
    #     if isinstance(temporal_method, str):
    #         temporal_method = get_encryption_method_by_name(temporal_method)
    #     encryption_method = temporal_method    
        
    key = None
    # key_f = MAP_ENCRYPTION_CODE_TO_KEY.get(encryption_method, None)
    # if key_f != None:
    #     key = key_f()
    if encryption_method == ENCRYPTION_METHOD_FERNET:
        key = _CACHE_get("ENC_CACHED_GENERATED_KEY_fernet", singleton = True)
    if encryption_method == ENCRYPTION_METHOD_AES:
        key = _CACHE_get("ENC_CACHED_GENERATED_KEY_aes", singleton = True)
    if encryption_method == ENCRYPTION_METHOD_CHACHA20:
        key = _CACHE_get("ENC_CACHED_GENERATED_KEY_chacha20", singleton = True)
    if encryption_method == ENCRYPTION_METHOD_OBFUSCATE_BASE64:
        key = _CACHE_get("ENC_CACHED_GENERATED_KEY_obfuscated_base64", singleton = True)
    return key, encryption_method
    
# =============================================================================
# Guarda la clave en un fichero
# =============================================================================
def saveEncryptionKey(key, filename = 'key.key'):
    """
    (c) Seachad
    
    Description:
    ---------------------- 
    saveEncryptionKey
    
    Guarda la ultima clave guardada en un fichero key.key
    
    Extended Description:
    ---------------------
    _extended_summary_
    
    
    Args:
    -----
        - key (_type_): _description_
        - filename (str, optional): _description_. Defaults to 'key.key'.
    """
    
    writeRawFile(key, filename, "wb")
    

# =============================================================================
# Recupera la clave guardada en un fichero
# =============================================================================
def retrieveEncryptionKey(filename = 'key.key'):
    """
    (c) Seachad
    
    Description:
    ---------------------- 
    retrieveEncryptionKey
    
    Recuperan la clave usada en el sistema, la última clave usada está guardad en el fichero key.key
    
    Extended Description:
    ---------------------
    _extended_summary_
    
    
    Args:
    -----
        - filename (str, optional): _description_. Defaults to 'key.key'.
    
    Returns:
    --------
        _type_: _description_
    """
    
    key = readRawFile(filename, "rb" )    

    return key

def obfuscate(data, is_in_bytes: bool = False) -> str:
    """
    Obfuscates a given string or bytes using base64 encoding.
    
    :param data: The original data to obfuscate (string or bytes).
    :param is_bytes: A boolean indicating if the input data is in bytes format.
    :return: The obfuscated string.
    """
    import base64
    if is_in_bytes:
        encoded_bytes = base64.b64encode(data)
    else:
        encoded_bytes = base64.b64encode(data.encode('utf-8'))
    
    encoded_str = encoded_bytes.decode('utf-8')
    return encoded_str

def deobfuscate(data: str, is_in_bytes: bool = False):
    """
    Deobfuscates a given string using base64 decoding.
    
    :param data: The obfuscated data to deobfuscate (string).
    :param is_bytes: A boolean indicating if the output should be in bytes format.
    :return: The original data in the requested format (string or bytes).
    """
    import base64
    decoded_bytes = base64.b64decode(data.encode('utf-8'))
    
    if is_in_bytes:
        return decoded_bytes
    else:
        return decoded_bytes.decode('utf-8')


from Crypto.Cipher import AES

import base64

def pad(data):
    return data + b"\0" * (AES.block_size - len(data) % AES.block_size)

def unpad(data):
    return data.rstrip(b"\0")

def encrypt_aes(key, plaintext):
    data = pad(plaintext.encode())
    iv = get_random_bytes(AES.block_size)
    cipher = AES.new(key, AES.MODE_CBC, iv)
    encrypted = cipher.encrypt(data)
    encrypted = base64.b64encode(iv + encrypted).decode('utf-8')
    return encrypted

def decrypt_aes(key, ciphertext):
    raw = base64.b64decode(ciphertext)
    iv = raw[:AES.block_size]
    encrypted = raw[AES.block_size:]
    cipher = AES.new(key, AES.MODE_CBC, iv)
    decrypted = unpad(cipher.decrypt(encrypted))
    decrypted = decrypted.decode('utf-8')
    return decrypted



from cryptography.hazmat.primitives.ciphers import Cipher, algorithms, modes
from cryptography.hazmat.backends import default_backend
import os
import base64

def encrypt_chacha20(key, plaintext):
    nonce = os.urandom(16)
    algorithm = algorithms.ChaCha20(key, nonce)
    cipher = Cipher(algorithm, mode=None, backend=default_backend())
    encryptor = cipher.encryptor()
    encrypted = encryptor.update(plaintext.encode())
    encrypted = base64.b64encode(nonce + encrypted).decode('utf-8')
    return encrypted

def decrypt_chacha20(key, ciphertext):
    raw = base64.b64decode(ciphertext)
    nonce = raw[:16]
    encrypted = raw[16:]
    algorithm = algorithms.ChaCha20(key, nonce)
    cipher = Cipher(algorithm, mode=None, backend=default_backend())
    decryptor = cipher.decryptor()
    decrypted = decryptor.update(encrypted)
    decrypted = decrypted.decode('utf-8')
    return decrypted

def get_user_env(key, default = None):
    import os
    return os.getenv(key, default)

# ENCRYPTION_SUPPORT
CACHE_KEY = get_user_env("CM_ENCRYPTION_PASSWORD")

ENCRYPTION_SUPPORT = get_user_env("CM_ENCRYPTION_SUPPORT", "False")
if ENCRYPTION_SUPPORT.lower() == "true":
    ENCRYPTION_SUPPORT = True
else:
    ENCRYPTION_SUPPORT = False
ENCRYPTION_PREFIX_MARK = "@SCHD18"    
encryption_mark = os.getenv("CM_ENCRYPTION_MARK", ENCRYPTION_PREFIX_MARK)
encrytion_mark = eval(f"b'{encryption_mark}'")  

def generate_encryption_prefix_mark(encription_method = ENCRYPTION_METHOD_FERNET):
    return f"{ENCRYPTION_PREFIX_MARK}_{encription_method}@"

def is_string_encrypted(string, remove_encryption_mark = False):
    """
    (c) Seachad
    
    Description:
    ---------------------- 
    is_string_encrypted
    
    Devuelve si el string está encriptado, mirando si comienza con una marca de encriptación    
    devuelve el string sin la marca si así se le pide
    devuelve también el método de encriptación usado
    
    Extended Description:
    ---------------------
    _extended_summary_
    
    
    Args:
    -----
        - is_encrypted: boolean
        - string (_type_): with or without the encryption prefix
        - encryption method (int): encryption method used
    """
    
    is_encrypted = False
    encryption_method = None
    clean_string = string
    # miramos si el string trae una marca de encriptación, y devolvemos si está encriptado, la string (original o quitando la marca) y el tipo de encriptación usado originalmente
    # comienza el string con ENCRYPTION_PREFIX_MARK?
    if string.find(ENCRYPTION_PREFIX_MARK) != -1:
        is_encrypted = True
        if remove_encryption_mark:
            # quitamos hasta la segunda aparición de @ empezando por el principio de string
            temporal_string = string.replace(f"{ENCRYPTION_PREFIX_MARK}_", "")
            encryption_method = int(temporal_string[0])
            clean_string = remove_encryption_mark(string)
    
    return is_encrypted, clean_string, encryption_method    

def remove_encryption_mark(s: str) -> str:
    """
    Removes the prefix '@SCHD18_' followed by a digit and ending with '@' from the input string if it exists.
    
    :param s: The input string.
    :return: The string with the prefix removed if it existed, otherwise the original string.
    """
    prefix = ENCRYPTION_PREFIX_MARK
    # if s.startswith(prefix) and len(s) > len(prefix) + 1 and s[len(prefix)].isdigit() and s[len(prefix) + 1] == '@':
    #     return s[len(prefix) + 2:]
    return s[len(prefix) + 2:]   

# =============================================================================
# Encripta una cadena usando una key. 
# =============================================================================

def encrypt(r_data, key, is_in_bytes = False, returns_string = False, encryption_method = ENCRYPTION_METHOD_FERNET):
    """
    ojo! key tiene que venir de una función previa key = 
    (password = encrypt_key)
    bytes == True significa que la información YA viene en bytes (por ejemplo, el pickle de un objeto)
    
    # !!! pendiente!
    02/11/2023
    ojo! si lo que me viene ya está encriptado (porque lo puedo desencriptar o porque tiene una marca binaria antes)
    entonces no vuelvo a encriptarlo
    
    Ejemplos:
    ---------
    
    >>> # ejemplo para ENCRYPTION_METHOD_OBFUSCATE_BASE64, se usa igual para todos los métodos
    >>> CACHED_GENERATED_KEY = common.generateEncryptionKey(password = CACHE_KEY, encryption_method = ENCRYPTION_METHOD_OBFUSCATE_BASE64)    
    >>> key_encrypt = CACHED_GENERATED_KEY
    >>> encrypt_encrypted.append(common.encrypt(test_string, key_encrypt, returns_string = True, encryption_method = ENCRYPTION_METHOD_OBFUSCATE_BASE64))
 
    """
    # encryption_method = encryption_get_method_to_encrypt()
    # include_encryption_prefix = get_include_encryption_prefix()
    include_encryption_prefix = False
    
    if encryption_method == ENCRYPTION_METHOD_FERNET:
        if key == None:
            key = _CACHE_get("ENC_CACHED_GENERATED_KEY_fernet", singleton=True)            
            
        if r_data == None: 
            return r_data 
        try:
            fernet = Fernet(key) 
        except Exception as e:
            tc._print(f"error {e}", tc.R)
            a=999
        if is_in_bytes == True: 
            b = r_data 
        else: 
            l_data = str(r_data) 
            b = bytes(l_data, 'utf-8')    

        encrypted = fernet.encrypt(b)    

        if returns_string: 
            b_encrypted = encrypted.decode("utf-8") 
            b64_encrypted = base64.b64encode(b_encrypted.encode()).decode() 
            encrypted = b64_encrypted 
            if include_encryption_prefix:
                encrypted = f"{generate_encryption_prefix_mark(encryption_method)}{encrypted}"

    if encryption_method == ENCRYPTION_METHOD_AES:
        if key == None:
            key = _CACHE_get("ENC_CACHED_GENERATED_KEY_aes", singleton=True)
        encrypted = encrypt_aes(key, r_data)
        if include_encryption_prefix:
            encrypted = f"{generate_encryption_prefix_mark(encryption_method)}{encrypted}"        
        pass
    
    if encryption_method == ENCRYPTION_METHOD_CHACHA20:
        if key == None:
            key = _CACHE_get("ENC_CACHED_GENERATED_KEY_chacha20", singleton=True)
        encrypted = encrypt_chacha20(key, r_data)
        if include_encryption_prefix:
            encrypted = f"{generate_encryption_prefix_mark(encryption_method)}{encrypted}"        
        pass
    
    if encryption_method == ENCRYPTION_METHOD_OBFUSCATE_BASE64:
        if key == None:
            key = _CACHE_get("ENC_CACHED_GENERATED_KEY_obfuscated_base64", singleton=True)
        encrypted = obfuscate(r_data, is_in_bytes = is_in_bytes)
        if include_encryption_prefix:
            encrypted = f"{generate_encryption_prefix_mark(encryption_method)}{encrypted}"        
        pass

    return encrypted  

def isBase64(sb):
    try:
        if isinstance(sb, str):
            # If there's any unicode here, an exception will be thrown and the function will return false
            sb_bytes = bytes(sb, 'ascii')
        elif isinstance(sb, bytes):
            sb_bytes = sb
        else:
            raise ValueError("Argument must be string or bytes")
        return base64.b64encode(base64.b64decode(sb_bytes)) == sb_bytes
    except Exception:
        return False


def decrypt(e_data, key, returns_string = False, encryption_method = ENCRYPTION_METHOD_FERNET):  
    """ 
    ojo! key tiene que venir de una función previa key = generateEncryptionKey(password = encrypt_key) 
    returns_string devuelve una cadena que empaqueta la información para ser guardada en un campo TEXT de BBDD, por ejemplo 

    Ejemplos:
    ---------
    >>> encrypt_decrypted.append(common.decrypt(encrypted, key_encrypt, returns_string = True, encryption_method = ENCRYPTION_METHOD_OBFUSCATE_BASE64))

    """     
    # si hay un método temporal, manda
    
    if e_data == None: 
        return e_data     
    
    if encryption_method == ENCRYPTION_METHOD_FERNET:
        # if key == None:
        #     key = _CACHE_get("ENC_CACHED_GENERATED_KEY_fernet", singleton=True)
        # si viene None no lo encriptamos 
        if e_data == None: 
            return e_data 
        
        decrypted = e_data 

        if returns_string: 
            # sólo se pueden encriptar cadenas 
            if not isinstance(e_data, str): 
                return e_data 
            # trabajamos con bytes para poder guardar la información en tablas 
            if isBase64(decrypted): 
                try: 
                    b64_from_table = base64.b64decode(decrypted).decode() 
                    b64_bytes = b64_from_table.encode("utf-8") 
                    e_data = b64_bytes 
                except Exception as e: 
                    a = 999 
                    pass 

        # is_encrypted ya devuelve la desencritación si puede o False si no está encriptado
        decrypted_value = is_encrypted(e_data, key) 
        if decrypted_value: 
            decrypted = decrypted_value 

        if returns_string: 
            if isinstance(decrypted, bytes): # chequear si son bytes lo que tengo que devolver, para convertirlo a string 
                decrypted = decrypted.decode("utf-8") 

    if encryption_method == ENCRYPTION_METHOD_AES:
        # if key == None:
        #     key = _CACHE_get("ENC_CACHED_GENERATED_KEY_aes", singleton=True)
        decrypted = decrypt_aes(key, e_data)
        pass
    
    if encryption_method == ENCRYPTION_METHOD_CHACHA20:
        # if key == None:
        #     key = _CACHE_get("ENC_CACHED_GENERATED_KEY_chacha20", singleton=True)
        decrypted = decrypt_chacha20(key, e_data)
        pass
    
    if encryption_method == ENCRYPTION_METHOD_OBFUSCATE_BASE64:
        # if key == None:
        #     key = _CACHE_get("ENC_CACHED_GENERATED_KEY_obfuscated_base64", singleton=True)
        decrypted = deobfuscate(e_data)
        pass

    return decrypted 

def is_encrypted(e_data, key ):
    """
    solo para ENCRYPTION_METHOD_FERNET
    ojo! key tiene que venir de una función previa key = generateEncryptionKey(password = encrypt_key)
    returns 
        - decrypted information si la información está encriptada previamente o 
        - False si no está encriptada
    """

    # probar este código
    # source: https://stackoverflow.com/questions/61137072/how-to-check-if-a-key-is-already-encrypted-while-using-fernet

    if not isinstance(e_data, bytes): 
        # sólo puede estar encriptado si son bytes, si recibimos otra cosa, no está encriptado
        return False 

    # key = Fernet.generate_key()
    try:
        f = Fernet(key) 
    except Exception as e:
        a = 999
 
    try: 
        decrypted = f.decrypt(e_data, None) 
        return decrypted 
    except InvalidToken: 
        return False 

# =============================================================================
# Ofusca una string
# Mantiene el mismo número de caracteres - sólo ofusca letras y números, el resto lo deja como está 
# =============================================================================
# def ofuscate(data):
#     """ 
#     Desription: 
#     -----------

#     recibe una cadena y la ofusca (cambia letras y números)
#     mantiene el mismo formato y longitud

#     """

#     # data = "bb26d856-8d57-486d-9615-238e89c398b2"
#     # _print(data)

#     numeros = ["0", "1", "2", "3", "4", "5", "6", "7", "8", "9"]
#     import codecs
#     # ofusco las letras
#     data = codecs.encode(data, 'rot13')
#     # ahora ofusco los números
#     import random
#     ofuscated_data = ""
#     for idx in range(len(data)):
#         c = data[idx]
#         if c in numeros:
#             c = random.randint(0,9)

#         ofuscated_data += str(c)        
#     # _print(ofuscated_data)

#     return data

def compress(obj):


    """
    (c) Seachad
    
    Description:
    ---------------------- 
    compress
    
    Comprime un objeto en memoria
    
    Extended Description:
    ---------------------
    _extended_summary_
    
    
    Args:
    -----
        - obj (_type_): _description_
    """    

    import zlib
    import pickle
    return zlib.compress(pickle.dumps(obj, pickle.HIGHEST_PROTOCOL))

    pass

def uncompress(obj):
    """
    (c) Seachad
    
    Description:
    ---------------------- 
    uncompress
    
    Descomprime un objeto y lo devuelve
    
    Extended Description:
    ---------------------
    _extended_summary_
    
    
    Args:
    -----
        - obj (_type_): _description_
    """    
    import zlib
    import pickle
    return pickle.loads(zlib.decompress(obj))


# =============================================================================
# Encripta un fichero con la clave que se le proporciona
# =============================================================================
def encryptFile(infilename = "configuration_machine.json", outfilename = "VM1.encrypted", key = "p"):
    
    # with open(infilename, 'rb') as f:
    #     data = f.read()

    data = lowlevel_file_opperation(infilename, FO_READ, 'rb')
    
    fernet = Fernet(key)
    encrypted = fernet.encrypt(data)
    
    # with open(outfilename, 'wb') as f:
    #     f.write(encrypted)
    lowlevel_file_opperation(infilename, FO_WRITE, 'wb', encrypted)

# =============================================================================
# Encripta un fichero con la clave que se le proporciona
# =============================================================================
def encrypt_to_file(infilename = "configuration_machine.json", data = "", key = "p"):
    
    # with open(infilename, 'rb') as f:
    #     data = f.read()
    
    fernet = Fernet(key)
    encrypted = fernet.encrypt(data)
    
    # with open(outfilename, 'wb') as f:
    #     f.write(encrypted)
    lowlevel_file_opperation(infilename, FO_WRITE, 'wb', encrypted)



FO_WRITE = 1
FO_READ = 0
def lowlevel_file_opperation(filename, logical_mode = FO_WRITE, mode = "rb", buffer = ""):
    
    data = None
    
    # f = open(filename, mode)
    # try:
    #     if logical_mode == FO_WRITE:
    #         f.write(buffer)
    #     if logical_mode == FO_READ:
    #         data = f.read()
    # except:
    #     raise Exception(f"error en fichero {filename}")

    # f.close()
    
    if logical_mode == FO_WRITE:
        writeRawFile(buffer, filename, "wb")
    if logical_mode == FO_READ:
        data = readRawFile(filename, "rb")
    
    
    
    return data

# =============================================================================
# Desencripta un fichero con la clave que se le proporciona
# =============================================================================
def decrypFile(infilename = "configuration_machine.json", key = "p"):
    """ Desencripta un fichero con la clave que se le proporciona """
    
    # with open(infilename, 'rb') as f:
    #     data = f.read()

    decrypted = None
    data = lowlevel_file_opperation(infilename, FO_READ, 'rb')
    if not data == None:
        fernet = Fernet(key)
        decrypted = fernet.decrypt(data)
    
    return decrypted

# =============================================================================
# Encripta un fichero con la clave que se le proporciona
# =============================================================================
def encryptJSONFile(filename = "configuration_machine.json", data = "", key = "p"):
    """ encripta un fichero JSON con la clave que se le proporciona """
    

    jsondata = json.dumps(data)
    encrypted = encrypt(jsondata, key)
    
    # with open(filename, 'wb') as f:
    #     f.write(encrypted)

    lowlevel_file_opperation(filename, FO_WRITE, "wb", encrypted)

# =============================================================================
# Desencripta un fichero con la clave que se le proporciona
# =============================================================================
def decryptJSONFile(infilename = "configuration_machine.json", key = "p"):
    """ Desencripta un fichero JSON con la clave que se le proporciona """

    # with open(infilename, 'rb') as f:
    #     data = f.read()
    # f = open(infilename, 'rb')
    # data = f.read()        
    # f.close()
    
    data = readRawFile(infilename, "rb")
    
    
    fernet = Fernet(key)
    decrypted = fernet.decrypt(data)
    
    decrypted = decrypted.decode("ascii")
    
    # decrypted = decrypted.replace("'", '"')
    
    # decrypted = decrypted.replace('\'','\"')
    
    decrypted = json.loads(decrypted)
    
    return decrypted


# =============================================================================
# Ejemplo de uso de encriptación y desencriptación
# =============================================================================
def _testEncriptarYDesencriptar():
    # secuencia de trabajo
    # 1) generar clave
    # 2) guardar clave
    # 3) encriptar y desencriptar ficheros con la clave guardada, dándole como parámetro el fichero dónde está guardada    
    
    
    # =============================================================================
    # Generamos la clave
    # =============================================================================
    key = generateEncryptionKey()
    # =============================================================================
    # Guardamos y recuperamos la clave en un fichero (sólo si la queremos guardar, pero no es aconsejable)
    # =============================================================================
    filename = "key.key"
    saveEncryptionKey(key, filename = filename)
    key = retrieveEncryptionKey(filename = filename)
    # =============================================================================
    # Encriptamos un fichero
    # =============================================================================
    infilename = "Documento.txt"
    outfilename = "Documento.txt.encrypted"
    encryptFile(infilename, outfilename, key)
    
    # =============================================================================
    # Desencriptamos el fichero
    # =============================================================================
    infilename = "Documento.txt.encrypted"
    data = decrypFile(infilename, key)

""" PERSISTENCIA CON DOTENV ENCRIPTADO """ 
""" DOTENV """
try:            
    import dotenv
except:
    import subprocess
    name = "python-dotenv"
    
    def install(name):
        subprocess.call(['pip', 'install', name])
    
    install(name)    
    import dotenv  

# import os
# from dotenv import load_dotenv, set_key, find_dotenv

dotenv_file = dotenv.find_dotenv()
dotenv.load_dotenv(dotenv_file)

""" CONFIG PARSER """
try:            
    import configparser
except:
    import subprocess
    name = "configparser"
    
    def install(name):
        subprocess.call(['pip', 'install', name])
    
    install(name)    
    import configparser

class CONFIG:

    config_file = "admin_library.cfg" # fichero por defecto de configuración
    encrypt_key = "M1Encr1ptKe1" # clave por defecto, en las funciones get_env y set_env pueden emplearse claves locales, adicionalmente

    def __init__(self, config_filename = "admin_library.cfg", encrypt_key = "M1Encr1ptKe1"):
        self.config_file = config_filename
        self.encrypt_key = encrypt_key

    # # =============================================================================
    # # Recupera una clave del fichero .env (con su valor encriptado en bytes)
    # # =============================================================================
    def get_env(self, parent_key, key, encrypted = False, encrypt_key = None):

        config = configparser.ConfigParser()
        config.read(self.config_file)
        
        l_encrypt_key = self.encrypt_key        
        value = None
        if parent_key in config:
            value = config[parent_key].get(key, None)
            if not value == None:
                if encrypted:
                    if not encrypt_key == None:
                        l_encrypt_key = encrypt_key                    
                    # value = value.replace("b'", "")
                    # value = value[:-1]
                    # convierte a bytes para desencriptar
                    try:
                        value = bytes(value, "ascii")
                        value = decrypt(value, generateEncryptionKey(password = l_encrypt_key))
                        value = value.decode("ascii")    
                    except:
                        value = config[parent_key].get(key, None) 
            
        return value

    # =============================================================================
    # Escribe una clave con su valor encriptado en bytes en .env    
    # =============================================================================
    def set_env(self, parent_key, key, value, encrypted = False, encrypt_key = None):
        config = configparser.ConfigParser()
        config.read(self.config_file)
        
        l_encrypt_key = self.encrypt_key
        if encrypted:
            if not encrypt_key == None:
                l_encrypt_key = encrypt_key
            value = encrypt(value, generateEncryptionKey(password = l_encrypt_key))
            # convierte a string para guardar
            value = value.decode("ascii")    
            
        if not parent_key in config:
            config[parent_key] = {}
        
        config[parent_key][key] = value

        with open(self.config_file, 'w') as configfile: # file con gestión de contexto
            config.write(configfile)

""" guarda y recupera environments encriptados """

def set_environment(diccionario = None, name = "ENV", filename = "env.encrypted"):
    """
    genera un environment de ejecución para API_ROBOT

    Parameters
    ----------
    diccionario : TYPE, optional
        DESCRIPTION. The default is None.

    Returns
    -------
    None.

    """
    if not diccionario == None:
        import ar_managers
       
        em = ar_managers._persistence_manager_json_encrypted(_name = name, _filename = filename)
        """ grabamos la info """
        em._put(diccionario)


def get_environment(name = "ENV", filename = "env.encrypted"):
    """
    carga un environment de ejecución para mongodb, cargará mongodb.encrypted (o el fichero que se le pida)
    con los accesos a bases de datos y tablas, con claves encriptadas de acceso

    Parameters
    ----------
    name : TYPE, optional
        DESCRIPTION. The default is "ENV".

    Returns
    -------
    None.

    """
    import ar_managers
    em = ar_managers._persistence_manager_json_encrypted(_name = name, _filename = filename)

    """ recuperamos la info """
    diccionario = em._get()
    return diccionario


# =============================================================================
# Wrapper de llamada síncrona y asíncrona a scripts de Python
# Acepta el script y los parametros por separado
# =============================================================================
def callToScript(script_to_call = "", parameters= "", pasync = False):
    #cadena_ejecucion = f'{script_to_call} "{parameters}"'
    cadena_ejecucion = f"python \"{script_to_call}\" \"{parameters}\""
    return callToScriptComplete(script_to_call_complete = cadena_ejecucion, pasync = pasync)

# =============================================================================
# Wrapper de llamada síncrona y asíncrona a scripts de Python
# Recibe el comando y parámetros todo en una cadena
# =============================================================================
def callToScriptComplete(script_to_call_complete = "", pasync = False):
    #cadena_ejecucion = f'{script_to_call} "{parameters}"'
    cadena_ejecucion = script_to_call_complete
    if pasync == True:
        _print(f"callToScriptComplete: Ejecución ASYNC de {cadena_ejecucion}")
        resultado = os.system(cadena_ejecucion)
        #from subprocess import Popen
        import subprocess

        p = subprocess.Popen([cadena_ejecucion],
                     stdout=subprocess.PIPE, 
                     stderr=subprocess.PIPE,
                     shell=True,
                     creationflags=subprocess.CREATE_NEW_CONSOLE,
                     close_fds = True)
    else:
        _print(f"callToScriptComplete: Ejecución SYNC de {cadena_ejecucion}")
        resultado = os.system(cadena_ejecucion)
    return

# ---- .
# ---- * TRABAJAR CON LISTAS Y DICCIONARIOS -----------------------------------------------------------------------------------------------


# =============================================================================
# Obtiene un elemento de una lista, si el indice no existe devuelve default valuee
# =============================================================================
def getSafeListIndex(lista, index, default_value = ""):
    retorno = default_value
    try:
        retorno = lista[index]
    except:
        retorno = default_value
        
    return retorno        

# =============================================================================
# Obtiene un elemento de un diccionario, si la clave no existe devuelve default valuee
# =============================================================================
def getSafeDictionaryKey(ldict, key, default_value = ""):
    retorno = default_value

    if key in ldict:
        retorno = ldict[key]

        
    return retorno  


def test_getSafeListIndex():
    lista = [1, 10, 20]
    _print(f"{getSafeListIndex(lista, 5, 'kk')}")    
    
    
def find_key(input_dict, value):
    """ devuelve la clave para un valor encontrado en un diccionario """
    return next((k for k, v in input_dict.items() if v == value), None)
    
def get_value_in_dictionary(d, v = None):
    """ 
    busca si el valor v está en d, si es así devuelve True, en otro caso devuelve False
    devuelve también la clave dónde ha encontrado el valor
    sólo se permiten valores atómicos (no listas, diccionarios o cosas por el estilo 
    opera en modo recursivo         

    Parameters
    ----------
    d : TYPE dict
        DESCRIPTION. Diccionario en el que buscar
    v : TYPE, optional
        DESCRIPTION. The default is None. valor que queremos encontrar

    Raises
    ------
    Exception
        DESCRIPTION.

    Returns
    -------
    TYPE
        DESCRIPTION.
    TYPE
        DESCRIPTION.

    """
    
    retorno = False
    
    if not isinstance(d, dict):
        raise Exception(f"get_value_in_dictionary: {d} no es un diccionario")
        return None, None
    
    if not isinstance(v, (str, int)):
        raise Exception(f"get_value_in_dictionary: {v} no es un valor válido")
        return None, None
    
    if v in d.values(): # si está el valor
        return True, find_key(d, v)
    
    for lk,lv in d.items():
        # if lk == k: # si está la clave
        #     retorno = True
        #     break
        if isinstance(lv,list):
            for i in lv: # cada i es un diccionario
                retorno = get_value_in_dictionary(i, v)
                if retorno == True:
                    return True, find_key(i, v)
        if isinstance(lv,dict): # si tengo un diccionario como valor, también busco
            retorno = get_value_in_dictionary(lv, v)
            if retorno == True:
                return True, find_key(lv, v)             
        
    return retorno, None
    
    pass    
    
def get_key_value_in_dictionary(d, k, v):
    """
    busca un par clave,valor en un diccionario, profundizando si tiene listas o diccionarios anidados
    devuelve la primera clave que encuentra con el valor que se le pide

    Parameters
    ----------
    d : TYPE dict
        DESCRIPTION. diccionario a inspeccionar
    k : TYPE str
        DESCRIPTION. clave a buscar
    v : TYPE str/int
        DESCRIPTION. valor a buscar

    Raises
    ------
    Exception
        DESCRIPTION.

    Returns
    -------
    TYPE
        DESCRIPTION.

    """
    retorno = False
    
    if not isinstance(d, dict):
        raise Exception(f"get_key_in_dictionary: {d} no es un diccionario")
        return None
    
    if not isinstance(k, str):
        raise Exception(f"get_key_in_dictionary: {k} no es una clave válida (no es str)")
        return None
    
    if k in d: # si está la clave
        if d[k] == v:
            return True
    
    for lk,lv in d.items():
        # if lk == k: # si está la clave
        #     retorno = True
        #     break
        if isinstance(lv,list):
            for i in lv: # cada i es un diccionario
                if isinstance(i, dict): # sólo lo mando a resolver si es un diccionario
                    retorno = get_key_value_in_dictionary(i, k, v)
                    if retorno == True:
                        return True
        if isinstance(lv,dict): # si tengo un diccionario como valor, también busco
            retorno = get_key_value_in_dictionary(lv, k, v)
            if retorno == True:
                return True                
        
    return retorno
    

def get_key_in_dictionary(d, k = None):
    """ busca si la clave k está en d, si es así devuelve True, en otro caso devuelve False 
    devuelve también el valor encontrado
    opera en modo recursivo
    """
    
    retorno = False
    
    if not isinstance(d, dict):
        # raise Exception(f"get_key_in_dictionary: {d} no es un diccionario")
        return None, None
    
    if not isinstance(k, str):
        raise Exception(f"get_key_in_dictionary: {k} no es una clave válida (no es str)")
        return None, None
    
    if k in d: # si está la clave
        return True, d[k]
    
    for lk,lv in d.items():
        # if lk == k: # si está la clave
        #     retorno = True
        #     break
        if isinstance(lv,list):
            for i in lv: # cada i es un diccionario
                retorno, valor = get_key_in_dictionary(i, k)
                if retorno == True:
                    return True, valor
        if isinstance(lv,dict): # si tengo un diccionario como valor, también busco
            retorno, valor = get_key_in_dictionary(lv, k)
            if retorno == True:
                return True, valor               
        
    return retorno, None
    
    pass

def get_key_dictionary_in_dictionary(d, k = None):
    """ busca si la clave k está en d, si es así devuelve True, en otro caso devuelve False 
    devuelve también el valor encontrado
    opera en modo recursivo
    
    devuelve el valor y el diccionario que la contiene
    
    """
    
    retorno = False

    if not isinstance(k, str): # la clave tiene que ser un string!!!
        raise Exception(f"get_key_in_dictionary: {k} no es una clave válida (no es str)")
        return None, None, None
        
    if isinstance(d, list): # si es una lista, rastreo todos los elementos
        for l in d:
            retorno, valor, dic = get_key_dictionary_in_dictionary(l, k)
            if retorno == True: # ha encontrado algo!
                return True, valor, dic
    
    if not isinstance(d, dict): # si no es un diccionario, no puedo seguir!
        return None, None, None

    if k in d: # si está la clave
        return True, d[k], d
    
    for lk,lv in d.items():
        if isinstance(lv,list):
            for i in lv: # cada i es un diccionario
                retorno, valor, dic = get_key_dictionary_in_dictionary(i, k)
                if retorno == True:
                    return True, valor, dic
        if isinstance(lv,dict): # si tengo un diccionario como valor, también busco
            retorno, valor, dic = get_key_dictionary_in_dictionary(lv, k)
            if retorno == True:
                return True, valor, dic               
        
    return retorno, None, None

def new_list(dict1, list2, mode = 0):
    """
    hace merge de dos listas (normalmente contiene diccionarios como elementos)

    Parameters
    ----------
    dict1 : TYPE dict
        DESCRIPTION.
    list2 : TYPE list
        DESCRIPTION.
    mode : TYPE, optional
        DESCRIPTION. The default is 0. 0 = inserta elementos por arriba (respecta el orden original), -1 = inserta elementos al final

    Returns
    -------
    None. Nueva lista creada

    """
    list_to_return = []
    
    
    """ lo que viene en dict1 es un dictionary """
    list_to_return = list2.copy() 
    if mode == 0:
        list_to_return.insert(0, dict1)
    if mode == -1:
        list_to_return.append(dict1) 
    
    
    
    """ esto sería si dict1 es una lista """
    # if mode == 0:
    #     list_to_return = list2.copy()        
    #     for i in range(len(list1)-1, -1, -1): # rastreo la lista desde el final hasta el principio
    #         list_to_return.insert(0, list1[i]) # inserto al principio
    # if mode == -1:
    #     list_to_return = list2.copy()        
    #     for i in list1: # rastreo la lista desde el final hasta el principio
    #         list_to_return.append(i) # inserto al principio
        
    return list_to_return  

# ---- .
# ---- * EXCEL -----------------------------------------------------------------------------------------------


import pandas as pd

# @ar_decorator_functions.catch_exceptions
def readEXCEL(filename, sheet_name = None):
    """
    lee un fichero EXCEL y lo devuelve como un DF

    Parameters
    ----------
    filename : TYPE
        DESCRIPTION.


    Returns
    -------
    None.

    """

    df = None

    if not sheet_name == None:
        df = pd.read_excel(filename, sheet_name, engine='openpyxl') # ojo, habría que instalar esta librería para dar soporte a XLSX...
    else:
        df = pd.read_excel(filename, engine='openpyxl')
    
    # df.describe()
    
    return df
    
def f_deal_with_excel(excel_filename, excel_filename_list = [], df_list = [], replace_sheets = True):
    """
    en un excel crear los sheets en base a los nombres de excel files o dataframes que recibe
    si el sheet existe, se lo carga en base a lo que diga el parámetro replace_sheet
    si el excel no existe, lo crea y guarda la información que le llega
    los sheets previos que pueda tener el excel se mantienen

    A tener en cuenta en esta versión:
        el fichero excel_filename, por cada sheet nueva, se carga entero de nuevo

    Parameters
    ----------
    excel_filename : TYPE
        DESCRIPTION. fichero excel con el path completo
    excel_filename_list : TYPE, optional
        DESCRIPTION. The default is []. cada elemento es un diccionario que contiene el nombre del sheet y el nombre del excel a cargar
    replace_sheets : TYPE, optional
        DESCRIPTION. The default is True. si está a True, reemplaza el sheet si este existe

    Returns
    -------
    None.

    """
    

    import pandas as pd
    from pandas import ExcelWriter
    from openpyxl import load_workbook
    
    retorno = False


    for el in excel_filename_list:    

        book = None
        
        sheet_name = el["sheet_name"]
        filename = el["excel_filename"]
        try:
            df = readEXCEL(filename)        
            _print(f"lectura de [{filename}] exitosa", G)
            try:
                book = load_workbook(excel_filename)
            except:
                # el fichero no existe, lo creamos
                _print(f"el fichero [{excel_filename}] no existe, se creará vacío", B)                 
            
            with pd.ExcelWriter(excel_filename, engine='openpyxl') as writer:
                if book is not None:
                    writer.book = book
                try:
                    # intento borrar la sheet que me dicen, si es que existe
                    if replace_sheets:
                        book.remove(book[sheet_name])
                except:
                    _print(f"la sheet [{sheet_name}] no existe, la creamos nueva", B)                 

                df.to_excel(writer, sheet_name, index=False)
                _print(f"grabada la sheet [{sheet_name}] en [{excel_filename}]", Y)                 
        except:
            _print(f"lectura de [{filename}] exitosa", R)            
    
    
    for el in df_list:    

        book = None
        
        sheet_name = el["sheet_name"]
        df = el["df"]
        try:
            book = load_workbook(excel_filename)
        except:
            # el fichero no existe, lo creamos
            _print(f"el fichero [{excel_filename}] no existe, se creará vacío", B)                 
        
        with pd.ExcelWriter(excel_filename, engine='openpyxl') as writer:
            if book is not None:
                writer.book = book
            try:
                # intento borrar la sheet que me dicen, si es que existe
                if replace_sheets:                
                    book.remove(book[sheet_name])
            except:
                _print(f"la sheet [{sheet_name}] no existe, la creamos nueva", B)                 

            df.to_excel(writer, sheet_name, index=False)
            _print(f"grabada la sheet [{sheet_name}] en [{excel_filename}]", Y)                 
    
    return retorno
    pass
    
# ---- .
# ---- * CSV -----------------------------------------------------------------------------------------------

def filename_parts(filename):
    """

    Description: 
    ------------
    filename_parts Devuelve las partes de un fichero en la forma (directorio, filename, extension, exists) - el último parámetro determina si existe ya

    Extended Description: 
    ---------------------
    _extended_summary_

    Args:
    -----
        - filename (str): filename conteniendo el path completo

    Raises:
    -------
        Exception: _description_
        Exception: _description_
        Exception: _description_
        Exception: _description_
        Exception: _description_
        Exception: _description_
        Exception: _description_

    Returns:
    --------
        tuple: (directorio, filename, extension, exists)
    """

    from pathlib import Path
    p = Path(filename)
    return p.parents[0], p.stem ,p.suffix, p.exists()


import pandas as pd
import csv

tipos = ["pandas", "list", "dict", "raw", "csv", "json"] # tipos de devolución de fichero

# @ar_decorator_functions.catch_exceptions
def deal_with_file(filename, extension = ".json", create_directory = True):
    """
    se encarga de controlar que el fichero tiene la extensión que se le pide y crea el directorio si no existe

    Parameters
    ----------
    filename : TYPE
        DESCRIPTION.
    extension : TYPE, optional
        DESCRIPTION. The default is ".json".

    Returns
    -------
    filename : TYPE
        DESCRIPTION.

    """ 
    filename = Path(filename)
    path = filename.parents[0]
    if create_directory:    
        dealWithDirectory(path, True)

    """ si no tiene la extensión, se la pongo """
    ff = str(filename).lower()
    res = ff.find(extension)
    if res == -1: # no la tiene
        filename = Path(str(filename) + extension)  
        
    return filename
    pass

# @ar_decorator_functions.catch_exceptions
# def readCSV_ORIGINAL(filename, tipo = "pandas", delimiter = ",", encoding = None):
#     """
#     lee un CSV que llega con el path completo y devuelve un df de pandas

#     Parameters
#     ----------
#     filename : TYPE str
#         DESCRIPTION. path completo con nombre de fichero

#     tipo : TYPE str
#         DESCRIPTION. posibles valores "pandas", "dict", ... devuelve la información en el formato que se le pide

#     delimiter : TYPE str
#         DESCRIPTION. delimitador a utilizar, "," por defecto
        
#     Returns
#     -------
#     TYPE si tipo == "pandas" pandas dataframe o False si no se ha podido leer, si tipo == "dict", entonces dictionary 
#         DESCRIPTION.

#     """
#     if not tipo in tipos:
#         raise Exception(f"el tipo {tipo} no se reconoce - posibles tipos {tipos}")
#         return False
    
#     if tipo == "pandas":
#         # try:
#         if not encoding == None:
#             return pd.read_csv(filename, sep = delimiter, low_memory = False, encoding = encoding)
#         else:    
#             return pd.read_csv(filename, sep = delimiter, low_memory = False)
#         # except:
#         #     cadena = f"el fichero {filename} no se ha podido leer"
#         #     _flush_error(cadena)             
#         #     raise Exception(cadena)
#         #     return False

#     if tipo == "csv":
#         # try:
#         with open(filename) as csv_file:
#             csv_reader = csv.reader(csv_file, delimiter=delimiter)

#         return csv_reader
#         # except:
#         #     cadena = f"el fichero {filename} no se ha podido leer"
#         #     _flush_error(cadena)             
#         #     raise Exception(cadena)
#         #     return False



#     if tipo == "raw":
#         # try:
#         # lowlevel_file_opperation(infilename, FO_WRITE, 'wb', encrypted)            
#         with open(filename) as csv_file:
#             csv_reader = csv.reader(csv_file, delimiter=delimiter)
#         return csv_reader
#         # except:
#         #     cadena = f"el fichero {filename} no se ha podido leer"
#         #     _flush_error(cadena)             
#         #     raise Exception(cadena)
#         #     return False

        
#     if tipo == "dict":
#         # try:
#         with open(filename, mode='r') as csv_file:
#             csv_reader = csv.DictReader(csv_file)
#         return csv_reader
#         # except:
#         #     cadena = f"el fichero {filename} no se ha podido leer"
#         #     _flush_error(cadena)             
#         #     raise Exception(cadena)
#         #     return False
        
#     pass



# FUNCIONES DE SOPORTE A API_ROBOT PARA TRABAJAR CON FICHEROS CSV (PARSER)
def encrypt_csv_to_file(df, filename, key):
    """
    Encrypts a pandas DataFrame and saves it to a file.

    :param df: pandas DataFrame to be encrypted.
    :param filename: Name of the file to save the encrypted content.
    :param key: Encryption key used by Fernet.
    :return: None
    Generated by chatGPT, owner Seachad - (FGV) on 22-09-2023 - 12:00:01 (CET+1)
    """
    cipher = Fernet(key)
    csv_content = df.to_csv(index=False).encode()
    encrypted_content = cipher.encrypt(csv_content)
    
    with open(filename, 'wb') as file:
        file.write(encrypted_content)

def decrypt_file_to_df(filename, key):
    """
    Decrypts an encrypted file and returns its content as a pandas DataFrame.

    :param filename: Name of the encrypted file.
    :param key: Encryption key used by Fernet.
    :return: pandas DataFrame containing the decrypted content.
    Generated by chatGPT, owner Seachad - (FGV) on 22-09-2023 - 12:05:01 (CET+1)
    """
    cipher = Fernet(key)
    
    with open(filename, 'rb') as file:
        encrypted_content = file.read()
    
    decrypted_content = cipher.decrypt(encrypted_content).decode()
    return pd.read_csv(pd.StringIO(decrypted_content))


# @ar_decorator_functions.catch_exceptions
def readCSV(filename, tipo = "pandas", delimiter = ",", encoding = "utf-8", decimals_delimiter = "."):
    """
    lee un CSV que llega con el path completo y devuelve un df de pandas

    Parameters
    ----------
    filename : TYPE str
        DESCRIPTION. path completo con nombre de fichero

    tipo : TYPE str
        DESCRIPTION. posibles valores "pandas", "dict", ... devuelve la información en el formato que se le pide

    delimiter : TYPE str
        DESCRIPTION. delimitador a utilizar, "," por defecto
        
    Returns
    -------
    TYPE si tipo == "pandas" pandas dataframe o False si no se ha podido leer, si tipo == "dict", entonces dictionary 
        DESCRIPTION.

    """

    # si el filename no existe returns False
    if not os.path.exists(filename):
        _print(f"el fichero {filename} no se ha podido leer", R)
        return False


    raws = []
    
    if not tipo in tipos:
        raise Exception(f"el tipo {tipo} no se reconoce - posibles tipos {tipos}")
        return False
    
    if tipo == "pandas":
        try:
            if not encoding == None:
                return pd.read_csv(filename, sep = delimiter, low_memory = False, encoding = encoding, decimal = decimals_delimiter)
            else:    
                return pd.read_csv(filename, sep = delimiter, low_memory = False, decimal = decimals_delimiter)
        except:
            cadena = f"el fichero {filename} no se ha podido leer"
            # _flush_error(cadena)             
            raise Exception(cadena)
            return False

    
    if tipo == "csv" or tipo == "list":
        # try:
        
        with open(filename, encoding = "utf-8") as csv_file:
            csv_reader = csv.reader(csv_file, delimiter=delimiter)
            
            line_count = 0
            for row in csv_reader:
                raws.append(row)
                if line_count == 0:
                    # _print(f'Column names are {", ".join(row)}', Y)
                    line_count += 1
                else:
                    if not line_count%100:
                        _print(f'\t{line_count} read from file {filename}.', Y)
                    line_count += 1
            
            
        return raws
        # except:
        #     cadena = f"el fichero {filename} no se ha podido leer"
        #     _flush_error(cadena)             
        #     raise Exception(cadena)
        #     return False

    if tipo == "raw":
        # try:
        # lowlevel_file_opperation(infilename, FO_WRITE, 'wb', encrypted)            
        with open("filename", "rb") as f:
            bytes_read = f.read()
        return bytes_read 
        # except:
        #     cadena = f"el fichero {filename} no se ha podido leer"
        #     _flush_error(cadena)             
        #     raise Exception(cadena)
        #     return False

        
    if tipo == "dict":
        # try:
        dicts = []
        with open(filename, mode='r', encoding = encoding) as csv_file:
            reader = csv.DictReader(csv_file)
            for row in reader:
                dicts.append(row)
            
        return dicts
        # except:
        #     cadena = f"el fichero {filename} no se ha podido leer"
        #     _flush_error(cadena)             
        #     raise Exception(cadena)
        #     return False
        
    if tipo == "json":
        # try:
        return readJSON(filename)    
        # except:
        #     cadena = f"el fichero {filename} no se ha podido leer"
        #     _flush_error(cadena)             
        #     raise Exception(cadena)
        #     return False        
        
        
    pass


def writeCSV(cadena, filename, tipo = "raw", delimiter = ",", mode = "w" ):
    """
    escribe un csv 

    recibe una lista de cadenas con sus elementos separados por "delimiter"

    Parameters
    ----------
    splitted_lines : TYPE list of strings
        DESCRIPTION. registros a escribir
    filename : TYPE str
        DESCRIPTION. path completo con el nombre del fichero a escribir

    Returns
    -------
    None.

    """
    
    filename = deal_with_file(filename, extension = ".csv")
    
    if not tipo in tipos:
        raise Exception(f"el tipo {tipo} no se reconoce - posibles tipos {tipos}")
        return False    
    
    if tipo == "raw" or tipo == "csv":
 
        with open(filename, mode=mode, newline='', encoding='utf-8') as csv_file:    
            if isinstance(cadena, list):
                el = cadena[0]
                if isinstance(el, list): # si contiene listas (son registros)
                    csv_writer = csv.writer(csv_file, delimiter=delimiter, quotechar='"', quoting=csv.QUOTE_MINIMAL)    
                    
                if isinstance(el, dict): # si contiene diccionarios
                    fieldnames = el.keys() # considero que las claves del primer diccionario son los headers
                    csv_writer = csv.DictWriter(csv_file, fieldnames=fieldnames)            

                for r in cadena:
                    csv_writer.writerow(r)
                    
            if isinstance(cadena, str): # si viene todo en una cadena (puede ser que se envíe así como resultado de la respuesta del API)
                # data = "".join(cadena)
                cadena = cadena.replace("\r\n", "\n")
                splitted_lines = cadena.splitlines()   
                reader = csv.reader(splitted_lines)
                parsed_csv = list(reader)
                csv_writer = csv.writer(csv_file, delimiter=delimiter)                   
                for l in parsed_csv:
                    csv_writer.writerow(l)   
                    
    elif tipo == "pandas":
        if isinstance(cadena, str): # si viene todo en una cadena (puede ser que se envíe así como resultado de la respuesta del API)
            with open(filename, mode=mode, newline='', encoding='utf-8') as csv_file:    
                # data = "".join(cadena)
                cadena = cadena.replace("\r\n", "\n")
                splitted_lines = cadena.splitlines()   
                reader = csv.reader(splitted_lines)
                parsed_csv = list(reader)
                csv_writer = csv.writer(csv_file, delimiter=delimiter)                   
                for l in parsed_csv:
                    csv_writer.writerow(l)    
                
        if isinstance(cadena, pd.DataFrame):
            cadena.to_csv(filename, index=False, sep=delimiter) 
                
    else:
        cadena = f"el fichero {filename} no se ha podido escribir. el tipo {tipo} no está implementado"
        raise Exception(cadena)        
        return False 


# ---- .
# ---- * JSONS -----------------------------------------------------------------------------------------------

REPORT_PRE_PROCESS_ = 1
REPORT_PROCESS_ = 2


# =============================================================================
# 20200803
# Devuelve True si el JSON es sintácticamente válido, False en otro caso
# =============================================================================
# @ar_decorator_functions.catch_exceptions
def validateJSON(jsonData):
    import json

    # sólo puedo recibir lista, diccionario o string
    if not isinstance(jsonData, (list, dict, str)):
        return False
    
#    try:
#        json.loads(jsonData)
#    except ValueError as err:
#        _print(err, R)
#        return False
    try:
        json.loads(jsonData)
    except:
        cadena = f"el json {jsonData} tiene errores!"  
        _flush_error(cadena)
        raise Exception(cadena)
        return False
    
    return True


# =============================================================================
# Actualiza un JSON en disco, pasándole una lista con claves y valores
# =============================================================================
execution_report_json = [
    {
     "REPORT_PROCESS_SCRIPT" : "IPL",
     "REPORT_PROCESS_API" : "API",
     "REPORT_PROCESS_TENANT_ID" : "",
     "REPORT_PROCESS_TIME" : get_str_timestamp(),
     "REPORT_RESULT" : 200
    }
    ]    


# =============================================================================
# Lee un JSON desde disco y devuelve un formato JSON (debe llevar la extensión .json al final)
# =============================================================================
# @ar_decorator_functions.catch_exceptions
def readJSON(filename = None, 
             _resolve_patterns = False, 
             _resolve_dates = False, 
             _introspect = True,
             _additional_dictionaries_to_look_in = None,
             binary = False,
             **kwargs):
    """
    (c) Seachad
    
    Description:
    ---------------------- 
    readJSON
    
    reads the json filename (with path included)
    
    Extended Description:
    ---------------------
    _extended_summary_

    Example:
    -------
    
    >>> resultado = readJSON.(d="C:/dict.json")
    
    this will return a usual json. Maybe:
    
    >>> resultado
    >>> [
    >>> "key" : "value",
    >>> "date" : "@N@",
    >>> "pattern_to_resolve" : "{key} and {date}"
    >>> ]    
        
    
    >>> resultado = readJSON.(d="C:/dict.json", resolve_dictionary_patterns = True)
    
    this will return a "resolved" dictionary
    
    >>> resultado
    >>> [
    >>> "key" : "value",
    >>> "date" : "@N@",
    >>> "pattern_to_resolve" : "value and 29/01/2021"
    >>> ]    
    
    Args:
    -----
        - filename (json path, optional): _description_. Defaults to None.
        - _resolve_patterns (if true it solves any pattern in the dictionary found in the form, optional): _description_. Defaults to False.
        - _resolve_dates (_type_, optional): _description_. Defaults to False.
        - _introspect (if True will look INSIDE itself searching for variables defined inside filename.json, optional): _description_. Defaults to True.
        - _additional_dictionaries_to_look_in (if additional dictionary o list of dictionaries added, will be included in the list to resolve patterns, optional): _description_. Defaults to None.

    Returns
    -------
    None si no ha podido leerlo, en otro caso devuelve el JSON.

    """
    _recursive = _introspect

    if binary == False:

        """ si no tiene la extensión, se la pongo """
        ff = str(filename).lower()
        res = ff.find(".json")
        if res == -1: # no la tiene
            filename = Path(str(filename) + ".json")

        # veo si me lo valida como un JSON
        l_JSON = False
        # with open(filename) as json_file:
        #     l_JSON = json.load(json_file)
        #     json_file.close()    
            
        try:
            data = readRawFile(filename, "r")
        except:
            return None
        
        if data == None:
            return []
        
        # limpio todo lo que aparezca antes del primer "[" porque TIENE que ser un JSON
        data = clean_unkind_JSON_data(data)

        l_JSON = json.loads(data)    
        
        # resolve patterns {}
        if _resolve_patterns:
            # creo una copia del diccionario original en el diccionario a buscar
            if _recursive:
                to_search_l_JSON = copy.deepcopy(l_JSON) 
            else:
                to_search_l_JSON = []
            # si tengo diccionarios en los que mirar los añado a la lista de diccionarios en los que buscar
            if not _additional_dictionaries_to_look_in == None:
                to_search_l_JSON.append(_additional_dictionaries_to_look_in)
            # como es un JSON puede tener varios diccionarios en la lista, con lo que busco, uno a uno, en todos ellos
            resulting_JSON = []
            for d in l_JSON:
                resolved_diccionary = rfs.resolve_dictionary_patterns(d = d, 
                                        dictionary_to_search_on = to_search_l_JSON, **kwargs)
                resulting_JSON.append(resolved_diccionary)
            l_JSON = resulting_JSON
        

        # resolve dates 
        # language "@:xxxx@"
        if _resolve_dates:
            l_JSON = operate_dates_on_diccionary(d = l_JSON)
            
        return l_JSON
    
    if binary == True:
        try:
            import msgpack
            # Read from a file
            filename = str(filename).replace(".json", "")
            with open(f'{filename}.msgpack', 'rb') as file:
                binary_data = file.read()    
                
            # Deserialize data from MessagePack
            l_JSON = msgpack.unpackb(binary_data, raw=False)                
            
            # resolve patterns {}
            if _resolve_patterns:
                # creo una copia del diccionario original en el diccionario a buscar
                if _recursive:
                    to_search_l_JSON = copy.deepcopy(l_JSON) 
                else:
                    to_search_l_JSON = []
                # si tengo diccionarios en los que mirar los añado a la lista de diccionarios en los que buscar
                if not _additional_dictionaries_to_look_in == None:
                    to_search_l_JSON.append(_additional_dictionaries_to_look_in)
                # como es un JSON puede tener varios diccionarios en la lista, con lo que busco, uno a uno, en todos ellos
                resulting_JSON = []
                for d in l_JSON:
                    resolved_diccionary = rfs.resolve_dictionary_patterns(d = d, 
                                            dictionary_to_search_on = to_search_l_JSON, **kwargs)
                    resulting_JSON.append(resolved_diccionary)
                l_JSON = resulting_JSON
            

            # resolve dates 
            # language "@:xxxx@"
            if _resolve_dates:
                l_JSON = operate_dates_on_diccionary(d = l_JSON)
                
            return l_JSON
                
                        
        except Exception as e:
            tc._print(f"Error: {e} BINARY!!!", tc.R)
            return None
        
    

def test_readJSON():
    filename = "D:/OneDrive - Seachad/07 - DESARROLLOS INTERNOS/1000 - MAQUINA_DESARROLLO/API_ROBOT_EXECUTION/Process Production.json"
    list_of_changes = {}
    resultado = readJSON(filename, _resolve_patterns = True, 
                         _resolve_dates = True, 
                         _additional_dictionaries_to_look_in = [access], list_of_changes = list_of_changes)    
    
    a = 1
    pass

# 20210506.001
def clean_unkind_JSON_data(data):
    ldata = data
    try:
        ldata = data
        substring = "["
        index = data.find(substring)
        ldata = data[index:]
    except Exception as e:
        a = 999
    
    return ldata


# =============================================================================
# Escribe en disco el registro de proceso (deprecated)
# =============================================================================
def updateJSON(jsonFilename = "", jsonRecord = ""):
    """
    DEPRECATED - antigua función de registro de proceso, ya no se utiliza
    Actualiza un JSON buscando por su clave principal
    Si encuentra la clave principal, actualiza su contenido
    En otro caso crea la clave principal

    Parameters
    ----------
    jsonFilename : TYPE, optional
        DESCRIPTION. The default is "".
    jsonUpdate : TYPE, optional
        DESCRIPTION. The default is "".
    script_key : TYPE, optional
        DESCRIPTION. The default is "".
    api_key    

    Returns
    -------
    None.

    """
    lista = []
    updatable = False

    # leemos el JSON
    try:
        
        # json_file =  open(jsonFilename, mode = "r")
        # lista = json.load(json_file)        
        # json_file.close()
        
        lista = readJSON(jsonFilename, "r")
        
        updatable = True
    except:
        writeReadableJSON(execution_report_json, jsonFilename)
        return
        
    if jsonRecord == "":
        return lista
    
    found = False    
    if updatable == True:    
        # procesamos el JSON con claves y valores a actualizar    
        for i in lista:

            if jsonRecord["REPORT_PROCESS_SCRIPT"] == i["REPORT_PROCESS_SCRIPT"] and jsonRecord["REPORT_PROCESS_API"] == i["REPORT_PROCESS_API"] and jsonRecord["REPORT_PROCESS_TENANT_ID"] == i["REPORT_PROCESS_TENANT_ID"]: # existe la clave?
                found = True
                for k,v in jsonRecord.items():
                    i[k] = v
               
    if found == False:
        lista.append(jsonRecord)         
        
    writeReadableJSON(lista, Path(jsonFilename))   
      
    return lista    
    


def appendJSON(filename, record):
    """
    escribe un nuevo registro en un json previo existente, si el archivo no existe, lo crea

    Parameters
    ----------
    filename : TYPE
        DESCRIPTION.
    record : TYPE
        DESCRIPTION.

    Returns
    -------
    None.

    """
    # chequeo si el fichero existe y chequeo la extensión
    filename = deal_with_file(filename, extension = ".json")
    # leo el json
    json_read = readJSON(filename)
    # añado el registro
    json_read.append(record)
    # actualizo el json resultante
    writeReadableJSON(json_read, filename)

TYPE_CLASS  = 1
TYPE_FUNCTION = 2
TYPE_NATIVE = 3
def say_type(obj):
    # Check if obj is a class
    type_obj_str = str(type(obj))
    if type_obj_str == "<class 'function'>":
        return TYPE_FUNCTION
    else:
        # if starts with "<class" and ends with ">"
        if type_obj_str.startswith("<class") and type_obj_str.endswith(">"):
                return TYPE_CLASS
    # Otherwise
    return TYPE_NATIVE

class CustomEncoder(json.JSONEncoder):
    # !!! activar si queremos ver qué se está procesando. Útil para problemas en alguna clave
    # def encode(self, obj):
    #     # Si el objeto es un diccionario, imprime cada clave
    #     tc._print(f"CustomEncoder:encode {obj}", tc.Y)
    #     if isinstance(obj, dict):
    #         for key in obj:
    #             tc._print(f"CustomEncoder:encode -> Procesando la clave: {key} - value {obj[key]}", tc.Y)
    #     # Llama al método encode del padre para continuar con la serialización
    #     return super().encode(obj)
        
    def default(self, obj):
        # Intenta serializar el objeto con el método predeterminado
        # Otherwise use default encoding
        try:        
            type_obj = say_type(obj)
            if type_obj == TYPE_CLASS:
                return f"obj->{str(obj.__class__.__name__)}"
            if type_obj == TYPE_FUNCTION:
                return f"func->{str(obj.__name__)}"

            result = super().default(self, obj)  
            return result
        except Exception as e:
            tc._print(f"CustomEncoder: ERROR {e} in obj {obj}", tc.R)

def writeReadableJSON( dictionary, 
                      filename,
                      binary = False 
                      ):
    """
    escribe un json en formate "readable"
    en el filename que contiene el path completo, crea la estructura de directorios si esta no existe    
    chequea si tiene la extensión json, y si no se la pone

    incopora funcionalidad de retry, por si el fichero se encuentra bloqueado por alguien

    Parameters
    ----------
    dictionary : TYPE dict
        DESCRIPTION.Diccionario a escribir
    filename : TYPE str
        DESCRIPTION. Nombre del fichero de salida

    Returns
    -------
    None.

    """

    filename = deal_with_file(filename, extension = ".json")
    cadena = {}
    def default(obj):
        # solver problems of unknonw objects (like Request, MetaData, etc...)
        return f"unknown object"
    if binary == False:
        try:
            cadena = json.dumps(dictionary, indent=4, cls=CustomEncoder)
        except Exception as e:
            tc._print(f"common.writeReadableJSON: {filename} Error {e}", R)
            raise
        writeRawFile(cadena, filename, "w")
    if binary == True:
        try:
            import msgpack
            # Serialize data to MessagePack (binary format)
            # cadena = json.dumps(dictionary, indent=4, cls=CustomEncoder)
            packed_data = msgpack.packb(dictionary, use_bin_type=True)    
            # Save to a file
            filename = str(filename).replace(".json", "")
            with open(f'{filename}.msgpack', 'wb') as file:
                file.write(packed_data)            
        except Exception as e:
            tc._print(f"common.writeReadableJSON: {filename} Error {e} BINARY!", R)
            raise   




def writeJSON(dictionary, filename, binary = False):
    writeJSONFile( dictionary, filename, binary = binary)

def writeJSONFile( dictionary, filename, binary = False):
    writeReadableJSON( dictionary, filename, binary = binary)

# def fileExtension(extensionToLookFor = ".JSON"):
    
#     import re
#     regex_patron = r".JSON"
#     insensible = re.compile(regex_patron, re.I) # modifico para que sea insensible a mayúsculas y minúsculas
#     json_present = insensible.findall( str(filename) )
#     if len(json_present) == 0: # no contiene la cadena .json
#         filename = Path(str(filename) + ".json")




# =============================================================================
# JSON to disk (no easy readable)
# recibe Path de pathlib
# =============================================================================
def jsonToDisk(filename, res, rescode = 200):
    """
    Esta función escribe un JSON a disco, con el nombre de fichero que se le envía. Admite opcionalmente un código de error (devuelto por llamada a API). Si el error es diferente a 200 (todo OK)
    añade al fichero una cadena conteniendo información sobre el error.

    Parameters
    ----------
    filename : TYPE
        DESCRIPTION.
    res : TYPE
        DESCRIPTION.
    rescode : TYPE, optional
        DESCRIPTION. The default is 200.

    Returns
    -------
    None.

    """
    # global lg
    # logger = lg.arguments(currentframe(), verbose = False)

    data = res

    if rescode != 200:
        filename = f"{filename}_ERR[{rescode}]"
    try:
            
        # file = open(str(filename) + '.json', mode = 'w')
        # json.dump(data, file, indent=4)
        # file.close()
        
        data = json.dumps(data)
        writeReadableJSON(data, filename)
            
    except Exception as e:
        raise Exception(f"Error en jsonToDisk {e}")
        
        # lg._Debug(logger, f"Error {e}")
        return -1


# @ar_decorator_functions.catch_exceptions
def file_delete(filename):
    """
    borra un fichero del sistema

    Parameters
    ----------
    filename : TYPE str 
        DESCRIPTION. contiene el nombre y la ruta del fichero a borrar

    Returns
    -------
    None.

    """
    import os
    if os.path.exists(filename):
      os.remove(filename)
      return True
    else:
      # _print(f"delete_file: The file {filename} does not exist")    
      return False

# @ar_decorator_functions.catch_exceptions
def file_copy(filename_from, filename_to):
    """
    copia un fichero a otro directorio

    Parameters
    ----------
    autoexplicativos

    Returns
    -------
    None.

    """
    import os
    if os.path.exists(filename_from):
      os.system(f'copy {filename_from} {filename_to}')
      return True
    else:
      # _print(f"delete_file: The file {filename} does not exist")    
      return False

# =============================================================================
# Construye un registro para actualizar en el report de ejecución (deprecated)
# =============================================================================
def _buildRecordReportJson(
        registroProcesoFilename = "ReportJSON.json",
        script = "not informed",
        api = "not informed",
        tenant_id = "not informed",
        ltime = get_str_timestamp(),
        result = "NOK"
        ):
    
    # copiamos el template
    record_execution_report_json = execution_report_json[0]
    
    # actualizamos datos
    record_execution_report_json["REPORT_PROCESS_SCRIPT"] = script
    record_execution_report_json["REPORT_PROCESS_API"] = api
    record_execution_report_json["REPORT_PROCESS_TENANT_ID"] = tenant_id
    record_execution_report_json["REPORT_PROCESS_TIME"] = ltime
    record_execution_report_json["REPORT_RESULT"] = result
        
    
    return updateJSON(jsonFilename = registroProcesoFilename, jsonRecord = record_execution_report_json)
# ---- .
# ---- * LOW LEVEL FILES -----------------------------------------------------------------------------------------------
# ---- .

# =============================================================================
# Necesario para escribir ficheros encriptados
# =============================================================================
# @ar_decorator_functions.catch_exceptions
@retry(wait_random_min=1000, wait_random_max=2000, stop_max_attempt_number=10)
def writeRawFile(data, filename, mode = "wb"):
    # global lg
    # logger = lg.arguments(currentframe())

    # _print("Retrying...", Y)

    filename = Path(filename)
    path = filename.parents[0]
    dealWithDirectory(path, True)

    cadena = data

    # cooked filename (si no lleva .JSON se lo ponemos al final)
    # import re
    # regex_patron = r".JSON"
    # insensible = re.compile(regex_patron, re.I) # modifico para que sea insensible a mayúsculas y minúsculas
    # json_present = insensible.findall( str(filename) )
    # if len(json_present) == 0: # no contiene la cadena .json
    #     filename = Path(str(filename) + ".json")

    if mode != "wb":
        f=open(filename,mode, encoding='utf-8') # fichero de configuración de llamadas al API
    else:
        f=open(filename,mode)
    f.write(cadena)
    f.close()    

# =============================================================================
# Necesario para escribir ficheros encriptados
# =============================================================================
# @ar_decorator_functions.catch_exceptions
@retry(wait_random_min=1000, wait_random_max=2000, stop_max_attempt_number=3)
def readRawFile(filename, mode = "rb"):
    # global lg
    # logger = lg.arguments(currentframe())

    cadena = None
    filename = Path(filename)
    # path = filename.parents[0]
    # dealWithDirectory(path, True)

    # cadena = data

    # cooked filename (si no lleva .JSON se lo ponemos al final)
    # import re
    # regex_patron = r".JSON"
    # insensible = re.compile(regex_patron, re.I) # modifico para que sea insensible a mayúsculas y minúsculas
    # json_present = insensible.findall( str(filename) )
    # if len(json_present) == 0: # no contiene la cadena .json
    #     filename = Path(str(filename) + ".json")
    try:
        f = open(str(filename), mode) # fichero de configuración de llamadas al API
        cadena = f.read()
        f.close()
    except:
        tc._print(f"No se ha podido leer el fichero {str(filename)}", tc.R)

    
    return cadena

# @ar_decorator_functions.catch_exceptions
@retry(wait_random_min=1000, wait_random_max=2000, stop_max_attempt_number=10)
def appendRawFile(filename, record):
    """
    escribe un nuevo registro en un raw file previo existente, si el archivo no existe, lo crea

    Parameters
    ----------
    filename : TYPE
        DESCRIPTION.
    record : TYPE
        DESCRIPTION.

    Returns
    -------
    None.

    """
    # chequeo si el fichero existe y chequeo la extensión
    filename = deal_with_file(filename, extension = "")
    # leo el json
    # data0 = readRawFile(filename)
    # # añado el registro
    # data = data0 + record
    # actualizo el json resultante
    writeRawFile(record, filename, mode = "a")

# ---- .
# ---- * MONGODB low level --------------------------------------------------------------




    
# ---- .
# ---- * TRABAJAR CON DIRECTORIOS -----------------------------------------------------------------------------------------------

# Chequea si un path existe. Lo crea si se le pide, con create = True

def dealWithDirectory(filename_path, create = False):


    """
    Chequea si un directorio existe y lo crear si la variable create está a True

    Parameters
    ----------
    filename_path : TYPE
        DESCRIPTION.
    create : TYPE, optional
        DESCRIPTION. The default is False.

    Returns
    -------
    None.

    """

    exists = False
    if Path(filename_path).exists():
        exists = True
    else:
        if create == True:
            Path(filename_path).mkdir(parents=True, exist_ok=True)
    return exists

# =============================================================================
# Devuelve el path y el fichero por separado
# =============================================================================
def file_get_components(filename):
    # filename = '912c2842-05da-43f3-9441-4e0f4846fbc7/otrosub/CSPFacturacion.xlsx'
    fileoutput = Path(filename)
    path = fileoutput.parents[0] # obtengo el path
    filename = fileoutput.parts[-1] # obtengo el nombre del fichero
    return path, filename

# @ar_decorator_functions.catch_exceptions
def import_module(module_name):
    """
    (c) Seachad
    
    Description:
    ---------------------- 
    import_module
    
    importa un módulo con su path completo
    el módulo funciona en python como si fuese un módulo cargado por la forma normal

    Ejemplo:
    >>> modulo = import_module("C:/Users/Usuario/Documents/Pyt.py")
    >>> res = modulo.funcion("kk)
    
    Extended Description:
    ---------------------
    _extended_summary_
    
    
    Args:
    -----
        - module_name (module name with complete path to load): path y nombre de módulo completo
    
    Returns:
    --------
        _type_: _description_
    """    
    import importlib
    me = None
    module_text = module_name
    path_to_extension, module_exension = file_get_components(module_text)
    # añado, como directorios de búsqueda, todos los hijos que tenga (por si hay librerías o packages, al modo Python...)
    if not path_to_extension in sys.path:
        sys.path.append(str(path_to_extension))  
        children_paths = [x[0] for x in os.walk(path_to_extension)]   
        for cd in children_paths:
            sys.path.append(cd)

    me = importlib.import_module(module_exension)      
       
    return me
    


# ---- * Garbage collection ------------------------------------------------------
def namestr(obj, namespace):
    return [name for name in namespace if namespace[name] is obj]

def control_references(object_of_interest):
    """ muestra los referrers a un objeto """
    import gc
    gc.collect() #make sure all garbage cleared before collecting referrers.
    referrers = gc.get_referrers(object_of_interest)
    _print( "get_references in gc - GLOBALS ----------", B)    
    namespace =  globals()    
    for referrer in referrers:
        _print( namestr(referrer, namespace), B)
        
    _print( "get_references in gc - LOCALS ----------", B)    
    namespace =  locals()    
    for referrer in referrers:
        _print( namestr(referrer, namespace), B)        

# ---- .
# ---- * SYSTEM ----------------------------------------------------------------

# trabajar con los contadores de referencias de Python (como tengo diccionarios con punteros a clases, siempre está la referencia interna aumentada en uno
# y las instancias que se inscriben en el diccionario nunca reciben la llamada a __del__() porque el didcionario sigue manteniendo un puntero a esa copia )

""" incrementar, reducir y obtener contador de referencias de python """

import ctypes
from ctypes import wintypes

def disable_quick_edit_mode():
    """
    deshabilita la propiedad de QuickEditMode en una consola de Windows, que hace que
    al tocar en la ventana con el ratón, ésta pare la ejecución

    Returns
    -------
    last_error : TYPE
        DESCRIPTION.

    """
    kernel32 = ctypes.WinDLL('kernel32')
    dword_for_std_input_handle = ctypes.wintypes.DWORD(-10)
    dword_for_enable_extended_flags = ctypes.wintypes.DWORD(0x0080)
    std_input_handle = kernel32.GetStdHandle(dword_for_std_input_handle)
    kernel32.SetConsoleMode(std_input_handle, dword_for_enable_extended_flags)
    last_error = kernel32.GetLastError()
    return last_error

def test_disable_quick_edit_mode():
    disable_quick_edit_mode()



_decref = ctypes.pythonapi.Py_DecRef
_decref.argtypes = [ctypes.py_object]
_decref.restype = None

_incref = ctypes.pythonapi.Py_IncRef
_incref.argtypes = [ctypes.py_object]
_incref.restype = None

def _countref(o):
    
    import ctypes
    # my_var = 'hello python'
    # my_var_address = my_var
    
    l = ctypes.c_long.from_address(id(o)).value-2 # le quito 2 porque se crean dos referencias temporales al pasarlo a la función  
    
    return l    


""" recoge variables de entorno preguntando al sistema (os.environ sólo accede a las que están vivas en la sesión) """   
import winreg
def get_user_env(name):
    key = winreg.CreateKey(winreg.HKEY_CURRENT_USER, r"Environment")
    retorno = ""
    try:
        retorno = winreg.QueryValueEx(key, name)[0]
    except:
        retorno = None
    return retorno

def set_user_env(name, value):
    value = str(value)
    key = name
    execute_string = f"cmd /c setx {key} \"{value}\""
    retorno = os.system(execute_string) 
    return retorno

import winreg
def get_environment_variables( prefix = None, path = winreg.HKEY_CURRENT_USER):
    """
    devuelve una lista de diccionarios con todas las variables de entorno que cumplen con el patrón
    Si se le envía None, entonces devolverá el contenido de TODAS las variables de entorno

    Parameters
    ----------
    prefix : TYPE, optional
        DESCRIPTION. The default is "CM_extensions_".

    Returns
    -------
    None.

    """
    
    key = winreg.CreateKey(path, r"Environment")
    count = winreg.QueryInfoKey(key)[1]
    list_environment_variables = []
    include = False
    for c in range(count):
        name, value, tt = winreg.EnumValue(key, c)
        include = True
        if not prefix == None:
            if not prefix in name:
                include = False
        if include:
            d = {}
            d["name"] = name
            d["value"] = value
            list_environment_variables.append(d.copy())

    return list_environment_variables



# =============================================================================
# saca un "shot" de error en un archivo
# =============================================================================
def _error_flush(cadena, auto_destroy = False):
    _flush_error(cadena, auto_destroy = auto_destroy)
    
def _flush_error(cadena, auto_destroy = False): 
    """
    "escupe" un fichero en el directorio CM_err_folder o crea una directorio "C:/API_ROBOT_ERR" 

    Parameters
    ----------
    cadena : TYPE
        DESCRIPTION.
    auto_destroy : TYPE, optional
        DESCRIPTION. The default is False. Si está a True, el fichero no persistirá

    Returns
    -------
    None.

    """

    import builtins

    access = builtins.access

    # import IPL
    access = IPL.get_access()

    import ar_managers
    from pathlib import Path
    name = IPL.timestamp()
    filename = f"ERR_{IPL.timestamp()}.txt"
    log_path = access.get("CM_err_folder", "C:/API_ROBOT_ERR")
    filename = Path(log_path) / Path(filename)
    m = ar_managers._persistence_manager_raw(_name = name, _filename = str(filename), _auto_destroy = auto_destroy)   

    m._put(cadena)
    _print(cadena)
    # quitar...
    # _input("ERROR???")    


# =============================================================================
# Recibe un string y mira a ver si hay algo entre {} -> chequea si hay una variable de entorno con ese nombre y sustituye
# =============================================================================
def look_for_environment_variable(cadena):
    
    import re

    lista_de_elementos_entre_llaves = []
    
    """ localizar todo lo que esté entre {} """
    # detect delimiters of date commands
    start_delimiter = "{"
    end_delimiter = "}"

    # is it a date language specific command?
    # it needs to start with @D:
    patron_comando = r"{start_delimiter}[\w\d+-:]+{end_delimiter}".format(
        start_delimiter = start_delimiter,
        end_delimiter = end_delimiter) # buscamos secuencias de palabras dígitos, signos + y menos y : entre @

    regex_patron = patron_comando
    comandos = re.findall(regex_patron, cadena)
    
    if len(comandos)==0: # no es un comando
        return cadena    
    
    """ una vez obtenidos los comandos los busco
    1) en access
    2) en variables de entorno
    """
    lista_cambios = {}
    
    for e in comandos:
        mem_e = e
        # le quitamos las llaves
        e = e.replace("{", "")
        e = e.replace("}", "")
        valor, params = _resolve_dynamic_fstring(
                     clave = e, # clave que quiero resolver (si la clave es un diccionario, resolverá todos sus elementos y devolverá un diccionario)
                     # context = None, # contexto de queries (contiene toda la información lógica y el acceso a access con la información ya descargada de ENV y de ENCRYPT data)
                     list_of_dictionaries = access, # orden de diccionarios que trataremos para resolver la clave, desde el primero al último (se pasa por clave y la función los resuelve)
                     force_until_all_resolved = True, # por defecto dejo que se devuelvan parámetros sin resolver (cuidado con este flag, si queda algo sin resolver contínuamente, puede dar un reventado)
                     )
        
        if valor == None: # no lo ha encontrado, buscamos en las variables de entorno
            valor = get_user_env(e)
    
        if valor == None:
            valor = mem_e
    
        lista_cambios[mem_e] = valor
        _print(f"{valor}")
    # ahora cambiamos los valores y devolvemos el resultado
    for k,v in lista_cambios.items():
        cadena = cadena.replace(k,v)
    
    return cadena


# ---- * SORT listas de diccionarios por múltiples claves  --------------------------------------------------------------------------------------------------------------------------------------------------------


from operator import itemgetter as i
from functools import cmp_to_key

def cmp(x, y):
    """
    Replacement for built-in function cmp that was removed in Python 3

    Compare the two objects x and y and return an integer according to
    the outcome. The return value is negative if x < y, zero if x == y
    and strictly positive if x > y.

    https://portingguide.readthedocs.io/en/latest/comparisons.html#the-cmp-function
    """

    return (x > y) - (x < y)

def multikeysort(items, columns):
    comparers = [
        ((i(col[1:].strip()), -1) if col.startswith('-') else (i(col.strip()), 1))
        for col in columns
    ]
    def comparer(left, right):
        comparer_iter = (
            cmp(fn(left), fn(right)) * mult
            for fn, mult in comparers
        )
        return next((result for result in comparer_iter if result), 0)
    return sorted(items, key=cmp_to_key(comparer))


# import api_robot_integration as ar_int

# ---- .
# ---- * LENGUAJE DE FECHAS 

from datetime import datetime
# =============================================================================
# Deja una fecha en formato educado
# =============================================================================
def clean_datetime_string(datetime_string):
    # limpiamos espacios

    datetime_string = datetime_string.rstrip()
    datetime_string = datetime_string.lstrip()
    
    # quitamos T y Z
    datetime_string = datetime_string.replace("T", " ")
    datetime_string = datetime_string.replace("Z", "")
    
    # solo admitimos 6 posiciones tras el punto decimal
    punto_position = datetime_string.find(".")
    total_decimales = 6
    if not punto_position == -1: 
        decimals =  datetime_string[punto_position+1:][:6]

        left_datetime_string_punto_position = datetime_string[:punto_position+1]
        str_decimals = f"{int(decimals):<06d}"
        datetime_string = left_datetime_string_punto_position + str_decimals
        punto_position = len(datetime_string)
        total_decimales = 0
        # caracteres_a_coger = len(left_datetime_string_punto_position) + 
        
    # caracteres_a_coger = punto_position + total_decimales
    # datetime_string = datetime_string[:caracteres_a_coger+1]
    return datetime_string

# =============================================================================
# Construye los patterns para regex
# recibe datetime_string para saber si ha de añadir el .f% o no
# =============================================================================
def build_datetime_patterns(datetime_string):   
    
    # si tiene . es que viene con float en segundos
    agregar_float = False
    if datetime_string.find(".") > 0:
        agregar_float = True
        

    # nuevo detector de formatos (parsing)
    # REGEX PARSERS
    # _start = "^"
    # _end = "$"
    # _d = "([1-9]|0[1-9]|1[0-9]|2[0-9]|3[0-1])"
    # _m = "([1-9]|0[1-9]|1[0-2])"
    # _y = "\d{2}"
    # _Y = "\d{4}"
    # _dsep = "(\.|-|/)"
    # _H = "\d{2}"
    # _M = "\d{2}"
    # _S = "\d{2}"     
    # _hsep = ":"
    
    # DATE FORMAT
    _start = "^"
    _sstart = ""
    _end = "$"
    _send = ""
    if agregar_float:
        _end = ".\d{6}$"
        _send = ".%f"
        
    _d = "([1-9]|0[1-9]|1[0-9]|2[0-9]|3[0-1])"
    _sd = "%d"
    _m = "([1-9]|0[1-9]|1[0-2])"
    _sm = "%m"
    _y = "\d{2}"
    _sy = "%y"
    _Y = "\d{4}"
    _sY = "%Y"
    _dsep = "(\.|-|/)"
    
    _H = "\d{2}"
    _sH = "%H"
    _M = "\d{2}"
    _sM = "%M"
    _S = "\d{2}"     
    _sS = "%S"
    _hsep = ":"  

    _sdsep = "@"
    _shsep = "*"  

    # formato intermedio de dm y md
    dm = _d + _dsep + _m
    md = _m + _dsep + _d
    sdm = _sd + _sdsep + _sm
    smd = _sm + _sdsep + _sd

    # formato intermedio de hms
    hms = " " + _H + _hsep + _M + _hsep + _S 
    shms = " " + _sH + _shsep + _sM + _shsep + _sS 

    # parsers de sólo fecha
    ydm = _start + _y + _dsep + dm
    ymd = _start + _y + _dsep + md
    Ydm = _start + _Y + _dsep + dm
    Ymd = _start + _Y + _dsep + md

    sydm = _sstart + _sy + _sdsep + sdm
    symd = _sstart + _sy + _sdsep + smd
    sYdm = _sstart + _sY + _sdsep + sdm
    sYmd = _sstart + _sY + _sdsep + smd
    
    dmy = _start + dm + _dsep + _y
    mdy = _start + md + _dsep + _y
    dmY = _start + dm + _dsep + _Y
    mdY = _start + md + _dsep + _Y   
    
    sdmy = _sstart + sdm + _sdsep + _sy
    smdy = _sstart + smd + _sdsep + _sy
    sdmY = _sstart + sdm + _sdsep + _sY
    smdY = _sstart + smd + _sdsep + _sY   


    # parsers que incluyen HMS
    ydm_hms = ydm + hms
    ymd_hms = ymd + hms
    Ydm_hms = Ydm + hms
    Ymd_hms = Ymd + hms
    
    sydm_hms = sydm + shms
    symd_hms = symd + shms
    sYdm_hms = sYdm + shms
    sYmd_hms = sYmd + shms
    
    dmy_hms = dmy + hms
    mdy_hms = mdy + hms
    dmY_hms = dmY + hms
    mdY_hms = mdY + hms
    
    sdmy_hms = sdmy + shms
    smdy_hms = smdy + shms
    sdmY_hms = sdmY + shms
    smdY_hms = smdY + shms
  
    
    date_patterns = [
            (ydm + _end, sydm + _send),
            (ymd + _end, symd + _send),
            (Ydm + _end, sYdm + _send),
            (Ymd + _end, sYmd + _send),
            
            (dmy + _end, sdmy + _send),
            (mdy + _end, smdy + _send),
            (dmY + _end, sdmY + _send),
            (mdY + _end, smdY + _send),
            
            # parsers que incluyen HMS
            (ydm_hms + _end, sydm_hms + _send),
            (ymd_hms + _end, symd_hms + _send),
            (Ydm_hms + _end, sYdm_hms + _send),
            (Ymd_hms + _end, sYmd_hms + _send),
            
            (dmy_hms + _end, sdmy_hms + _send),
            (mdy_hms + _end, smdy_hms + _send),
            (dmY_hms + _end, sdmY_hms + _send),
            (mdY_hms + _end, smdY_hms + _send),
        ]
    
    return date_patterns, _dsep, _sdsep, _hsep, _shsep
  

# =============================================================================
# Convierte, si puede, un string a un objeto datetime    
# =============================================================================
def get_datetime_object( datetime_string ):  

    # recibe una cadena de texto con hipotético formato de fecha
    # y devuelve un objeto tipo datetime, si puede convertirlo
    # o un string si finalmente entiende que no es una fecha

    # Parameters
    # ----------
    # t : TYPE
    #     DESCRIPTION.

    # Returns
    # -------
    # None.


    datetime_string = clean_datetime_string(datetime_string)

    # probado con 
    t1 = "28/07/14"
    t2 = "07-28-2014"
    
    t3 = "07-28-14 10:40:25"
    t4 = "28-07-2014 10:40:25"
    
    t5 = "2020-11-26T05:13:53Z"
    
    t6 = "2014-07-28"
    t7 = "2014-28-07"

    from datetime import datetime
    # datetime.strptime(t1, f2)
    
    # si me viene un dato vacío, lo pongo a la menor fecha posible
    if datetime_string == "":
        datetime_string = "12/02/1582"
    

    patterns, _dsep, _sdsep, _hsep, _shsep = build_datetime_patterns(datetime_string)
    
    import re    
    retorno = datetime_string # si no ha habido conversión posible, devuelvo lo que llegó originalmente
    
    # cambio el separador neutro por el que corresponde
    
    cnt = 0
    for pattern in patterns:
        x=re.search(pattern[0], datetime_string) # formato sólo fecha 
        if not x == None:
            break
        cnt += 1

    # qué separador original de fecha tenía el string que nos han enviado?
    date_separators = ["/", "-", "|"]
    hour_separators = [":"]
    
    datetime_object = datetime_string
    
    if not x == None:
        format_date_time = patterns[cnt][1]


        for original_date_separator in date_separators:
            if original_date_separator in datetime_string:
                break  
        for original_hour_separator in hour_separators:
            if original_hour_separator in datetime_string:
                break                 
        
        format_date_time = format_date_time.replace(_sdsep, original_date_separator)
        format_date_time = format_date_time.replace(_shsep, original_hour_separator)
        
        try:
            datetime_object = datetime.strptime(datetime_string, format_date_time)        
        except:
            _print(f"no se puede aplicar {format_date_time} a {datetime_string}")        
            
        
        retorno = datetime_object         

    return retorno
    
def convert_stringdate_to_datetime(fecha_string):
    """ convierte una fecha en formato string en un objeto datetime.datetime de python """
    fs = clean_datetime_string(fecha_string)
    dt = get_datetime_object(fs)
    return dt  


def test_datetime_formats_probados():
    import apirobot_external_integration
    import terminal_colors as tc
        
    # todas estas fechas están probadas y funcionando...
    probar_fechas = [
        "28/07/14",
        "28-07-14",    
        "07-28-2014",
        "07-28-14 10:40:25",
        "28-07-2014 10:40:25",
        "2020-11-26T05:13:53Z",
        "2014-07-28",
        "2014-28-07",  
        "2019-09-30T21:36:45Z            ",
        "2021-03-14T00:55:55.8016757Z"
        ]
    
    
    for prueba_fecha in probar_fechas:
        
        tc.__trace(f"antes [{prueba_fecha}]", tc.Y)
        tc.__trace(f"antes [{prueba_fecha}]", tc.B)
        prueba_fecha = clean_datetime_string(prueba_fecha)
        tc.__trace(f"despues [{prueba_fecha}]", tc.Y)
        tc.__trace(f"resultado = {type(get_datetime_object(prueba_fecha))}")


    """ PROCESO LGV """
    fecha_string = "2021-01-13T11:06:28Z"
    _print(convert_stringdate_to_datetime(fecha_string))



def operate_dates_on_diccionary(d = None):
    """
    recibe un diccionario o una lista y devuelve las fechas correspondientes a cualquier campo que haya usado la nomenclatura de fechas

    Parameters
    ----------
    d : "dictionary to resolve dates on", optional
        DESCRIPTION. The default is None.

    Raises
    ------
    Exception
        DESCRIPTION.

    Returns
    -------
    d : TYPE (dict)
        DESCRIPTION.

    """
    
    if not isinstance(d, (dict,list)):
        raise Exception(f"operate_dates_on_diccionary:el valor enviado no es un diccionario ni una lista -> {d}")
        return d
    
    # esto es nuevo - 25/09/2021 para soportar una lista de diccionarios (un JSON?)
    # list
    if isinstance(d, list):
        for el in d:
            if isinstance(el, (list,dict)):
                _ = operate_dates_on_diccionary(el) 
            if isinstance(el, str):
                el = dealWithDates(el)     
    # dict
    if isinstance(d, dict):
        for k,v in d.items():
            # si es una lista, itero y envío para cada elemento diccionario
            if isinstance(v,list):
                for el in v:
                    if isinstance(el, dict):
                        _ = operate_dates_on_diccionary(el)                
            # si es un diccionario, lo envío            
            if isinstance(v,dict):
                _ = operate_dates_on_diccionary(v)               
            
            lv = dealWithDates(v)     
            if not lv == False:
                d[k]=lv
            
    return d

# =============================================================================
# Recibe un cadena con el lenguage de fechas y devuelve le fecha concreta ya sustituida
# =============================================================================
patron_year_start = "[0-9]{4}[\:][A-Z\d]+"
patron_constant_date_start = "[0-9]{4}[\-\:]{1}[0-9]{2}[\-\:]{1}[0-9]{2}"
patron_constant_date_start_mod = "\b(\d{2}-\d{2}-\d{4})"
patron_modifier = "[\:A-Z]+"
patron_sign = "[\+\-]{1}"
patron_start = r"[A-Z]+\d|[A-Z]+" # un tag es una secuencia de letras en mayúsculas que puede o no tener adicionalmente un dígito
patron_digits = "[\d]+"

# =============================================================================
# wrapper sencillito del dealWithDates(que devuelve si se ha cambiado o no y la fecha, todo en el mismo momento... :-())
# =============================================================================
def date_language(cadena, date_format = "%Y-%m-%d"):
    """
    devuelve la fecha original o cambiada si contiene lenguaje de fechas

    Parameters
    ----------
    cadena : TYPE
        DESCRIPTION.
    date_format : TYPE, optional
        DESCRIPTION. The default is "%Y-%m-%d".

    Returns
    -------
    timestamp : TYPE
        DESCRIPTION.

    """
    timestamp = cadena
    lres = dealWithDates(cadena, date_format)
    if lres:
        timestamp = lres    
    return timestamp

# =============================================================================
# Formula sencilla de detección de patrones para fechas, paths, ...
# =============================================================================
def detectPattern( start_delimiter = "@D:", end_delimiter = "@", cadena = ""):
    # detect delimiters of date commands
    start_delimiter = start_delimiter
    end_delimiter = end_delimiter

    path_expression = "[a-zA-Z]:\\((?:[a-zA-Z0-9() ]*\\)*).*"

    # is it a date language specific command?
    # it needs to start with @D:
    patron_comando = r"{start_delimiter}[\w\d.\*\\\/+-:]+{end_delimiter}".format(
        start_delimiter = start_delimiter,
        end_delimiter = end_delimiter) # buscamos secuencias de palabras dígitos, signos + y menos y : entre @

    regex_patron = patron_comando
    comandos = re.findall(regex_patron, cadena)
    if len(comandos)==0: # no es un comando
        return False  
    else:
        return comandos

# import re
# cadena = " start command pepito E:/API_ROBOT_temporal\/PROCESS_API_ROBOT_TEST*.*"
# start_delimiter = "@path:"
# end_delimiter = "@"
# found = detectPattern( start_delimiter = "@path:", end_delimiter = "@", cadena = cadena)
# from pathlib import Path
# path_ok = Path(cadena)

DEAL_WITH_DATES_STD_COMMANDS = {
        "NOW": "START", # fecha actual

        "CY" : "START", # Current Year
        "CM" : "START", # Current Month
        "CH" : "START", # Current Half
        "CQ" : "START", # Current Q

        "FJAN" : "START", # Fist day of January
        "FFEB" : "START", # Fist day of February
        "FMAR" : "START", # Fist day of March
        "FAPR" : "START", # Fist day of April
        "FMAY" : "START", # Fist day of May
        "FJUN" : "START", # Fist day of June
        "FJUL" : "START", # Fist day of July
        "FAGO" : "START", # Fist day of August
        "FSEP" : "START", # Fist day of September
        "FOCT" : "START", # Fist day of October
        "FNOV" : "START", # Fist day of November
        "FDEC" : "START", # Fist day of December

        "LJAN" : "START", # Last day of January
        "LFEB" : "START", # Last day of February
        "LMAR" : "START", # Last day of March
        "LAPR" : "START", # Last day of April
        "LMAY" : "START", # Last day of May
        "LJUN" : "START", # Last day of June
        "LJUL" : "START", # Last day of July
        "LAGO" : "START", # Last day of August
        "LSEP" : "START", # Last day of September
        "LOCT" : "START", # Last day of October
        "LNOV" : "START", # Last day of November
        "LDEC" : "START", # Last day of December

        "LHALF1" : "START", # Last day of first half
        "FHALF1" : "START", # First day of fisrt half
        "LHALF2" : "START", # Last day of second half
        "FHALF2" : "START", # First day of second half

        "LQUARTER1" : "START", # Last day of First Quarter
        "FQUARTER1" : "START", # First day of First Quarter
        "LQUARTER2" : "START", # Last day of Second Quarter
        "FQUARTER2" : "START", # First day of Second Quarter
        "LQUARTER3" : "START", # Last day of Third Quarter 
        "FQUARTER3" : "START", # First day of Third Quarter
        "LQUARTER4" : "START", # Last day of Fourth Quarter
        "FQUARTER4" : "START", # First day of Fourth Quarter

        "DAY" : "DATAFRAME", # Day dataframe
        "MONTH" : "DATAFRAME", # Month dataframe
        "HALF" : "DATAFRAME", # Half dataframe
        "QUARTER" : "DATAFRAME", # Quarter dataframe
        "YEAR" : "DATAFRAME", # Year dataframe

        "LASTDAY" : "MODIFIER", # Resolves Last day of the calculated date
        "FIRSTDAY" : "MODIFIER"} # Resolves First day of the calculated date
        
# @ar_decorator_functions.catch_exceptions
def dealWithDates(cadena, date_format = "%Y-%m-%d"):
    global lg
    cadena = str(cadena)
    """


    Parameters
    ----------
    cadena : TYPE
        DESCRIPTION. Contiene la cadena con el lenguaje de fechas @D:........@
    format : TYPE, optional
        DESCRIPTION. The default is "%Y-%m-%d". Se utilizará para formatear la cadena de vuelta y se le puede enviar cualquier modificador de decorado de fecha

    Ejemplos:

    >>> # significa ahora
    >>> cadena = "@D:NOW@"
    >>> common._print( f"cadena {cadena} = {common.dealWithDates(cadena)}")
    
    >>> cadena = "@D:NOW:LASTDAY@"
    >>> common._print( f"cadena {cadena} = {common.dealWithDates(cadena)}")    
   
    >>> # current year
    >>> cadena = "@D:CY@"
    >>> common._print( f"cadena {cadena} = {common.dealWithDates(cadena)}")
    >>> # current month
    >>> cadena = "@D:CM@"
    >>> common._print( f"cadena {cadena} = {common.dealWithDates(cadena)}")
    >>> # current H
    >>> cadena = "@D:CH@"
    >>> common._print( f"cadena {cadena} = {common.dealWithDates(cadena)}")
    >>> # current quarter
    >>> cadena = "@D:CQ@"
    >>> common._print( f"cadena {cadena} = {common.dealWithDates(cadena)}")

    >>> # ejemplos:

    >>> cadena = "@D:NOW-100DAY@" # -> esto sería START-SIGN-NUMBER-DATAFRAME
    >>> common._print( f"cadena {cadena} = {common.dealWithDates(cadena)}")

    >>> cadena = "@D:NOW-3DAY@" # -> esto sería START-SIGN-NUMBER-DATAFRAME
    >>> common._print( f"cadena {cadena} = {common.dealWithDates(cadena)}")

    >>> cadena = "@D:NOW-3DAY:LASTDAY@" # -> esto sería START-SIGN-NUMBER-DATAFRAME y MODIFIER
    >>> common._print( f"cadena {cadena} = {common.dealWithDates(cadena)}")

    >>> cadena = "@D:NOW:LFEB-35DAY@"
    >>> common._print( f"cadena {cadena} = {common.dealWithDates(cadena)}")

    >>> cadena = "@D:2030:FFEB-35DAY@"
    >>> common._print( f"cadena {cadena} = {common.dealWithDates(cadena)}")

    >>> #cadenas de fechas para producción en Cloudmore
    >>> cadena = "@D:NOW-6MONTH:LASTDAY@"
    >>> common._print( f"cadena {cadena} = {common.dealWithDates(cadena)}")    



    Returns
    -------
    TYPE
        DESCRIPTION.
        False si no es una orden de fechas
        Una cadena en la que se sustituye el comando por la fecha calculada con el formato enviado como parámetro en la función

    """

    # logger = lg.arguments(currentframe(), verbose = False)


    # detect delimiters of date commands
    start_delimiter = "@D:"
    end_delimiter = "@"

    # is it a date language specific command?
    # it needs to start with @D:
    patron_comando = r"{start_delimiter}[\w\d+-:]+{end_delimiter}".format(
        start_delimiter = start_delimiter,
        end_delimiter = end_delimiter) # buscamos secuencias de palabras dígitos, signos + y menos y : entre @

    regex_patron = patron_comando
    comandos = re.findall(regex_patron, cadena)
    if len(comandos)==0: # no es un comando
        return False

    # regex_patron = r"[0-9]{4}[\:][A-Z\d]+|[0-9]{4}[\-\:]{1}[0-9]{2}[\-\:]{1}[0-9]{2}|[A-Z]+|[\:A-Z]+|[\+\-]{1}|[\d]+"
    regex_patron_all = r"{patron_constant_date_start}|{patron_start}|{patron_year_start}|{patron_sign}|{patron_digits}|{patron_modifier}".format(
        patron_start = patron_start,
        patron_year_start = patron_year_start,
        patron_constant_date_start = patron_constant_date_start,
        patron_sign = patron_sign,
        patron_digits = patron_digits,
        patron_modifier = patron_modifier
        )

    # is a command of dates what I received?
    # result = re.findall(regex_patron, cadena)

    # if len(result)==0:
    #     return False

    std_commands = DEAL_WITH_DATES_STD_COMMANDS
    # Standard commands to create dates
    # std_commands = {
    #     "NOW": "START", # fecha actual

    #     "CY" : "START", # Current Year
    #     "CM" : "START", # Current Month
    #     "CH" : "START", # Current Half
    #     "CQ" : "START", # Current Q

    #     "FJAN" : "START", # Fist day of January
    #     "FFEB" : "START", # Fist day of February
    #     "FMAR" : "START", # Fist day of March
    #     "FAPR" : "START", # Fist day of April
    #     "FMAY" : "START", # Fist day of May
    #     "FJUN" : "START", # Fist day of June
    #     "FJUL" : "START", # Fist day of July
    #     "FAGO" : "START", # Fist day of August
    #     "FSEP" : "START", # Fist day of September
    #     "FOCT" : "START", # Fist day of October
    #     "FNOV" : "START", # Fist day of November
    #     "FDEC" : "START", # Fist day of December

    #     "LJAN" : "START", # Last day of January
    #     "LFEB" : "START", # Last day of February
    #     "LMAR" : "START", # Last day of March
    #     "LAPR" : "START", # Last day of April
    #     "LMAY" : "START", # Last day of May
    #     "LJUN" : "START", # Last day of June
    #     "LJUL" : "START", # Last day of July
    #     "LAGO" : "START", # Last day of August
    #     "LSEP" : "START", # Last day of September
    #     "LOCT" : "START", # Last day of October
    #     "LNOV" : "START", # Last day of November
    #     "LDEC" : "START", # Last day of December

    #     "LHALF1" : "START", # Last day of first half
    #     "FHALF1" : "START", # First day of fisrt half
    #     "LHALF2" : "START", # Last day of second half
    #     "FHALF2" : "START", # First day of second half

    #     "LQUARTER1" : "START", # Last day of First Quarter
    #     "FQUARTER1" : "START", # First day of First Quarter
    #     "LQUARTER2" : "START", # Last day of Second Quarter
    #     "FQUARTER2" : "START", # First day of Second Quarter
    #     "LQUARTER3" : "START", # Last day of Third Quarter 
    #     "FQUARTER3" : "START", # First day of Third Quarter
    #     "LQUARTER4" : "START", # Last day of Fourth Quarter
    #     "FQUARTER4" : "START", # First day of Fourth Quarter

    #     "DAY" : "DATAFRAME", # Day dataframe
    #     "MONTH" : "DATAFRAME", # Month dataframe
    #     "HALF" : "DATAFRAME", # Half dataframe
    #     "QUARTER" : "DATAFRAME", # Quarter dataframe
    #     "YEAR" : "DATAFRAME", # Year dataframe

    #     "LASTDAY" : "MODIFIER", # Resolves Last day of the calculated date
    #     "FIRSTDAY" : "MODIFIER"} # Resolves First day of the calculated date

    # =============================================================================
    # declare possible sintax combinations accepted - other approaches could work but this one is ok for the shake of understability
    # =============================================================================
    # ejemplos:
    # startdate = "startdate = @NOW@" -> esto sería START
    # startdate = "Startdate = @NOW-100DAY" -> esto sería START-SIGN-NUMBER-DATAFRAME
    # enddate = "enddate = @NOW-3DAY@" -> esto sería START-SIGN-NUMBER-DATAFRAME
    # endate1 = "enddate = @NOW-3DAY:LASTDAY@" -> esto sería START-SIGN-NUMBER-DATAFRAME y MODIFIER
    vocabulary = ["START", "SIGN", "DATAFRAME", "MODIFIER"]
    start_found = False
    check_sintax = [
        ["START"],
        ["START", "MODIFIER"],
        ["START", "SIGN", "NUMBER", "DATAFRAME"],
        ["START", "SIGN", "NUMBER", "DATAFRAME", "MODIFIER"],
        ["START", "MODIFIER", "SIGN", "NUMBER", "DATAFRAME"],
        ["START", "MODIFIER", "SIGN", "NUMBER", "DATAFRAME", "MODIFIER"]

        ]

    # =============================================================================
    # create the different combinations of sintax as keys
    # =============================================================================
    # to solve a problema that python has with list().index that doesn't work very well
    check_sintax_dict = {}
    key = ""
    for m in range(0, len(check_sintax)):
        key = ""
        for n in range(0,len(check_sintax[m])):
            key = key + check_sintax[m][n]
        check_sintax_dict[key] = key


    # =============================================================================
    # 0) Chequeamos que los comandos están correctamente escritos
    # para ello los extraemos y vemos si están en la lista de std_commands
    # =============================================================================
    # patron_comando = r"@[\w\d+-:]+@" # buscamos secuencias de palabras dígitos, signos + y menos y : entre @
    regex_patron = patron_comando
    comandos = re.findall(regex_patron, cadena)

    # clean delimiters
    regex_clean_start_delimiter = r"{start_delimiter}".format(
        start_delimiter = start_delimiter)
    regex_clean_end_delimiter = r"{end_delimiter}".format(
        end_delimiter = end_delimiter)

    dict_comandos = {}

    # clean delimiters in original parameter
    orig_cad = cadena
    cadena = re.sub(regex_clean_start_delimiter, "@", cadena)

    cadena_modificada = cadena
    for n in range(0,len(comandos)):

        dict_comandos[n] = re.sub(regex_clean_start_delimiter, "", comandos[n])
        dict_comandos[n] = re.sub(regex_clean_end_delimiter, "", dict_comandos[n])

        # sustituimos en la cadena original por un identificador para luego poder ir a por él
        dict_comandos[n] = dict_comandos[n].replace("+", "\+") # hay que obligar este cambio porque puede que la fecha quiera sumar DIAS o MESES, y en el patron de reg "+" significa más de una aparición, y quiero \+ como caracter...
        cadena_modificada = re.sub(dict_comandos[n],str(n),cadena_modificada)

    # =============================================================================
    # checking commands are in the vocabulary
    # we extract only commands and compare with the vocabulary declared
    # store all comands in a list
    # =============================================================================
    lista_de_comandos = []
    error = False
    # cogemos las palabras solamente (es decir, usamos como separador "+", "-" y cualquier secuencia de dígitos "\d")
    patron_tag = r"[A-Z]+\d{1}|[A-Z]+" # un tag es una secuencia de letras en mayúsculas que puede o no tener adicionalmente un dígito
    regex_patron = patron_tag
    for k,v in dict_comandos.items():
        comando = re.findall(regex_patron, v)
        for r in range(0, len(comando)):
            lista_de_comandos.append(comando[r])
            if comando[r] in std_commands:
                seguir = True
            else:
                # lg._Debug(logger, "El comando {} no existe".format(comando[r]) )
                raise Exception(f"El comando no existe {comando[r]}")
                seguir = False
            if seguir == False:
                error = True
                break

    # =============================================================================
    # SINTAX CHECKING
    # =============================================================================
    sintax = {}

    # declare different variables to get and store the information
    date = datetime.now()
    year = 0
    month = 0
    day = 0
    months = 0
    days = 0
    tokens = {}

    for k,v in dict_comandos.items():
        # v = dict_comandos[0] # PARA CADA ELEMENTO!!!!
        cadena_original = v # guardo la cadena original
        # deshago la cadena original y la sustituyo por sus elementos constituyentes
        # busco fecha completa, palabras todas en mayúsculas, signo, dígitos y un posible modificador

        final_commands = []
        # conseguimos los trozos (lógicos)
        comando = re.findall(regex_patron_all, cadena_original)
        # hacemos un split de los : para hacer los comandos correctamente (puede que un trozo lógico contenga un modifier)
        for r in range(0, len(comando)):
            split_command = re.split(r":", comando[r])
            for n in range(0,len(split_command)):
                if len(split_command[n])>0: # solo metemos aquéllo que nos devuelve algo con sentido (":LASTDAY" devuelve ['','LASTDAY'] así que quitaríamos el primero)
                    final_commands.append(split_command[n])



        comando = final_commands.copy()

        tokens[k] = comando.copy()
        

        for r in range(0, len(comando)):
            # busco en la cadena de sintaxis para construirlo
            # si son letras tengo que buscar en std_commands
            # en otro caso tengo que decir si son dígitos o es un signo y ponerlo
            # si alguna entrada ha quedado vacía significa que la sintaxis está mal puesta
            # r=0
            if comando[r] in std_commands:
                tokens[k][r] = std_commands[comando[r]]
                # ! evitamos múltiples START (sólo puede habe uno al comienzo), sin esto daba problemas en "@D:NOW+1000DAY@" porque confundía 1000 con un START (year)
                if std_commands[comando[r]] == "START":
                    start_found = True
            else:

                # no es una orden pero puede que hayan enviado una fecha concreta
                patron_year = r"[0-9]{4}"
                regex_patron = patron_year
                encontrado = re.findall(regex_patron, comando[r])
                if len(encontrado)>0:
                    if not start_found:
                        tokens[k][r] = "START" # ? sólo podemos encontrar start una vez, cualquier otra cosa aunque pueda parecer un start no lo es (por ejemplo, una sequencia de 4 digits)
                        start_found = True
                    else:
                        # is it a sign? or is it a couple of digits?
                        if comando[r] == "+" or comando[r] == "-":
                            tokens[k][r] = "SIGN"
                        else:
                            tokens[k][r] = "UNKNOWN"

                            regex_patron = r"[0-9]+"
                            resultado = re.findall(regex_patron, comando[r])
                            if len(resultado)>0:
                                tokens[k][r] = "NUMBER"
                        
                else:
                    # is it a sign? or is it a couple of digits?
                    if comando[r] == "+" or comando[r] == "-":
                        tokens[k][r] = "SIGN"
                    else:
                        tokens[k][r] = "UNKNOWN"

                        regex_patron = r"[0-9]+"
                        resultado = re.findall(regex_patron, comando[r])
                        if len(resultado)>0:
                            tokens[k][r] = "NUMBER"

            # =============================================================================
            # IF SINTAX OK
            # construimos una entrada KEY con toda la entrada de TOKENS y vemos si está en check_sintax_dict (pendiente)
            # we build the key and check with the possible sintax accepted
            # =============================================================================

            tokens_sintax = ""
            prior_token = ""
            date = False

            for k,v in tokens.items():
                # v= tokens[0]
                for n in range(0, len(v)):
                    # permitimos acumular START y MODIFIER (cuando aparece más de uno encadenado se considera OK por la sintaxis)
                    # n=0
                    if v[n] != prior_token:
                        tokens_sintax = tokens_sintax + v[n]
                    prior_token = v[n]

                if tokens_sintax in check_sintax_dict:
                    for r in range(0, len(comando)):
                        # r=5
                        pcommand = comando[r]
                        if tokens[k][r] == "START":
                            date_str = solveSTART(date, pcommand)
                        if tokens[k][r] == "MODIFIER":
                            date_str = solveMODIFIER(date, pcommand)
                        if tokens[k][r] == "SIGN":
                            sign = pcommand
                            number = comando[r+1]
                            dataframe = comando[r+2]
                            date_str = solveOPERATION(date, sign, number, dataframe)
                        date = datetime.strptime(date_str, "%Y-%m-%d")
                        
        # quiere cambiar cada aparición de fecha por su resultado, porque puede que en la cadena haya más de una...
        try:
            cadena_modificada = re.sub("@{k}@".format(k=k), "{date}".format(date=date.strftime(date_format)), cadena_modificada)
        except Exception as e:
            tc._print(f"error procesando tokens {tokens} en {cadena_modificada} -> {e}")
            a = 999

    return cadena_modificada


# =============================================================================
# Calcula el año y el mes quitándole a un año y un mes, un número de meses
# =============================================================================
def calcularDeltaMonths(currentyear = 2020, currentmonth = 3, nmonths = 22, sign = "-"):

    targetyear = currentyear
    intyears = 0 # número de años en integer
    
    if sign == "-":
        if nmonths >= currentmonth: # si vamos a quitar más meses de los que llevamos tenemos que calcular el año al que apuntamos
            res_months = nmonths - currentmonth
            intyears = int(res_months / 12)
            totyears = nmonths / 12
            if totyears > intyears: # como el número obtenido no es exacto, sumamos un año
                intyears += 1
        years = intyears
        targetyear = currentyear - years
        months = currentmonth - nmonths
        month = months % 12
        if month == 0:
            month = 12
    else:
        res_months = nmonths + currentmonth
        intyears = int(res_months / 12)
        totyears = int(nmonths / 12)
        
        # if totyears > intyears: # como el número obtenido no es exacto, sumamos un año
        #     intyears += 1
        years = intyears
        targetyear = currentyear + years
        months = currentmonth + nmonths
        month = months % 12
        # ! fix no funcionaba bien cuando se pasaba de año...
        if (currentmonth + nmonths)%12 == 0:
            if sign == "+":
                month = 12
                targetyear-=1
            if sign == "-":
                month = 1
                targetyear+=1
        # ! fin del fix no funcionaba bien cuando se pasaba de año...
                
        
    lastday, lmonth, error = lastdayofmonth(targetyear, month)
    return targetyear, month, lastday

# quitar días a una fecha
# from datetime import timedelta
# _print(datetime.now() - timedelta(days=18))

# =============================================================================
# Quita a una fecha un número de días concretos
# =============================================================================
def calcularDeltaDays(startdate, delta, sign = "-"):
    from datetime import timedelta
    resultado = 0
    if sign == "-":
        resultado = startdate - timedelta(days=delta)
    else:
        resultado = startdate + timedelta(days=delta)
    return resultado


# =============================================================================
# calcula el último día del mes de cualquier mes
# devuelve True si ha habido un error interno (mes enviado <0 o >12)
# =============================================================================
def lastdayofmonth(year, month):
    # year = 2019
    # month = 2
    import calendar
    lmonth = max(1,month)
    lmonth = min(lmonth, 12)
    error = False
    if lmonth != month:
        error = True

    return calendar.monthrange(year,lmonth)[1], lmonth, error


# =============================================================================
# Apply a modifier to a date and returns the new calculated date
# =============================================================================
def solveMODIFIER(date, pcommand):
    # preload of data
    currentdate = date
    # operación con cadena
    year = currentdate.year
    month = currentdate.month
    day = currentdate.day

    calculated_date = ""
    if pcommand == "LASTDAY":
        calculated_date = "{year}-{month}-{day}".format(
            year=year,
            month=month,
            day = lastdayofmonth(year,month)[0])

    if pcommand == "FIRSTDAY":
        calculated_date = "{year}-{month}-{day}".format(
            year=year,
            month=month,
            day = 1)

    return calculated_date
# =============================================================================
# Makes an operation with a date and returns the new calculated date
# =============================================================================
def solveOPERATION(date, sign, number, dataframe):
    # preload of data
    currentdate = date
    # operación con cadena
    year = currentdate.year
    month = currentdate.month
    day = currentdate.day

    # sign = pcommand
    number = int(number)

    if dataframe == "DAY":
        date = calcularDeltaDays(date, number, sign)
        currentdate = date
        year = currentdate.year
        month = currentdate.month
        day = currentdate.day
        
    if dataframe == "YEAR": # transformo el año en meses
        number = number * 12
        dataframe = "MONTH"

    if dataframe == "MONTH":
        year, month, lastday = calcularDeltaMonths(year, month, number, sign)
        day = lastday




    calculated_date = "{year}-{month}-{day}".format(
        year=year,
        month=month,
        day=day)

    return calculated_date

# =============================================================================
# Solve a start date that can be modified by operations and modifiers afterwords
# =============================================================================
def solveSTART(date, pcommand):

    if date == False:
        # preload of data
        currentdate = datetime.now()
    else:
        currentdate = date

    # precalculo fechas para operar a partir de ahí
    today = "{now}".format(now=currentdate.strftime("%Y-%m-%d"))
    today = currentdate.strftime("%Y-%m-%d")
    # operación con cadena
    year = currentdate.year
    month = currentdate.month
    day = currentdate.day
    # calculamos el h
    h = 1 if month > 1 and month < 7 else 2
    # calculamos el q
    q = 1
    if month > 1 and month < 4:
        q = 0
    else:
        if month > 3 and month < 7:
            q = 1
        else:
            if month > 7 and month < 10:
                q = 2
                if month > 9 and month < 13:
                    q = 3
    mes_inicio_q = [1,4,7,10]

    resolve_cad = True # le indicamos que llegue hasta el final para la resolución de comandos start
    # =============================================================================
    # # COMPLETE DATE
    # =============================================================================
    # do I have a complete date in the form YYYY-MM-DD
    regex_patron = patron_constant_date_start
    encontrado = re.findall(regex_patron, pcommand)
    cad = ""
    if len(encontrado)>0: # it's a YEAR:START pattern
        # extract all components
        date = datetime.strptime("{date}".format(
            date = pcommand), "%Y-%m-%d")
        cad = "{}-{}-{}".format(year,month,day)
        resolve_cad = False # is a final date


    # =============================================================================
    # # ONLY YEAR
    # =============================================================================

    # extract year
    patron_year = r"[0-9]{4}"
    regex_patron = patron_year
    encontrado = re.findall(regex_patron, pcommand)
    if len(encontrado)>0: # it's a YEAR:START pattern
        year = int(encontrado[0])
        month = 1
        day = 1
        cad = "{}-{}-{}".format(year,month,day)
        resolve_cad = False # is a year only, no resolve command


    # # another START statement?
    # regex_patron = patron_start
    # encontrado = re.findall(regex_patron, pcommand)
    # if len(encontrado)>0: # it's a YEAR:START pattern
    #     pcommand = encontrado[0]

    # =============================================================================
    # # OTHER START COMMANDS
    # =============================================================================
    if resolve_cad == True:
        cadena = "{" + pcommand + "}"
    else:
        cadena = "{year}-{month}-{day}"


    calculated_date = cadena.format(
        NOW=currentdate.strftime("%Y-%m-%d"),

        year=year,
        month=month,
        day=day,

        CY = "{year}-1-1".format(year=year), # devuelvo el primer día del año en curso
        CM = "{year}-{month}-1".format(year=year, month=month), # devuelvo el primer día del mes en curso
        CH = "{year}-{lmonth}-1".format(year=year, lmonth = lastdayofmonth(year,h*6)[1]), # devuelvo el primer día del H en curso
        CQ = "{year}-{lmonth}-1".format(year=year, lmonth = lastdayofmonth(year,mes_inicio_q[q])[1]), # devuelvo el primer día del Q en curso

        FJAN = "{year}-1-1".format(year=year),
        FFEB = "{year}-2-1".format(year=year),
        FMAR = "{year}-3-1".format(year=year),
        FAPR = "{year}-4-1".format(year=year),
        FMAY = "{year}-5-1".format(year=year),
        FJUN = "{year}-6-1".format(year=year),
        FJUL = "{year}-7-1".format(year=year),
        FAGO = "{year}-8-1".format(year=year),
        FSEP = "{year}-9-1".format(year=year),
        FOCT = "{year}-10-1".format(year=year),
        FNOV = "{year}-11-1".format(year=year),
        FDEC = "{year}-12-1".format(year=year),

        LJAN = "{year}-{lmonth}-{day}".format(year=year, lmonth = lastdayofmonth(year,1)[1], day = lastdayofmonth(year,1)[0]),
        LFEB = "{year}-{lmonth}-{day}".format(year=year, lmonth = lastdayofmonth(year,2)[1], day = lastdayofmonth(year,2)[0]),
        LMAR = "{year}-{lmonth}-{day}".format(year=year, lmonth = lastdayofmonth(year,3)[1], day = lastdayofmonth(year,3)[0]),
        LABR = "{year}-{lmonth}-{day}".format(year=year, lmonth = lastdayofmonth(year,4)[1], day = lastdayofmonth(year,4)[0]),
        LMAY = "{year}-{lmonth}-{day}".format(year=year, lmonth = lastdayofmonth(year,5)[1], day = lastdayofmonth(year,5)[0]),
        LJUN = "{year}-{lmonth}-{day}".format(year=year, lmonth = lastdayofmonth(year,6)[1], day = lastdayofmonth(year,6)[0]),
        LJUL = "{year}-{lmonth}-{day}".format(year=year, lmonth = lastdayofmonth(year,7)[1], day = lastdayofmonth(year,7)[0]),
        LAGO = "{year}-{lmonth}-{day}".format(year=year, lmonth = lastdayofmonth(year,8)[1], day = lastdayofmonth(year,8)[0]),
        LSEP = "{year}-{lmonth}-{day}".format(year=year, lmonth = lastdayofmonth(year,9)[1], day = lastdayofmonth(year,9)[0]),
        LOCT = "{year}-{lmonth}-{day}".format(year=year, lmonth = lastdayofmonth(year,10)[1], day = lastdayofmonth(year,10)[0]),
        LNOV = "{year}-{lmonth}-{day}".format(year=year, lmonth = lastdayofmonth(year,11)[1], day = lastdayofmonth(year,11)[0]),
        LDEC = "{year}-{lmonth}-{day}".format(year=year, lmonth = lastdayofmonth(year,12)[1], day = lastdayofmonth(year,12)[0]),

        LHALF1 = "{year}-{lmonth}-{day}".format(year=year, lmonth = lastdayofmonth(year,6)[1], day = lastdayofmonth(year,6)[0]),
        LHALF2 = "{year}-{lmonth}-{day}".format(year=year, lmonth = lastdayofmonth(year,12)[1], day = lastdayofmonth(year,12)[0]),
        FHALF1 = "{year}-6-1".format(year=year),
        FHALF2 = "{year}-12-1".format(year=year),

        LQUARTER1 = "{year}-{lmonth}-{day}".format(year=year, lmonth = lastdayofmonth(year,3)[1], day = lastdayofmonth(year,3)[0]),
        LQUARTER2 = "{year}-{lmonth}-{day}".format(year=year, lmonth = lastdayofmonth(year,6)[1], day = lastdayofmonth(year,6)[0]),
        LQUARTER3 = "{year}-{lmonth}-{day}".format(year=year, lmonth = lastdayofmonth(year,9)[1], day = lastdayofmonth(year,9)[0]),
        LQUARTER4 = "{year}-{lmonth}-{day}".format(year=year, lmonth = lastdayofmonth(year,12)[1], day = lastdayofmonth(year,12)[0]),

        FQUARTER1 = "{year}-3-1".format(year=year),
        FQUARTER2 = "{year}-6-1".format(year=year),
        FQUARTER3 = "{year}-9-1".format(year=year),
        FQUARTER4 = "{year}-12-1".format(year=year)

        )

    return calculated_date

# cadena = "@D:NOW@"
# _print( f"cadena {cadena} = {dealWithDates(cadena)}")


import winreg
def get_sys_env(name):
    key = winreg.CreateKey(winreg.HKEY_LOCAL_MACHINE, r"System\CurrentControlSet\Control\Session Manager\Environment")
    return winreg.QueryValueEx(key, name)[0]

def get_user_env(name, default = ""):
    retorno = os.getenv(name) # intenta en .env y en la copia local del environment (que puede que haya sufrido un cambio)
    if retorno is None: # ahora voy a las variables activas del sistema
        key = winreg.CreateKey(winreg.HKEY_CURRENT_USER, r"Environment")
        retorno = default
        try:
            retorno = winreg.QueryValueEx(key, name)[0]
        except:
            retorno = default
    return retorno

# ---- .
# ---- tests ------------------------------------------------------------------
def test_():
    
    cadena = "@D:NOW-24MONTH:LASTDAY@"
    _print( f"cadena {cadena} = {dealWithDates(cadena)}")    
    
    cadena = "@D:NOW+2YEAR:LASTDAY@"
    _print( f"cadena {cadena} = {dealWithDates(cadena)}")    




# =============================================================================
# Función para testear fechas
# =============================================================================
def test_dealWithDates():

    # significa ahora
    cadena = "@D:NOW@"
    _print( f"cadena {cadena} = {dealWithDates(cadena)}", Y)
    
    cadena = "@D:NOW-6MONTH:LASTDAY@"
    _print( f"cadena {cadena} = {dealWithDates(cadena)}")
   
    # current year
    cadena = "@D:CY@"
    _print( f"cadena {cadena} = {dealWithDates(cadena)}")
    # current month
    cadena = "@D:CM@"
    _print( f"cadena {cadena} = {dealWithDates(cadena)}")
    # current H
    cadena = "@D:CH@"
    _print( f"cadena {cadena} = {dealWithDates(cadena)}")
    # current quarter
    cadena = "@D:CQ@"
    _print( f"cadena {cadena} = {dealWithDates(cadena)}")

    # ejemplos:

    cadena = "@D:NOW@"
    _print( f"cadena {cadena} = {dealWithDates(cadena)}")

    cadena = "@D:NOW-100DAY@" # -> esto sería START-SIGN-NUMBER-DATAFRAME
    _print( f"cadena {cadena} = {dealWithDates(cadena)}")

    cadena = "@D:NOW-3DAY@" # -> esto sería START-SIGN-NUMBER-DATAFRAME
    _print( f"cadena {cadena} = {dealWithDates(cadena)}")

    cadena = "@D:NOW-3DAY:LASTDAY@" # -> esto sería START-SIGN-NUMBER-DATAFRAME y MODIFIER
    _print( f"cadena {cadena} = {dealWithDates(cadena)}")

    cadena = "@D:NOW:LFEB-35DAY@"
    _print( f"cadena {cadena} = {dealWithDates(cadena)}")

    cadena = "@D:2030:FFEB-35DAY@"
    _print( f"cadena {cadena} = {dealWithDates(cadena)}")

    #cadenas de fechas para producción en Cloudmore
    cadena = "@D:NOW-6MONTH:LASTDAY@"
    _print( f"cadena {cadena} = {dealWithDates(cadena)}")

    cadena = "@D:NOW:LASTDAY@"
    _print( f"cadena {cadena} = {dealWithDates(cadena)}")

# pruebas de cálculo de fechas aleatorias pero sin usar el lenguaje de fechas
    import random
    # para hacer pruebas
    years = list(range(2015,2020))
    currentmonth = 3
    months = list(range(1,40))
    for y in range(0, len(years)):
        currentyear = random.sample(years, 1)
        for m in range(0,len(months)):
            nmonths = random.sample(months, 1)

            _print("Año y mes resultado quitando a {currentyear}-{currentmonth} - {nmonths} - nos daría {targetyear}-{targetmonth}, siendo su último día {ultdia}".format(
                currentyear = currentyear,
                currentmonth = currentmonth,
                nmonths = nmonths[0],
                targetyear = calcularDeltaMonths(currentyear[0], currentmonth, nmonths[0])[0],
                targetmonth = calcularDeltaMonths(currentyear[0], currentmonth, nmonths[0])[1],
                ultdia = "{lmonth}-{day}".format(lmonth = lastdayofmonth(calcularDeltaMonths(currentyear[0], currentmonth, nmonths[0])[0],calcularDeltaMonths(currentyear[0], currentmonth, nmonths[0])[1])[1], day = lastdayofmonth(calcularDeltaMonths(currentyear[0], currentmonth, nmonths[0])[0],calcularDeltaMonths(currentyear[0], currentmonth, nmonths[0])[1])[0])
                ))

# import re
# =============================================================================
# Aplicar una expresión lógica sobre una cadena y devuelve el resultado
# =============================================================================
# def aplicarREGEX(regex_to_apply = "", change_to = "", string_in = ""):
#     string_out = re.sub(regex_to_apply, change_to, string_in )
#     return string_out

# cadena_a_cambiar = "D:/OneDrive - Seachad/06.- APICALL/Reports/bd59a770-df9d-4947-a787-b3be57b4b844/subscriptions/3be226db-7330-491e-8bc6-eca10000d669/M_RoleAssignments"
# cadena_resultado = "D:/OneDrive - Seachad/06.- APICALL/Reports/bd59a770-df9d-4947-a787-b3be57b4b844_3be226db-7330-491e-8bc6-eca10000d669_M_RoleAssignments"

# import re
# regex_to_apply = "/subscriptions/|(/)(\w+)·"
# change_to = "\\1_\\2"
# string_in = f"·{cadena_a_cambiar}·" # <- incluimos unas marcas para garantizar que cambiamos lo que queremos cambiar
# string_out = re.sub(regex_to_apply, change_to, string_in )
# _print(f"[{string_in}] se convierte en: \n [{string_out}]")



if __name__ == "__main__":

    # test_dealWithDates()
    # test_compress_uncompress()
    pass




