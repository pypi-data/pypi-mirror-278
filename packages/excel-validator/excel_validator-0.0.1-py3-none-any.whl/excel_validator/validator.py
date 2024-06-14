# pylint: disable=line-too-long,invalid-name,consider-using-f-string
"""Module providing validation classes."""

import logging
import os
import sys
import re
import time
import itertools
from shutil import copyfile
import json
import jsonschema
import pandas as pd
import tempfile
from tqdm import tqdm
from openpyxl import reader, load_workbook
from frictionless import Package, Dialect, Resource, Schema, Checklist, Field, fields, settings, formats, describe
from ._version import __version__
from .report import ValidationReport
from .functions import setDescriptorValueDynamicString
from .functions import setDescriptorValueDynamicBoolean
from .functions import setDescriptorValueDynamicInteger
from .functions import setDescriptorValueDynamicNumber
from .functions import setDescriptorValueDynamicList

class Validate:

    """
    xlsx validator
    
    Parameters
    ----------------------
    filename : str
        location of the file that should be validated

    configuration: str
        location of the configuration file
        
    """

    ERROR_GENERAL              = "General Problem"
    ERROR_NO_WRITE_ACCESS      = "No Write Access"
    ERROR_NO_SHEET_NAMES       = "No Sheet Names"
    ERROR_NO_SHEET             = "No Sheet"
    ERROR_MISSING_SHEETS       = "Missing Sheets"
    ERROR_UNEXPECTED_SHEETS    = "Unexpected Sheets"
    ERROR_ORDER_SHEETS         = "Incorrect Order Sheets"
    ERROR_MISSING_COLUMNS      = "Missing Columns"
    ERROR_UNRECOGNIZED_COLUMNS = "Unrecognized Columns"
    ERROR_ORDER_COLUMNS        = "Incorrect Order Columns"

    WARNING_TYPE               = "Incorrect Type"
    WARNING_EMPTY_ROW          = "Empty Rows"
    WARNING_EMPTY_COLUMN       = "Empty Columns"
    WARNING_NO_SCHEMA          = "No Schema"
    WARNING_MODULE             = "Module"
    WARNING_EMPTY_PACKAGE      = "Empty Package"
    

    def __init__(self, filename: str, configuration: str, **args):
        #logging
        self._logger = logging.getLogger(__name__)

        reader.excel.warnings.simplefilter(action="ignore")

        #initialise
        self._wb = None
        
        #check file and define basepath for output
        filename = os.path.abspath(filename)
        assert os.path.isfile(filename) and os.access(filename, os.R_OK), \
                "File '%s' doesn't exist or isn't readable" % filename

        #include progress bar for cli call
        self._cli = bool(args.get("cli", False))
        
        if args.get("create", False):
            config_filename = str(args.get("create"))
            configDirectoryPattern = re.compile("^[a-zA-Z\_]+$")
            if configDirectoryPattern.match(config_filename):
                config_location = os.path.abspath(config_filename)
                if os.path.exists(config_filename):
                    raise Exception("couldn't create configuration, already exists")
                else:                    
                    self._filename = os.path.basename(filename)
                    with tempfile.TemporaryDirectory() as tmp_dir:
                        output_filename = os.path.join(tmp_dir,self._filename)
                        copyfile(filename,output_filename)
                        self._basepath = tmp_dir
                        self._createConfiguration(config_location)
                    self._basepath = None
            else:
                raise Exception("couldn't create configuration at %s" % config_filename)

        else:            
            #initialise for validation
            self._package = Package()
            self._name = args.get("name", os.path.basename(filename))
            self._report = ValidationReport(self._logger)
            self._expectedSheets = set()
            #check configuration
            configurationFilename = os.path.abspath(configuration)
            assert os.path.isfile(configurationFilename) and os.access(configurationFilename, os.R_OK), \
                    "Configuration file '{}' doesn't exist or isn't readable".format(configuration)
            self._parseConfiguration(configurationFilename)
            #add modules to path
            if self._modulesPath and not self._modulesPath in sys.path: sys.path.append(self._modulesPath)
            #start validation
            if self._config:
                self._filename = os.path.basename(filename)
                if not args.get("updateFile", False):
                    with tempfile.TemporaryDirectory() as tmp_dir:
                        output_filename = os.path.join(tmp_dir,self._filename)
                        copyfile(filename,output_filename)
                        self._basepath = tmp_dir
                        self._validate()
                    self._basepath = None
                else:
                    self._basepath = os.path.dirname(filename)
                    self._validate()
               
    def _createConfiguration(self, configurationLocation: str):
        #initialise
        progression_bar = tqdm(total=2, disable=(not self._cli), leave=False)
        progression_bar_size = 25
        progression_bar.set_description(str("Initialise").ljust(progression_bar_size))
        #check file
        package = Package()
        resourceFilename = os.path.join(self._basepath, self._filename)
        self._wb = load_workbook(os.path.join(resourceFilename))
        self._availableSheets = list(self._wb.sheetnames)
        #update
        progression_bar.reset(total=len(self._availableSheets)+2)        
        #initialise
        os.mkdir(configurationLocation)
        os.mkdir(os.path.join(configurationLocation,"schema"))
        configuration = {
            "settings": {
                "schemaPath": "schema",
                "allowAdditionalSheets": False,
                "requireSheetOrder": True,
                "adjustTypeForStringColumns": True,
                "removeEmptyRows": True,
                "removeEmptyColumns": True
            }, 
            "sheets":[]
        }
        resourceNames = set()
        progression_bar.update(1)
        for i,name in enumerate(self._availableSheets):
            #set progress
            n = progression_bar_size - 10
            sheet = "[%s...]"%name[:(n-3)] if len(name)>n else "[%s]"%name
            progression_bar.set_description("Analyse %s" % sheet.ljust(n+2))
            #filter
            self._removeEmptyRowsForSheet(name)
            self._removeEmptyColumnsForSheet(name)
            #describe
            resource = "sheet%s" % str(i).zfill(2)
            sheetConfig = {"name": name, "resource": resource, "schema": {}}
            dialect = Dialect()
            dialect.add_control(formats.ExcelControl(sheet=name, preserve_formatting=False))
            schemaResource = describe(resourceFilename, type="schema", dialect=dialect)
            schemaResource.to_json(os.path.join(configurationLocation,"schema","%s.json"%resource))
            sheetConfig["schema"]["file"] = "%s.json"%resource
            configuration["sheets"].append(sheetConfig)
            #update
            progression_bar.update(1)            
        #check configuration
        progression_bar.set_description(str("Validate configuration").ljust(progression_bar_size))
        with open(os.path.join(os.path.dirname(__file__),"config.json"), encoding="UTF-8") as configurationSchema:
            schema = json.load(configurationSchema)
        jsonschema.validate(configuration,schema)
        progression_bar.update(1)
        #store
        with open(os.path.join(configurationLocation,"config.json"), "w") as f:
            json.dump(configuration, f, ensure_ascii=False)
        
    def _parseConfiguration(self, configurationFilename: str):
        self._config = None
        self._allowedSheets = []
        self._resourceSheet = {}
        self._modulesPath = None
        self._schemaPath = None
        try:
            with open(configurationFilename, encoding="UTF-8") as configurationData:
                self._config = json.load(configurationData)
            with open(os.path.join(os.path.dirname(__file__),"config.json"), encoding="UTF-8") as configurationSchema:
                schema = json.load(configurationSchema)
            jsonschema.validate(self._config,schema)
            #set paths
            self._modulesPath = self._config.get("settings",{}).get("modulesPath",None)
            if self._modulesPath:
                if not os.path.isabs(self._modulesPath):
                    self._modulesPath = os.path.abspath(os.path.join(os.path.dirname(configurationFilename),self._modulesPath))
            self._schemaPath = self._config.get("settings",{}).get("schemaPath",None)
            if self._schemaPath:
                if not os.path.isabs(self._schemaPath):
                    self._schemaPath = os.path.abspath(os.path.join(os.path.dirname(configurationFilename),self._schemaPath))
            #loop over definitions
            for i in range(len(self._config["sheets"])):
                #require unique name
                assert self._config["sheets"][i]["name"] not in self._allowedSheets, "sheet name should be unique"
                self._allowedSheets.append(self._config["sheets"][i]["name"])
                if "resource" in self._config["sheets"][i]:
                    #require unique resource
                    assert self._config["sheets"][i]["resource"] not in self._resourceSheet, "resource name should be unique"
                    self._resourceSheet[self._config["sheets"][i]["resource"]] = self._config["sheets"][i]["name"]
                    #schema
                    if "schema" in self._config["sheets"][i]:
                        if "file" in self._config["sheets"][i]["schema"]:
                            if self._schemaPath:
                                schemaFilename = os.path.join(self._schemaPath,
                                                          self._config["sheets"][i]["schema"]["file"])
                            else:
                                schemaFilename = os.path.join(os.path.dirname(configurationFilename),
                                                          self._config["sheets"][i]["schema"]["file"])
                            with open(schemaFilename, encoding="UTF-8") as schemaData:
                                self._config["sheets"][i]["schema"]["data"] = json.load(schemaData)
                    #checklist
                    if "checklist" in self._config["sheets"][i]:
                        if "file" in self._config["sheets"][i]["checklist"]:
                            if self._schemaPath:
                                checklistFilename = os.path.join(self._schemaPath,
                                                          self._config["sheets"][i]["checklist"]["file"])
                            else:
                                checklistFilename = os.path.join(os.path.dirname(configurationFilename),
                                                          self._config["sheets"][i]["checklist"]["file"])
                            with open(checklistFilename, encoding="UTF-8") as checklistData:
                                self._config["sheets"][i]["checklist"]["data"] = json.load(checklistData)
            #reloop over definitions to get dependencies
            for i in range(len(self._config["sheets"])):
                if not "dependencies" in self._config["sheets"][i]:
                    self._config["sheets"][i]["dependencies"] = []
                #check schema
                if "resource" in self._config["sheets"][i] and "schema" in self._config["sheets"][i]:
                    #update dependencies
                    for entry in self._config["sheets"][i]["schema"]["data"].get("foreignKeys",[]):
                        if "reference" in entry and "resource" in entry["reference"]:
                            if not entry["reference"]["resource"]==self._config["sheets"][i]["resource"]:
                                if entry["reference"]["resource"] in self._resourceSheet:
                                    sheetName = self._resourceSheet[entry["reference"]["resource"]]
                                    if not sheetName in self._config["sheets"][i]["dependencies"]:
                                        self._config["sheets"][i]["dependencies"].append(sheetName)
                                        self._logger.debug("set dependency on '%s for '%s' based on frictionless schema",
                                            sheetName,self._config["sheets"][i]["name"])
                    for entry in self._config["sheets"][i]["schema"].get("dynamic",[]):
                        if "dynamicResources" in entry:
                            for value in entry["dynamicResources"].values():
                                if value["resource"] in self._resourceSheet:
                                    sheetName = self._resourceSheet[value["resource"]]
                                    if not sheetName in self._config["sheets"][i]["dependencies"]:
                                        self._config["sheets"][i]["dependencies"].append(sheetName)
                                        self._logger.debug("set dependency on '%s' for '%s' based on dynamic frictionless schema",
                                                           sheetName,self._config["sheets"][i]["name"])
                        if "linkedResources" in entry:
                            for value in entry["linkedResources"].values():
                                if value["resource"] in self._resourceSheet:
                                    sheetName = self._resourceSheet[value["resource"]]
                                    if not sheetName in self._config["sheets"][i]["dependencies"]:
                                        self._config["sheets"][i]["dependencies"].append(sheetName)
                                        self._logger.debug("set dependency on '%s' for '%s' based on dynamic frictionless schema",
                                            sheetName,self._config["sheets"][i]["name"])
        except Exception as ex:
            self._logger.error("Could not parse configuration: %s", str(ex))
            self._config = None

    def _computeExpectedSheets(self):
        self._expectedSheets = set()
        #recursively get dependencies
        def getAllDependencies(dependencies):
            checkDependencies = [sheetName for sheetName in dependencies]
            newDependencies = [sheetName for sheetName in dependencies]
            while len(checkDependencies)>0:
                newCheckDependencies = []
                for entry in self._config["sheets"]:
                    if ("dependencies" in entry) and (entry["name"] in checkDependencies):
                        for dependency in entry["dependencies"]:
                            if not dependency in newDependencies:
                                newDependencies.append(dependency)
                                newCheckDependencies.append(dependency)
                checkDependencies = newCheckDependencies
            return newDependencies
        #check defined sheets
        for i in range(len(self._config["sheets"])):
            if (self._config["sheets"][i]["name"] in self._availableSheets) or not self._config["sheets"][i]["optional"]:
                self._expectedSheets.add(self._config["sheets"][i]["name"])
                if "dependencies" in self._config["sheets"][i]:
                    dependencies = []
                    for sheetName in self._config["sheets"][i]["dependencies"]:
                        dependencies.append(sheetName)
                    self._expectedSheets.update(getAllDependencies(dependencies))

    def _computeValidationOrder(self):
        validationEntries = []
        validationSheets = []
        recheckList = []
        for entry in self._config["sheets"]:
            if "resource" in entry and "schema" in entry:
                if entry["name"] in self._availableSheets:
                    if len(set(entry["dependencies"]).difference(validationSheets))==0:
                        validationEntries.append(entry)
                        validationSheets.append(entry["name"])
                    else:
                        recheckList.append(entry)
                    recheck = True
                    while recheck:
                        recheck = False
                        newRecheckList = []
                        for entry in recheckList:
                            if len(set(entry["dependencies"]).difference(validationSheets))==0:
                                validationEntries.append(entry)
                                validationSheets.append(entry["name"])
                                recheck = True
                            else:
                                newRecheckList.append(entry)
                        recheckList = newRecheckList
        if len(recheckList)>0:
            recheckSheetList = [entry["name"] for entry in recheckList]
            self._report.addReportError("general", "can't solve dependencies for sheets","{}'".format("', '".join(recheckSheetList)))
        return validationEntries

    def _getSheetColumnNames(self, sheetName:str):
        self._logger.debug("get column names from sheet '{}'".format(sheetName))
        columnNames = []
        if not self._wb:
            resourceFilename = os.path.join(self._basepath, self._filename)
            self._wb = load_workbook(os.path.join(resourceFilename))
        if not sheetName in self._wb.sheetnames:
            self._logger.error("can't find sheet {} in resource".format(sheetName))
        else:
            ws = self._wb[sheetName]
            for row in ws.rows:
                columnNames = [cell.internal_value for cell in row]
                break
        return columnNames

    def _checkMissingColumns(self, sheetName: str, resource:Resource, reportId:str):
        adjusted_resource = False
        if not reportId in self._report.reports:
            self._logger.error("reportId {} not found".format(reportId))
        elif not sheetName in self._availableSheets:
            self._logger.error("sheetName {} not found".format(sheetName))
        else:
            try:
                headerCase = resource.dialect.header_case
                resourceColumnNames = self._getSheetColumnNames(sheetName)
                requiredNames = resource.schema.field_names
                if headerCase:
                    missingNames = set(requiredNames).difference(resourceColumnNames)
                    unrecognizedNames = set(resourceColumnNames).difference(requiredNames)
                else:
                    resourceColumnNamesUpper = [entry.upper() for entry in resourceColumnNames]
                    requiredNamesUpper = [entry.upper() for entry in requiredNames]
                    missingNames = set([x for x,y in zip(requiredNames,requiredNamesUpper)
                                        if not y in resourceColumnNamesUpper])
                    unrecognizedNames = set([x for x,y in zip(resourceColumnNames,resourceColumnNamesUpper)
                                             if not y in requiredNamesUpper])
                if len(missingNames)>0:
                    self._report.addReportError(reportId, Validate.ERROR_MISSING_COLUMNS, "'{}'".format("', '".join(missingNames)))
                    adjusted_resource = True
                if len(unrecognizedNames)>0:
                    self._report.addReportError(reportId, Validate.ERROR_UNRECOGNIZED_COLUMNS, "'{}'".format("', '".join(unrecognizedNames)))
                    adjusted_resource = True
                if headerCase:
                    recognizedResourceNames = [item for item in resourceColumnNames if item in requiredNames]
                    recognizedRequiredNames = [item for item in requiredNames if item in resourceColumnNames]
                else:
                    recognizedResourceNames = [item for item in resourceColumnNamesUpper if item in requiredNamesUpper]
                    recognizedRequiredNames = [item for item in requiredNamesUpper if item in resourceColumnNamesUpper]
                if not recognizedResourceNames==recognizedRequiredNames:
                    self._report.addReportError(reportId, Validate.ERROR_ORDER_COLUMNS, "expected: {}".
                                                format(", ".join(requiredNames)))
                    adjusted_resource = True
                #reconstruct schema frictionless
                if adjusted_resource:
                    originalFields = {fieldName:resource.schema.get_field(fieldName) for fieldName in requiredNames}
                    resource.schema.clear_fields()
                    for fieldName in resourceColumnNames:
                        if fieldName and fieldName in originalFields:
                            resource.schema.add_field(originalFields[fieldName])
                        else:
                            field = fields.AnyField(name=fieldName)
                            resource.schema.add_field(field)
                    resource.schema.primary_key = [fieldName for fieldName in resource.schema.primary_key
                                                   if resource.schema.has_field(fieldName)]
                    resource.schema.foreign_keys = [entry for entry in resource.schema.foreign_keys
                                                   if any([resource.schema.has_field(fieldName) for fieldName in entry["fields"]])]
            except Exception as e:
                self._report.addReportError(reportId, Validate.ERROR_GENERAL, "problem checking missing columns: {}".format(str(e)))
        return adjusted_resource

    def _removeEmptyRowsForSheet(self, sheetName:str, resource:Resource = None, reportId:str = None):
        try:
            if not resource is None:
                headerRows = resource.dialect.header_rows
            else:
                headerRows = [1]
            if not reportId is None:
                self._report.addReportDebug(reportId, "remove empty rows from sheet '{}'".format(sheetName))
            resourceFilename = os.path.join(self._basepath, self._filename)
            if not os.access(resourceFilename, os.W_OK):
                if not (resource is None or reportId is None):
                    self._report.addReportError(reportId, Validate.ERROR_NO_WRITE_ACCESS,
                            "no access to {}, can't try to remove empty rows".format(resource.name))
            else:
                if not self._wb:
                    self._wb = load_workbook(os.path.join(resourceFilename))
                if not sheetName in self._wb.sheetnames:
                    if not reportId is None:
                        self._report.addReportError(reportId, Validate.ERROR_NO_SHEET,
                                            "can't find {} in resource".format(sheetName))
                else:
                    ws = self._wb[sheetName]
                    if not reportId is None:
                        self._report.addReportDebug(reportId,
                                            "detect {} columns and {} rows in sheet".format(ws.max_column, ws.max_row))
                    deletablIds = []
                    for row in ws.rows:
                        values = [cell for cell in row if not (cell.internal_value is None or
                                                              str(cell.internal_value).isspace())]
                        if len(values)==0:
                            for cell in row:
                                rowId = cell.row
                                break
                            if not id in headerRows:
                                deletablIds.append(rowId)
                    if len(deletablIds)>0:
                        if not reportId is None:
                            self._report.addReportWarning(reportId, Validate.WARNING_EMPTY_ROW,
                                            "removed {} rows".format(len(deletablIds))) 
                        #try to delete efficient...
                        sortedDeletableIds = sorted(deletablIds, reverse=True)
                        deletableList = []
                        for i,rowId in enumerate(sortedDeletableIds):
                            deletableList.append(rowId)
                            if ((i+1)<len(sortedDeletableIds)) and sortedDeletableIds[i+1]==(rowId-1):
                                #just continue
                                pass
                            else:
                                ws.delete_rows(deletableList[-1],len(deletableList))
                    #always save
                    self._wb.save(filename = resourceFilename)
                    if not reportId is None:
                        self._report.addReportDebug(reportId,
                                            "updated {} after removing {} empty rows".format(resource.name,len(deletablIds))) 
        except Exception as e:
            if not reportId is None:
                self._report.addReportError(reportId, Validate.ERROR_GENERAL, "problem removing empty rows","{}".format(str(e)))

    def _removeEmptyColumnsForSheet(self, sheetName:str, resource:Resource = None, reportId:str = None):
        try:
            if not reportId is None:
                self._report.addReportDebug(reportId, "remove empty columns from sheet '{}'".format(sheetName))
            resourceFilename = os.path.join(self._basepath, self._filename)
            if not os.access(resourceFilename, os.W_OK):
                self._report.addReportError(reportId, Validate.ERROR_NO_WRITE_ACCESS,
                            "no access to {}, can't try to remove empty columns".format(resource.name))
            else:
                if not self._wb:
                    self._wb = load_workbook(os.path.join(resourceFilename))
                if not sheetName in self._wb.sheetnames:
                    if not (resource is None or reportId is None):
                        self._report.addReportError(reportId, Validate.ERROR_NO_SHEET, "can't find {} in resource".format(sheetName))
                else:
                    ws = self._wb[sheetName]
                    if not reportId is None:
                        self._report.addReportDebug(reportId,
                        "detect {} columns and {} rows in sheet".format(ws.max_column, ws.max_row))
                    deletablIds = []
                    for column in ws.columns:
                        values = [cell for cell in column if not (cell.internal_value is None or
                                                                  str(cell.internal_value).isspace())]
                        if len(values)==0:
                            for cell in column:
                                columnId = cell.column
                                break
                            deletablIds.append(columnId)
                    if len(deletablIds)>0:
                        if not reportId is None:
                            self._report.addReportWarning(reportId, Validate.WARNING_EMPTY_COLUMN,
                                            "removed {} columns".format(len(deletablIds))) 
                        #try to delete efficient...
                        sortedDeletableIds = sorted(deletablIds, reverse=True)
                        deletableList = []
                        for i in range(len(sortedDeletableIds)):
                            columnId = sortedDeletableIds[i]
                            deletableList.append(columnId)
                            if ((i+1)<len(sortedDeletableIds)) and sortedDeletableIds[i+1]==(columnId-1):
                                #just continue
                                pass
                            else:
                                ws.delete_cols(deletableList[-1],len(deletableList))
                    #always save
                    self._wb.save(filename = resourceFilename)
                    if not reportId is None:
                        self._report.addReportDebug(reportId,
                                            "updated {} after removing {} empty columns".format(resource.name,len(deletablIds))) 
        except Exception as e:
            if not reportId is None:
                self._report.addReportError(reportId, Validate.ERROR_GENERAL, "problem removing empty columns: {}".format(str(e)))

    def _updateTypeForStringColumns(self, sheetName:str, resource:Resource, reportId:str):
        try:
            headerRows = resource.dialect.header_rows
            columnNames = [field.name for field in resource.schema.fields if field.name and field.type=="string"]
            if len(columnNames)>0:
                self._report.addReportDebug(reportId,
                            "convert column(s) '{}' to string in '{}'".format("', '".join(columnNames),sheetName))
                resourceFilename = os.path.join(self._basepath, self._filename)
                if not os.access(resourceFilename, os.W_OK):
                    self._report.addReportError(reportId, Validate.ERROR_NO_WRITE_ACCESS,
                                "no access to {}, can't try to convert columns to string".format(resource.name))
                else:
                    if not self._wb:
                        self._wb = load_workbook(os.path.join(resourceFilename))
                    if not sheetName in self._wb.sheetnames:
                        self._report.addReportError(reportId, Validate.ERROR_NO_SHEET,
                            "can't find {} in resource".format(sheetName))
                    else:
                        ws = self._wb[sheetName]
                        totalUpdatedNumber=0
                        for column in ws.columns:
                            if column[0].value in columnNames:
                                updatedNumber = 0
                                for row,item in enumerate(column):
                                    if row+1 in headerRows:
                                        continue
                                    if not (item.internal_value is None or isinstance(item.value,str)):
                                        item.value = str(item.value)
                                        updatedNumber+=1
                                if updatedNumber>0:
                                    self._report.addReportWarning(reportId, Validate.WARNING_TYPE,
                                        "changed type for {} entries from column '{}'".format(
                                            updatedNumber,column[0].value))
                                totalUpdatedNumber+=updatedNumber
                        #always save
                        self._wb.save(filename = resourceFilename)
                        self._report.addReportDebug(reportId,
                                                "updated {} cells to string in {}".format(
                                                totalUpdatedNumber,resource.name))
        except Exception as e:
            self._report.addReportError(reportId, Validate.ERROR_GENERAL, "problem converting sheet columns to string: {}".format(str(e)))

    def _validate(self):
        self._logger.debug("start validation")

        #get available sheets
        self._report.addReport("general","General",False,ValidationReport.TYPE_GENERAL)
        progression_bar = tqdm(total=2, disable=(not self._cli), leave=False)
        progression_bar_size = 25
        progression_bar.set_description(str("Initialise validation").ljust(progression_bar_size))
        try:
            resourceFilename = os.path.join(self._basepath, self._filename)
            self._wb = load_workbook(os.path.join(resourceFilename))
            self._availableSheets = list(self._wb.sheetnames)
        except Exception as e:
            self._report.addReportError("general", Validate.ERROR_NO_SHEET_NAMES,
                                        "problem retrieving sheetnames from '{}': {}".format(
                                            os.path.basename(self._filename),str(e)))
            return

        #configuration
        allowAdditionalSheets = self._config["settings"].get("allowAdditionalSheets", False)
        requireSheetOrder = self._config["settings"].get("requireSheetOrder", False)        

        #compute the expected sheets, check if everything is included and compute the order for validation
        self._computeExpectedSheets()
        progression_bar.reset(total=len(self._expectedSheets)+2)
        
        missingSheets = [sheetName for sheetName in self._expectedSheets if not sheetName in self._availableSheets]
        if len(missingSheets)>0:
            self._report.addReportError("general", Validate.ERROR_MISSING_SHEETS, "'{}'".format("', '".join(missingSheets)))
        additionalSheets = [sheetName for sheetName in self._availableSheets if not sheetName in self._allowedSheets]
        if len(additionalSheets)>0:
            if not allowAdditionalSheets:
                self._report.addReportError("general", Validate.ERROR_UNEXPECTED_SHEETS, "'{}'".format("', '".join(additionalSheets)))
            else:
                self._report.addReportInfo("general", "ignoring additional sheet(s): '{}'".format("', '".join(additionalSheets)))
        if requireSheetOrder:
            requiredOrder = [entry["name"] for entry in self._config["sheets"] if entry["name"] in self._availableSheets]
            availableOrder = [sheetName for sheetName in self._availableSheets if sheetName in requiredOrder]
            if not requiredOrder==availableOrder:
                self._report.addReportError("general", Validate.ERROR_ORDER_SHEETS, "'{}'".format("', '".join(availableOrder)))
        #get validation order (check dependencies)
        validationOrder = self._computeValidationOrder()
        progression_bar.reset(total=len(validationOrder)+2)
        #validate resources
        errorTypes = set()
        progression_bar.update(1)
        for entry in validationOrder:
            n = progression_bar_size - 13
            sheet = "[%s...]"%entry["name"][:(n-3)] if len(entry["name"])>n else "[%s]"%entry["name"]
            progression_bar.set_description("Validating %s" % sheet.ljust(n+2))
            resource_validation = self._validateResource(entry)
            if resource_validation:
                errorTypes.update([item.type for item in resource_validation.tasks[0].errors])
            progression_bar.update(1)            
        progression_bar.set_description("Validating package")
        progression_bar.update(1)
        #validate package
        self._validatePackage(skip_errors = list(errorTypes))
        progression_bar.close()

    def _validateResource(self, entry, skip_errors:list=None):
        reportId = "resource:{}".format(entry["resource"])
        self._report.addReport(reportId, entry["name"], True, ValidationReport.TYPE_RESOURCE)
        self._report.addReportDebug(reportId,"define resource from sheet '{}'".format(entry["name"]))

        adjustTypeForStringColumns = self._config["settings"].get("adjustTypeForStringColumns", False)
        removeEmptyRows = self._config["settings"].get("removeEmptyRows", False)
        removeEmptyColumns = self._config["settings"].get("removeEmptyColumns", False)

        dialectArguments = {}
        dialectArguments["header"] = entry.get("header",self._config["settings"].get(
            "header",settings.DEFAULT_HEADER))
        dialectArguments["header_rows"] = entry.get("headerRows",self._config["settings"].get(
            "headerRows",settings.DEFAULT_HEADER_ROWS))
        dialectArguments["header_join"] = entry.get("headerJoin",self._config["settings"].get(
            "headerJoin",settings.DEFAULT_HEADER_JOIN))
        dialectArguments["header_case"] = entry.get("headerCase",self._config["settings"].get(
            "headerCase",settings.DEFAULT_HEADER_CASE))
        dialectArguments["comment_char"] = entry.get("commentChar",self._config["settings"].get(
            "commentChar",None))
        dialectArguments["comment_rows"] = entry.get("commentRows",self._config["settings"].get(
            "commentRows",[]))
        dialectArguments["skip_blank_rows"] = entry.get("skipBlankRows",self._config["settings"].get(
            "skipBlankRows",False))
        dialect = Dialect(**dialectArguments)
        dialect.add_control(formats.ExcelControl(sheet=entry["name"], preserve_formatting=False))
        resource = Resource(basepath=self._basepath, path=self._filename, dialect=dialect)
        #set name and remove if exists
        if self._package.has_resource(entry["resource"]):
            self._package.remove_resource(entry["resource"])
        resource.name = entry["resource"]
        resource.schema = Schema.from_descriptor(entry["schema"]["data"])
        #dynamic
        if "dynamic" in entry["schema"]:
            resources = {}
            for dynamic in entry["schema"]["dynamic"]:
                dynamicResources = {}
                mappings = dynamic.get("mappings",{})
                newFields = []
                #get resources
                if "dynamicResources" in dynamic:
                    for name,dynamicResource in dynamic["dynamicResources"].items():
                        if not dynamicResource["resource"] in resources:
                            resources[dynamicResource["resource"]] = pd.DataFrame(self._package.get_resource(
                                dynamicResource["resource"]).extract().get(dynamicResource["resource"],[]))
                        resourceData = resources[dynamicResource["resource"]].copy()
                        if "condition" in dynamicResource:
                            for condition in dynamicResource["condition"]:
                                resourceData = resourceData[resourceData[condition["field"]]==condition["value"]]
                        dynamicResources[name] = resourceData
                if "linkedResources" in dynamic:
                    for name,linkedResource in dynamic["linkedResources"].items():
                        if not linkedResource["resource"] in resources:
                            resources[linkedResource["resource"]] = pd.DataFrame(self._package.get_resource(
                                linkedResource["resource"]).extract().get(linkedResource["resource"],[]))
                #create fields
                dynamicResourcesList = sorted(list(dynamicResources.keys()))
                dynamicResourcesIterators = [dynamicResources[key].iterrows() for key in dynamicResourcesList]
                linkedResources = {}
                recomputeLinkedResources = False
                for dynamicEntry in itertools.product(*dynamicResourcesIterators):
                    dynamicResourcesEntry = dict(map(lambda k,v : (k,v[1]), dynamicResourcesList,dynamicEntry))
                    if "linkedResources" in dynamic and (len(linkedResources)==0 or recomputeLinkedResources):
                        linkedResources = {}
                        for name,linkedResource in dynamic["linkedResources"].items():
                            resourceData = resources[linkedResource["resource"]].copy()
                            if ("condition" in linkedResource) and (resourceData.shape[0]>0):
                                for condition in linkedResource["condition"]:
                                    if isinstance(condition["value"], dict):
                                        recomputeLinkedResources = True
                                        value = dynamicResourcesEntry.get(
                                                condition["value"]["dynamicResource"],{}).get(condition["value"]["field"],None)
                                    else:
                                        value = condition["value"]
                                    resourceData = resourceData[resourceData[condition["field"]].values==value]
                            linkedResources[name] = resourceData
                    for fieldEntry in dynamic["fields"]:
                        fieldDescriptor = {}
                        for key in ["name","type","rdfType","title","format","example","description"]:
                            fieldDescriptor = setDescriptorValueDynamicString(
                                key,fieldDescriptor,fieldEntry,
                                dynamicResourcesEntry,mappings,linkedResources,self._logger)
                        if "constraints" in fieldEntry:
                            fieldDescriptor["constraints"] = {}
                            for key in ["enum"]:
                                fieldDescriptor["constraints"] = setDescriptorValueDynamicList(
                                 key,fieldDescriptor["constraints"],fieldEntry["constraints"],
                                    dynamicResourcesEntry,mappings,linkedResources,self._logger)
                            for key in ["required"]:
                                fieldDescriptor["constraints"] = setDescriptorValueDynamicBoolean(
                                 key,fieldDescriptor["constraints"],fieldEntry["constraints"],
                                    dynamicResourcesEntry,mappings,linkedResources,self._logger)
                            for key in ["minLength","maxLength"]:
                                fieldDescriptor["constraints"] = setDescriptorValueDynamicInteger(
                                 key,fieldDescriptor["constraints"],fieldEntry["constraints"],
                                    dynamicResourcesEntry,mappings,linkedResources,self._logger)
                            for key in ["minimum","maximum"]:
                                fieldDescriptor["constraints"] = setDescriptorValueDynamicNumber(
                                 key,fieldDescriptor["constraints"],fieldEntry["constraints"],
                                    dynamicResourcesEntry,mappings,linkedResources,self._logger)
                            for key in ["pattern"]:
                                fieldDescriptor["constraints"] = setDescriptorValueDynamicString(
                                 key,fieldDescriptor["constraints"],fieldEntry["constraints"],
                                    dynamicResourcesEntry,mappings,linkedResources,self._logger)
                        newFields.append(Field.from_descriptor(fieldDescriptor))
                #update schema
                position = dynamic.get("position","after")
                field = dynamic.get("field",None)
                if field is None:
                    if position=="before":
                        for i,newField in enumerate(newFields):
                            resource.schema.add_field(newField,position=1+i)
                    elif position=="after":
                        for i,newField in enumerate(newFields):
                            resource.schema.add_field(newField)
                elif resource.schema.has_field(field):
                    pos = min([i+1 for i,schemaField in enumerate(resource.schema.fields)
                                    if schemaField.name==field])
                    if position=="before":
                        for i,newField in enumerate(newFields):
                            resource.schema.add_field(newField, position=pos+i)
                    elif position=="after":
                        for i,newField in enumerate(newFields):
                            resource.schema.add_field(newField, position=pos+i+1)

        #check types, empty rows and columns
        pick_errors = []
        pick_errors += ["type-error"] if adjustTypeForStringColumns else []
        pick_errors += ["blank-row"] if removeEmptyRows else []
        pick_errors += ["extra-label"] if removeEmptyColumns else []
        if len(pick_errors)>0:
            resource_validation = resource.validate(checklist=Checklist(pick_errors=pick_errors, skip_errors=skip_errors))
            if not resource_validation.valid:
                errorTypes = set([item.type for item in resource_validation.tasks[0].errors])
                if adjustTypeForStringColumns and "type-error" in errorTypes:
                    self._updateTypeForStringColumns(entry["name"],resource,reportId)
                if removeEmptyRows and "blank-row" in errorTypes:
                    self._removeEmptyRowsForSheet(entry["name"],resource,reportId)
                if removeEmptyColumns and "extra-label" in errorTypes:
                    self._removeEmptyColumnsForSheet(entry["name"],resource,reportId)
        #validate
        checklist = Checklist.from_descriptor(entry["checklist"]["data"]) if "checklist" in entry else Checklist()
        checklist.skip_errors = skip_errors
        resource_validation = resource.validate(checklist=checklist)
        self._report.setFrictionless(reportId, resource_validation)
        self._report.addReportDebug(reportId, resource_validation.stats)
        if resource_validation.valid and skip_errors is None:
            self._report.addReportInfo(reportId,"succesfull frictionless validation '{}' sheet".format(entry["name"]))
        #check missing columns and possibly revalidate
        elif skip_errors is None and not resource_validation.valid:
            if self._checkMissingColumns(entry["name"],resource,reportId):
                resource_validation = resource.validate(checklist=checklist)
                self._report.setFrictionless(reportId, resource_validation)
                self._report.addReportDebug(reportId, resource_validation.stats)
        #try to add resource
        self._package.add_resource(resource)                
        if "modules" in entry:
            for module in entry["modules"]:
                try:
                    moduleObject = __import__(module["name"])
                    validationModule = moduleObject.ValidationModule(self._package,resource.name)
                except ModuleNotFoundError:
                    self._report.addReportWarning(reportId, Validate.WARNING_MODULE, 
                                                  "could not import module %s" % str(module["name"]))
        #return validation
        return resource_validation

    def _validatePackage(self, skip_errors:list=None):
        reportId = "package"
        self._report.addReport(reportId, "Package", True, ValidationReport.TYPE_PACKAGE)
        #validate
        if len(self._package.resources)>0:
            package_validation = self._package.validate(checklist=Checklist(skip_errors=skip_errors))
            self._report.setFrictionless(reportId, package_validation)
            self._report.addReportDebug(reportId, package_validation.stats)
        else:
            self._report.addReportWarning(reportId, Validate.WARNING_EMPTY_PACKAGE, "no sheets found to be validated")

    def createPackageJSON(self, filename:str = None):
        """
        Create frictionless package data
        """
        if not filename is None:
            try:
                return self._package.to_json(filename)
            except:
                self._logger.error("can't store package to '%s'",filename)
                return self._package.to_json()
        else:
            return self._package.to_json()

    def createReport(self):
        """
        Create JSON object with report
        """
        reportObject = {"name": os.path.basename(self._name),
                        "version": __version__, 
                        "valid": self.valid, 
                        "reports": self._report.createReportObjects()}
        return reportObject

    def createTextReport(self, filename:str=None, textWidth=100, examples=3, warnings=True):
        """
        Create text version report
        """
        #create text
        name = os.path.basename(self._name)
        reportText = "=== {} '{}' (version {}) ===\n".format(
            ("VALID" if self._report.valid else "INVALID"),name,__version__)
        reportText = reportText + self._report.createTextReport(textWidth, examples, False, warnings)
        return reportText

    def createMarkdownReport(self, filename:str=None, textWidth=100, examples=3, warnings=True):
        """
        Create markdown version report
        """
        #create text
        name = os.path.basename(self._name)
        reportText = "# {}: {}\n\nVersion {}\n\n".format(
            ("VALID" if self._report.valid else "INVALID"),name,__version__)
        reportText = reportText + self._report.createTextReport(textWidth, examples, True, warnings)
        return reportText

    @property
    def valid(self):
        """
        Valid
        """
        return self._report.valid

    @property
    def reports(self):
        """
        Reports
        """
        return self._report.reports

    def __getitem__(self, key):
        return self._report[key]
