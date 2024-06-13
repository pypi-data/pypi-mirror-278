import ast
from io import BytesIO
from math import log10
import matplotlib.pyplot as plt
import matplotlib.ticker as mticker
import numpy as np
from sklearn.decomposition import PCA
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import (
    Alignment, 
    Border, 
    Font, 
    PatternFill,
    Side
)
import os
import pandas as pd
from PIL import Image as PILImage, ImageDraw
from process_twarc.corpus_analysis.tabulate import df_to_excel_with_styling
from process_twarc.util import get_embedding, make_dir, get_files, concat_dataset
import seaborn as sns
from transformers import AutoModel, AutoTokenizer

def  plot_tSNE(
        path_to_model: str, 
        vocab_dict: dict,
        embedding_method: str,
        colors: list=plt.get_cmap('tab10')(np.linspace(0, 1, 10)),
        background: bool=True, 
        title: str="",
        show_plot: bool=True,
        path_to_output_plot: str="",
        path_to_output_excel: str="",
):
    """Provided with a dictionary of target vocab and a model, this function
    produces a t-SNE plot of the embeddings of the tokens in the vocabulary dictionary.
    

    Args:
        path_to_model (str): Path to the model to use for generating embeddings.
        vocab_dict (dict): A dictionary where the keys are the categories of the tokens and the values are dictionaries with the following keys:
            - tokens: A list of tokens in the category.
            - table_values: A list of values to display in the table. If not provided, the tokens will be used.
        embedding_method (str): The method to use for generating embeddings. Options are 'token' and CLS.
        colors (list): A list of colors to use for the different categories. Defaults to the 'tab10' colormap.
        background (bool, optional): Whether to set the background color to a very light grey. Defaults to False.
        title (str, optional): The title of the plot. Defaults to an empty string.
        show_plot (bool): Whether to display the plot. Defaults to True.
        path_to_output_plot (str, optional): The path to save the plot, (foo.png)
        path_to_output_excel (str, optional): The path to save the excel file, (foo.xlsx)
    """
    model = AutoModel.from_pretrained(path_to_model)
    tokenizer = AutoTokenizer.from_pretrained(path_to_model)

    def get_embeddings(vocab_dict):
        for _, tokens_info in vocab_dict.items():
            tokens_info["embeddings"] = [get_embedding(token, tokenizer, model, embedding_method) for token in tokens_info["tokens"]]
        return vocab_dict

    vocab_dict = get_embeddings(vocab_dict)
    all_embeddings = np.concatenate([np.array(tokens_info["embeddings"]) for tokens_info in vocab_dict.values()])

    pca = PCA(n_components=2)
    embeddings_2d = pca.fit_transform(all_embeddings)

    plt.figure(figsize=(12, 10))
    ax = plt.gca()  # Get the current Axes instance
    
    if background:
        ax.set_facecolor('#f0f0f0')  # Set the background color to a very light grey

    offset = 0

    for (category, tokens_info), color in zip(vocab_dict.items(), colors):
        embeddings = tokens_info["embeddings"]
        (alpha, color) = (1.0, color) 
        plt.scatter(embeddings_2d[offset:offset+len(embeddings), 0], embeddings_2d[offset:offset+len(embeddings), 1], alpha=alpha, color=color, edgecolors='black', s=100, label=category)
        offset += len(embeddings)

    ax.set_axisbelow(True)  # Ensure gridlines are behind the plot points
    ax.grid(color='white', linestyle='-', linewidth=0.7)  # Add white gridlines

    # Remove axis numbers but keep the tick marks
    ax.tick_params(axis='both', which='both', length=6, labelleft=False, labelbottom=False)

    # Set the title if provided, with a larger font size
    if title:
        plt.title(title, fontsize=14)  # Adjust fontsize as needed

    plt.legend()
    if path_to_output_plot:
        os.makedirs(os.path.dirname(path_to_output_plot), exist_ok=True)
        plt.savefig(path_to_output_plot)
    if show_plot:
        plt.show()

    if path_to_output_excel:
        def dict_to_dataframe(vocab_dict):
            result = {}
            max_length = 0
            
            for value in vocab_dict.values():
                current_length = len(value.get('table_values', value['tokens']))
                if current_length > max_length:
                    max_length = current_length
            
            # Loop through each key-value pair in the input dictionary and standardize the length
            for key, value in vocab_dict.items():
                chosen_list = value.get('table_values', value['tokens'])
                
                length_difference = max_length - len(chosen_list)
                result[key] = chosen_list + [None] * length_difference
            
            df = pd.DataFrame.from_dict(result, orient='index').transpose()
            
            return df
        df = dict_to_dataframe(vocab_dict)
        make_dir(path_to_output_excel)
        df_to_excel_with_styling(df, path_to_output_excel)


def plot_rank_frequency_distribution(
        data_dict: dict,
        log_scale: bool=True,
        title: str="Rank-Frequency Plot",
        xlabel: str="",
        ylabel: str="",
        xlim: tuple=tuple(),
        ylim: tuple=tuple(),
        path_to_output: str=""):
    plt.figure(figsize=(10, 6))
    ax = plt.gca()
    ax.set_facecolor('#f0f0f0')
    """The purpose of these plot is to analyze distributions againnst Zipfian predictions.

    Configured to plot rank-freuqncy of tokens or characters.

    Args:
    """

    # List of marker styles for different plots
    markers = ['o', 's', '^', 'd', '*', 'x', '+', 'v', '<', '>']  # You can add more if needed

    for (name, data), marker in zip(data_dict.items(), markers):
        df = data["data"]
        label = data.get("label", name)
        do_best_fit = data.get("do_best_fit", False)
        best_fit_label = data.get("best_fit_label", f'{name} best-fit line')
        if log_scale:
            plt.loglog(df['rank'], df['frequency'], marker=marker, linestyle='-', label=label, markersize=1)  # Smaller markers

            if do_best_fit:
                log_rank = np.log(df['rank'])
                log_freq = np.log(df['frequency'])
                slope, intercept = np.polyfit(log_rank, log_freq, 1)
                best_fit_line = np.exp(intercept) * df['rank']**(slope)
                plt.loglog(df['rank'], best_fit_line, label=best_fit_label, linestyle='--')
        else:
            plt.plot(df['rank'], df['frequency'], marker=marker, linestyle='-', label=label, markersize=1)  # Smaller markers

            if do_best_fit:
                slope, intercept = np.polyfit(df['rank'], df['frequency'], 1)
                best_fit_line = intercept + slope * df['rank']
                plt.plot(df['rank'], best_fit_line, label=best_fit_label, linestyle='--')

    if not xlabel:
        xlabel = "Rank (log)" if log_scale else "Rank (linear)"
    if not ylabel:
        ylabel = "Frequency (log)" if log_scale else "Frequency (linear)"
    plt.title(title)
    plt.xlabel(xlabel)
    plt.ylabel(ylabel)
    plt.grid(True, which="both", ls="--", color='white')
    plt.legend()

    # Setting x and y limits if specified
    if xlim:
        plt.xlim(xlim)
    if ylim:
        plt.ylim(ylim)

    if path_to_output:
        make_dir
        os.makedirs(os.path.dirname(path_to_output))
        plt.savefig(path_to_output)
    plt.show()

def plot_pos_confusion_matrix(
        path_to_vocab: str,
        path_to_UPOS_data: str,
        path_to_JPOS_data: str,
        predicted_tags: str, 
        observed_tags: str, 
        tok_choice: str, 
        omit_subwords: bool=False, 
        category: str="all",
        path_to_output: str=""
        ):
    
    def _load():
        vocab = pd.read_csv(path_to_vocab)
        merge_ = lambda path, pos: vocab.merge(pd.read_csv(path), on=["token", pos], how="left")
        vocab = merge_(path_to_UPOS_data, "UPOS")
        vocab = merge_(path_to_JPOS_data, "JPOS")
        vocab.dropna(subset = ["token", predicted_tags, observed_tags], inplace=True)
        vocab = vocab[vocab["token"].apply(lambda x: len(x) > 0)]
        if omit_subwords:
            vocab = vocab[~vocab["token"].str.contains("##")]

        for column in list(set([predicted_tags, observed_tags])):
            vocab[column] = vocab[column].apply(lambda x: ast.literal_eval(x))
            vocab = vocab[vocab[column].apply(lambda x: len(x) > 0)]
        
        vocab["predictions"] = vocab.apply(lambda x: x[predicted_tags][0], axis=1)
        return vocab
        
    vocab = _load()
    
    # Ensure the correct mapping of predicted and true tags
    pos_tags = {
        "UPOS": ["ADJ", "ADP", "ADV", "AUX", "CCONJ", "DET", "INTJ", "NOUN", "NUM", "PART", "PRON", "PROPN", "PUNCT", "SCONJ", "SYM", "VERB"],
        "JPOS": ["補助記号", "感動詞", "助動詞", "名詞", "助詞", "接尾辞", "接続詞", "動詞", "副詞", "形容詞", "接頭辞", "代名詞", "形状詞", "連体詞", "記号"]
    }

    jpos_to_eng = {
    "補助記号": "PUNCT",  # Punctuation
    "感動詞": "INTJ",  # Interjection
    "助動詞": "AUX",  # Auxiliary verb
    "名詞": "NOUN",  # Noun
    "助詞": "PART",  # Adposition (preposition and postposition)
    "接尾辞": "SUFF",  # Suffix
    "接続詞": "CONJ",  # conjunction
    "動詞": "VERB",  # Verb
    "副詞": "ADV",  # Adverb
    "形容詞": "ADJ",  # Adjective
    "接頭辞": "PREFIX",  # Prefix
    "代名詞": "PRON",  # Pronoun
    "形状詞": "ADJ-NOUN",  # Adjectival noun
    "連体詞": "DET",  # Determiner
    "記号": "SYM",  # Symbol
}
    lexical = ["ADJ", "ADV", "NOUN", "PROPN", "VERB", "INTJ", "感動詞", "名詞", "接尾辞", "動詞", "副詞", "形容詞", "接頭辞", "形状詞"]
    functional = ["ADP", "AUX", "CCONJ", "DET", "PART", "PRON", "SCONJ", "助詞", "接続詞", "連体詞","代名詞", "助動詞"]

    vocab = vocab[vocab[tok_choice].fillna(False)]

    token_counts = vocab[f"{tok_choice}_count"]
    prob_values = vocab[pos_tags[observed_tags]].multiply(token_counts, axis=0)

    observed_labels = sorted(pos_tags[observed_tags])  # Use the full list of possible true tags
    if observed_tags == "JPOS":
        observed_labels = sorted(pos_tags[observed_tags], key=lambda x: jpos_to_eng[x])
    pred_labels = sorted(pos_tags[predicted_tags])  # Use the full list of possible predicted tags
    if predicted_tags == "JPOS":
        pred_labels = sorted(pos_tags[predicted_tags], key=lambda x: jpos_to_eng[x])

    if category == 'lexical':
        observed_labels = [tag for tag in observed_labels if tag in lexical]
        pred_labels = [tag for tag in pred_labels if tag in lexical]
    elif category == 'functional':
        observed_labels = [tag for tag in observed_labels if tag in functional]
        pred_labels = [tag for tag in pred_labels if tag in functional]


    # Initialize confusion matrix dimensions correctly
    nub_observed_tags = len(observed_labels)
    num_pred_tags = len(pred_labels)
    cm = np.zeros((num_pred_tags, nub_observed_tags))  # Swap dimensions to match indexing

    # Updated indexing mapping
    observed2idx = {tag: i for i, tag in enumerate(observed_labels)}
    pred2idx = {tag: i for i, tag in enumerate(pred_labels)}

    for _, row in vocab.iterrows():
        if row["predictions"] in pred2idx:  # Check if the predicted tag exists in the mapping
            pred_idx = pred2idx[row["predictions"]]
            for tag, prob in row[prob_values.columns].items():
                if tag in observed2idx:  # Ensure the true tag is in the mapping
                    observed_idx = observed2idx[tag]
                    cm[pred_idx, observed_idx] += prob


    # Normalize the confusion matrix, avoiding division by zero
    cm_sum = cm.sum(axis=1, keepdims=True)
    cm_normalized = cm / np.where(cm_sum == 0, 1, cm_sum)  # Avoid division by zero

    if observed_tags == "JPOS":
        observed_labels = [jpos_to_eng[tag] for tag in observed_labels]
    if predicted_tags == "JPOS":
        pred_labels = [jpos_to_eng[tag] for tag in pred_labels]
    plt.figure(figsize=(12, 10))
    sns.heatmap(cm_normalized, annot=True, fmt=".2f", cmap="YlGnBu", xticklabels=observed_labels, yticklabels=pred_labels)
    plt.title(f'Predicted ({predicted_tags}) vs. Observed ({observed_tags}) Tags Confusion Matrix')
    plt.ylabel(f'Predicted Tags ({predicted_tags})')
    plt.xlabel(f'Observed Tags ({observed_tags})')
    plt.show()
    if path_to_output:
        plt.savefig(path_to_output)

def plot_polar_datetime_histogram(
        arr: np.ndarray, 
        title: str= "Relative Contribution per Time Unit",
        path_to_output: str=""
):
    """Given an array of hours or days of the week, this function creates a 
    polar histogram to show the relative contribution of each time unit.
    
    
    Args:
        arr (np.ndarray): An array of hours or days of the week.
        title (str): The title of the plot. Defaults to "Relative Contribution per Time Unit".
        path_to_output (str): The path to save the plot, (foo.png)
    """
    # Define a fixed order for the days of the week
    days_order = ["Sunday", "Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday"]
    
    # Check if the array contains strings (e.g., days of the week)
    if arr.dtype.type is np.str_:
        # Map days of the week to integers based on the fixed order
        day_to_num = {day: i for i, day in enumerate(days_order)}
        mapped_arr = np.vectorize(day_to_num.get)(arr)
        n_unique = len(days_order)
        xticklabels = days_order
    else:
        # Use the array as is for hours
        mapped_arr = arr
        n_unique = len(np.unique(arr))
        xticklabels = range(n_unique)
    
    hist, bins = np.histogram(mapped_arr, bins=np.arange(n_unique+1), density=True)

    # Compute the width of each bin in radians
    width = 2 * np.pi / n_unique

    # Create a polar subplot
    fig, ax = plt.subplots(subplot_kw={'polar': True})

    ax.grid(True, axis="y", color="red", linestyle='-', linewidth=0.5, alpha=0.25)
    ax.set_facecolor('#f0f0f0')

    bars = ax.bar(bins[:-1] * np.pi / (n_unique/2), hist, width=width, bottom=0.0, edgecolor='white', color = "#7BD3EA", linewidth=0.5)

    ax.set_theta_zero_location('N')
    ax.set_theta_direction(-1)

    ax.set_xticks(np.pi/(n_unique/2) * np.linspace(0, n_unique-1, n_unique))
    ax.set_xticklabels(xticklabels, fontsize=11)

    yticks = ax.get_yticks()
    ax.set_yticklabels(['{:,.0%}'.format(y) for y in yticks][:-2], fontsize=11)
    ax.xaxis.grid(False)

    plt.title(title)
    if path_to_output:
        make_dir(os.path.dirname(path_to_output))
        plt.savefig(path_to_output)
    plt.show()

def plot_histogram_of_months(
        arr: np.ndarray,
        title: str = "Histogram of Corpus Contribution by Month",
        xlabel: str = "Month, Year",
        ylabel: str = "Corpus Contribution (Millions of Tweets)",
        ylabel2: str = "Percentage (%)",
        path_to_output: str=""

):
    # Convert to Pandas Series and then to datetime
    months_series = pd.Series(arr)
    months_series = pd.to_datetime(months_series, format='%B %Y')

    # Count occurrences and sort
    months_counts = months_series.value_counts().sort_index()

    # Plotting
    plt.figure(figsize=(10, 6))
    ax = plt.subplot(111)  # Explicitly create axes to modify
    ax.set_facecolor('#f0f0f0')  # Set the face color to light grey

    # Convert dates to "Month Year" format for x-tick labels
    months_counts.index = months_counts.index.strftime('%B %Y')
    months_counts.plot(kind='bar', ax=ax, color='#7BD3EA', edgecolor='black', width = 0.9)

    # Gridlines
    ax.grid(color='red', axis='y',linestyle='-', linewidth=0.5, alpha = 0.25)

    # Axis labels and title
    plt.xlabel(xlabel)
    plt.ylabel(ylabel)
    plt.title(title)
    plt.xticks(rotation=45, ha='right')

    # Secondary Y-axis for relative values
    ax2 = ax.twinx()
    total = months_counts.sum()
    relative_counts = months_counts.apply(lambda x: x / total * 100)
    ax2.plot(ax.get_xticks(), relative_counts, color="r", marker=None, linestyle='None')
    ax2.set_ylabel(ylabel2)

    # Use a formatter to specify the y-axis labels on the right side
    ax2.yaxis.set_major_formatter(mticker.PercentFormatter())

    plt.tight_layout()
    if path_to_output:
        make_dir(os.path.dirname(path_to_output))
        plt.savefig(path_to_output)
    plt.show()

def preprocess_data_for_histogram_of_users(
    preprocessed_dir: str,
    path_to_users_table: str,
    masks: list,
    count_unknown_users: bool = False,
    log_segments: int = 1,
    count_target: str = "tweets",
    bucket_column: str = "count"
):

    def load_dataset():
        df =  concat_dataset(
            file_paths=get_files(preprocessed_dir),
            columns=["tweet_id", "author_id"],
            masks=masks
        )
        unknown_users = len(df[df["author_id"].isna()])
        df = df.groupby("author_id").agg({"tweet_id": "count"}).reset_index()
        df.rename(columns={"tweet_id": "count"}, inplace=True)
        return df, unknown_users
    
    def load_user_data():
        users = pd.read_csv(path_to_users_table)
        users["user_id"] = users["user_id"].astype(str)
        return users
    
    # def combine_last_rows(users_agg: pd.DataFrame ):
    #         # Ensure there are at least two rows
    #         if len(users_agg) > 1:
    #             # Find the indices for the last and penultimate rows
    #             #check if the last row is very small, if so, combine it with the penultimate row
    #             if users_agg.at[users_agg.index[-1], "count"] < 0.01 * users_agg["count"].sum():
    #                 penultimate_idx = users_agg.index[-2]
    #                 last_idx = users_agg.index[-1]
    #                 for col in users_agg.columns:
    #                     if col not in ["count", ("group", "min"), ("group", "observed_min")]:
    #                         users_agg.at[penultimate_idx, col] = users_agg.at[last_idx, col]
                    
    #                 # Sum the 'count_column' values of the penultimate and last row, update the penultimate row
    #                 users_agg.loc[penultimate_idx, "count"] = users_agg.loc[penultimate_idx, "count"] + users_agg.loc[last_idx, "count"]
                    
    #                 # Drop the last row
    #                 users_agg = users_agg.drop(last_idx)
    #         return users_agg
    
    def preprocess():
            df, unknown_users = load_dataset()
            users = load_user_data()
            
            merged = users.merge(df, left_on="user_id", right_on="author_id", how="left")
            merged = merged.dropna()
            merged[("group", "level")] = merged[bucket_column].apply(lambda x: -1 if x == 1 else int(log10(x-1)*log_segments))
            merged[("group", "min")] = merged[("group", "level")].apply(lambda x: (int(10**(x/log_segments))+1) if x != -1 else 0)
            merged[("group", "max")] = merged[("group", "level")].apply(lambda x: (int(10**(((x+1)/log_segments)))) if x != -1 else 1)
            merged[("group", "observed_min")] = merged.groupby(("group", "level"))[bucket_column].transform("min").astype(int)
            merged[("group", "observed_max")] = merged.groupby(("group", "level"))[bucket_column].transform("max").astype(int)

            count_method = "count" if count_target == "users" else "sum"
            users_agg = merged.groupby([
                ("group", "level"), 
                ("group", "min"), 
                ("group", "max"), 
                ("group", "observed_min"), 
                ("group", "observed_max")
                ]).agg({"count": count_method}).reset_index()
            
            # users_agg = combine_last_rows(users_agg)

            if count_unknown_users:
                unknown = pd.DataFrame({
                        ("group", "level"): [-1],
                        ("group", "min"): [0],
                        ("group", "max"): [1],
                        ("group", "observed_min"): [0],
                        ("group", "observed_max"): [1],
                        ("count"): [unknown_users]
                    })
                users_agg = pd.concat([users_agg, unknown], ignore_index=True)
                
            users_agg[("group", "level")] = users_agg.index+1
            users_agg.rename(columns={"count": ("absolute", "sum")}, inplace=True)
            users_agg[("absolute", "cum_sum")] = users_agg[("absolute", "sum")].cumsum()
            users_agg[("relative", "sum")] = users_agg[("absolute", "sum")] / users_agg[("absolute", "sum")].sum()
            users_agg[("relative", "cum_sum")] = users_agg[("absolute", "cum_sum")] / users_agg[("absolute", "sum")].sum()

            return users_agg
                    
    return preprocess()

def users_agg_to_excel(
        users_agg:pd.DataFrame,
        colors,
        path_to_output_table: str="test.xlsx"):
    
    def preprocess_data(users_agg=users_agg, colors=colors):
        df = users_agg.copy()
        df[("group", "color")] = [tuple(int(c*255) for c in color[:3]) for color in colors]
        df[("range", "min-max")] = df[("group", "observed_min")].astype(str) + " - " + df[("group", "observed_max")].astype(str)
        if df.at[df.index[-1], ("range", "min-max")] == "0 - 1":
            df.at[df.index[-1], ("range", "min-max")] = "Unknown"


        df = df[[
            ("group", "color"),
            ("group", "level"),
            ("range", "min-max"),
            ("absolute", "sum"),
            ("absolute", "cum_sum"),
            ("relative", "sum"),
            ("relative", "cum_sum")
        ]]

        def format_number(x):
            if x < 1000:
                return str(x)
            elif x < 1e6:
                return '{:.0f}k'.format(x/1000)
            else:
                return '{:.1f}M'.format(x/1e6)
        for column in [("absolute", "sum"), ("absolute", "cum_sum"), ("relative", "sum"), ("relative", "cum_sum")]:
            if column[0] == "absolute":
                df[column] = df[column].apply(format_number)
            else:
                df[column] = df[column].apply(lambda x: '{:.2f}%'.format(x*100))

        df = df.rename(columns = {
            ("group", "color"): ("", ""),
            ("group", "level"): ("", ""),
            ("range", "min-max"): ("Range", "Min - Max"),
            ("absolute", "sum"): ("Absolute", "Sum"),
            ("absolute", "cum_sum"): ("Absolute", "Cum. Sum"),
            ("relative", "sum"): ("Relative", "Sum"),
            ("relative", "cum_sum"): ("Relative", "Cum. Sum")})

        df.columns = pd.MultiIndex.from_tuples(df.columns)
        return df
    
    def init_workbook(path_to_output_table=path_to_output_table):
        df = preprocess_data()
        make_dir(os.path.dirname(path_to_output_table))
        with pd.ExcelWriter(path_to_output_table, engine='openpyxl') as writer:
            df.to_excel(writer)
        
        wb = load_workbook(path_to_output_table)
        ws = wb.active
        return wb, ws
    
    def draw_triangles(
            ws, 
            colors,
            size=(20, 20),
            img_col=3):
        
        for i, color in enumerate(colors):
            # Convert color to RGB
            rgb_color = tuple(int(c * 255) for c in color[:3])

            # Create a new image with a white background
            shape_img = PILImage.new("RGB", size, "white")
            draw = ImageDraw.Draw(shape_img)

            # Coordinates for the vertices of the right triangle
            triangle_vertices = [(0, size[1] - 1), (size[0] - 1, size[1] - 1), (0, 0)]

            # Draw the right triangle with the right angle at the bottom left
            draw.polygon(triangle_vertices, fill=rgb_color, outline=rgb_color)

            # Convert the PIL image to a byte array
            img_byte_arr = BytesIO()
            shape_img.save(img_byte_arr, format='PNG')
            img_byte_arr.seek(0)  # Rewind to the beginning of the byte array

            # Convert byte array to an openpyxl Image
            img = openpyxl.drawing.image.Image(img_byte_arr)

            # Clear the cell contents before inserting the image
            # Adjust the row index as necessary
            cell_coordinate = ws.cell(row=i + 3, column=img_col).coordinate
            ws[cell_coordinate] = ""  # Clear the text in the cell

            # Insert the image into the cleared cell
            ws.add_image(img, cell_coordinate)
        return ws

    def resize_columns(ws, dimensions_dict= {'A': 3, 'B': 3, 'C': 4}):
        for col, width in dimensions_dict.items():
            ws.column_dimensions[col].width = width
        return ws

    # Define a function to convert RGB to Hex (since RGBColor isn't directly used in openpyxl for colors)
    def rgb_to_hex(rgb_color):
        return f"{rgb_color[0]:02X}{rgb_color[1]:02X}{rgb_color[2]:02X}"

    # Define a function to create borders with specified colors
    def create_colored_border(rgb_color):
        hex_color = rgb_to_hex(rgb_color)
        return Border(bottom=Side(border_style="thick", color=hex_color))
    
    # Function to create a border with only the bottom line
    def bottom_border_only():
        return Border(bottom=Side(style='thin'))

    def apply_colored_fills_and_borders(ws, colors):
        for i, color in enumerate(colors):
            # Create the image as before...

            # Apply the colored bottom border to each cell in the row
            rgb_color = tuple(int(c * 255) for c in color[:3])
            hex_color = rgb_to_hex(rgb_color)
            fill_color = PatternFill(start_color='FF' + hex_color, end_color='FF' + hex_color, fill_type="solid")
            colored_border = create_colored_border(rgb_color)
            for col in range(2, ws.max_column + 1):

                cell = ws.cell(row=i + 4, column=col)  # Adjust row index as necessary

                if col == 2:
                # Apply color fill to the entire cell in column A
                    cell.fill = fill_color
                cell.border = colored_border
        return ws
    
    def format_header_rows(ws):
        navy_fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
        white_bold_font = Font(bold=True, color='FFFFFF')
        for row in range(1, 3):
            for col in ws.iter_cols(min_row=row, max_row=row, min_col=1, max_col=ws.max_column):
                for cell in col:
                    if row == 1:
                        # Remove all borders from cells in the first row
                        cell.border = Border()
                        # Set navy background and bold white text for the first row
                        if cell.column != 1:
                            cell.fill = navy_fill
                        
                            cell.font = white_bold_font
                    elif row == 2:
                        # Remove all borders then apply bottom border only for the second row
                        cell.border = bottom_border_only()
                        # Also apply navy background and bold white text for the second row
                        if cell.column != 1:
                            cell.fill = navy_fill
                            cell.font = white_bold_font
        return ws
    
    def reformat_text(ws, header_rows = [1]):
        for row in range(1, ws.max_row + 1):
            get_value = lambda column: str(ws.cell(row=row, column=column).value)

            def set_value(column, value):
                ws.cell(row=row, column=column).value = value
            def set_border(column, border):
                ws.cell(row=row, column=column).border = border
            
            set_value(1, "")
            set_value(2, "")
            set_border(1, Border())
            

        return ws

    def format_allignment(
            ws,
            col2allignment = {k:v for k, v in zip(range(3, 9), ["center"]*6)},
            start_from_row = 2
    ):

        for row in range(start_from_row, ws.max_row + 1):
            for col, allignment in col2allignment.items():
                ws.cell(row=row, column=col).alignment = Alignment(horizontal=allignment)
        return ws

    def format_process():
        wb, ws = init_workbook()
        ws = reformat_text(ws)
        ws = apply_colored_fills_and_borders(ws, colors)
        ws = format_header_rows(ws)
        ws = resize_columns(ws)

        ws = format_allignment(ws)
        ws.delete_rows(3)
        ws = draw_triangles(ws, colors)

        return wb
    
    wb = format_process()
    wb.save(path_to_output_table)


def plot_histogram_of_users_by_tweet_frequency(
        users_agg: pd.DataFrame,
        title: str = "Users Aggregated by Total Tweet Contributions",
        xlabel: str = "Number of Users",
        ylabel: str = "Range of Total Tweet Contribution",
        relative: bool = False,
        cumulative: bool = False,
        figure_size: tuple = (10, 3),
        path_to_output: str=""
):  
    
    unknown_users = True if users_agg.at[users_agg.index[-1], ("group", "min")] == 0 else False

    def remove_unknown_users(
            users_agg=users_agg,
            unknown_users=unknown_users
            ):
        if unknown_users:
            users_agg = users_agg.iloc[:-1]
        return users_agg
    
    def get_barh_values(users_agg):
        tup0 = "relative" if relative else "absolute"
        tup1 = "cum_sum" if cumulative else "sum"
        barh_values = users_agg[(tup0, tup1)].values
        yticks = np.arange(len(barh_values))
        return barh_values, yticks
    
    def get_yticklabels(users_agg):
        values = users_agg[("group", "max")].tolist()
        max_val = int(max(users_agg[("group", "observed_max")].tolist()))
        yticklabels = [min([value, max_val]) for value in values]
        yticklabels = [value if (value == 1 or value %10 ==0 or value == max_val) else None for value in yticklabels]
        return yticklabels

    def get_bars(ax):
        return [ax.barh(yticks, barh_values, height=1.0, edgecolor="black", color=colors)]
    
    def get_colors(users_agg):
        cmap = plt.get_cmap('winter')
        return cmap(np.linspace(0, 1, len(users_agg))[-1::-1])

    def preprocess_plot(users_agg):
        barh_values, yticks = get_barh_values(users_agg)
        yticklabels = get_yticklabels(users_agg)
        colors = get_colors(users_agg)
        return barh_values, yticks, yticklabels, colors
    
    def format_xticklabels(ax):
        if relative:
            ax.set_xticklabels(['{:.0f}%'.format(x*100) for x in ax.get_xticks()])
        else:
            ax.set_xticklabels(['{:.0f}k'.format(x/1000) for x in ax.get_xticks()])

    users_agg = remove_unknown_users()
    barh_values, yticks, yticklabels, colors = preprocess_plot(users_agg)

    fig, ax = plt.subplots(figsize=figure_size)
    get_bars(ax)
    ax.set_xlabel(xlabel)
    ax.set_ylabel(ylabel)
    ax.set_title(title)

    # Adjusting for the use of ax object
    ax.set_yticks(yticks + 0.5)
    ax.set_yticklabels(yticklabels)
    format_xticklabels(ax)

    ax.grid(axis='x', linestyle='--', color="red", alpha=0.25)

    plt.tight_layout()
    ax.set_facecolor("#f0f0f0")
    if path_to_output:
        make_dir(os.path.dirname(path_to_output))
        plt.savefig(path_to_output)
    plt.show()

    return users_agg, colors


def histogram_of_users_pipeline(
  preprocessed_dir: str,
  path_to_users_table: str,
  bucket_column: str,
  path_to_output_plot: str,
  path_to_output_table: str,
  plot_relative: bool=False,
  plot_cumulative: bool=False,
  figure_size: tuple=(10, 6),
  title: str="",
  xlabel: str="",
  ylabel: str="",
  log_segments: int=1,
  masks: list=[]      
):
    if bucket_column not in ["count", "sample_tweet_rate"]:
        raise ValueError("bucket_column must be either 'count' or 'sample_tweet_rate'")
    
    users_agg = preprocess_data_for_histogram_of_users(
        preprocessed_dir=preprocessed_dir,
        path_to_users_table=path_to_users_table,
        masks=masks,
        log_segments=log_segments,
        count_target="users",
        bucket_column=bucket_column
    )

    users_agg, colors = plot_histogram_of_users_by_tweet_frequency(
        users_agg,
        title=title,
        xlabel=xlabel,
        ylabel=ylabel,
        relative=plot_relative,
        cumulative=plot_cumulative,
        figure_size=figure_size,
        path_to_output=path_to_output_plot
    )

    users_agg_to_excel(
    users_agg,
    colors,
    path_to_output_table= path_to_output_table
    )
    return users_agg

def plot_stacked_histogram_of_corpus_by_user_contributions(
        data_dict: dict,
        title: str = "Corpus Aggregated by Users",
        xlabel: str = "Corpus Contribution",
        yticklabels: list = None,
        relative: bool = False,
        cumulative: bool = False,
        count_unknown: bool = True,
        path_to_output: str=""
):
    def preprocess(data_dict=data_dict):

        def get_barh_values(users_agg):
            tup0 = "relative" if relative else "absolute"
            tup1 = "cum_sum" if cumulative else "sum"
            return users_agg[(tup0, tup1)].values
        
        def get_color_range(data_dict=data_dict, count_unknown=count_unknown):
            cmap = plt.get_cmap('winter')
            n = max([len(data_dict[name]["users_agg"]) for name in data_dict.keys()])
            n -= 1 if count_unknown else n
            color_range = cmap(np.linspace(0, 1, n)[-1::-1])
            return color_range

        def get_colors(users_agg, color_range=get_color_range(data_dict), count_unknown=count_unknown):
            l = len(users_agg)
            l -= 1 if count_unknown else l

            colors = color_range[:l]
            if count_unknown:
                colors = np.vstack([colors, np.array([[1.0, 0.41, 0.71, 1]])])
            return colors
    
        def preprocess_plot(data_dict=data_dict):
            
            for name in data_dict.keys():
                def apply(key, fn, data_dict=data_dict, name=name):
                    data_dict[name][key] = fn(data_dict[name]["users_agg"])

                apply("barh_values", get_barh_values)
                apply("colors", get_colors)
                
            return data_dict
        
        return preprocess_plot()

    def plot(data_dict):
        # Number of datasets to plot
        n = len(data_dict)

        # Create a figure and axes object with adjusted figure size
        fig, ax = plt.subplots(figsize=(10, n * 1.2))  # Adjust the height based on the number of bars

        def get_bars(ax):
            all_bars = []
            y_position = 0  # Starting y-position for the first dataset
            for name in data_dict.keys():
                bottom, bars = 0, []
                
                yvalues = data_dict[name]["barh_values"]
                colors = data_dict[name]["colors"]
                for value, color in zip(yvalues, colors):
                    bar = ax.barh(y_position, value, left=bottom, height=0.8, edgecolor="black", color=color)
                    bottom += value
                    bars.append(bar)
                all_bars.append(bars)
                y_position += 1  # Increment y-position for the next dataset

            return all_bars


        bars = get_bars(ax)

        ax.set_xlabel(xlabel)
        # Set y-ticks to the middle of each histogram
        ax.set_yticks(range(n))
        ax.set_yticklabels(yticklabels if yticklabels else data_dict.keys())  # Assuming the keys of data_dict are descriptive
        ax.set_title(title)
        
        if relative:
            # Adjust x-tick labels to show percentages
            ax.set_xticklabels(['{:.0f}%'.format(x*100) for x in ax.get_xticks()])
        else:
            ax.set_xticklabels(['{:.0f}M'.format(x/1e6) for x in ax.get_xticks()])

        ax.grid(axis='x', linestyle='--', color="red", alpha=0.25)

        plt.tight_layout()
        ax.set_facecolor("#f0f0f0")

        if path_to_output:
            make_dir(os.path.dirname(path_to_output))
            plt.savefig(path_to_output)
        plt.show()


    data_dict = preprocess()
    plot(data_dict)
    return data_dict

def stacked_histogram_of_corpus_contributions_pipeline(
        preprocess_args: dict,
        corpus_args: dict,
        title = "",
        xlabel = "",
        yticklabels = None,
        relative = False,
        cumulative = False,
        count_unknown = True,
        path_to_output_plot = "",
        table_dir = ""
):
    data_dict = {}
    for name, args in corpus_args.items():
        data_dict[name]= {"users_agg": preprocess_data_for_histogram_of_users(**{**preprocess_args, **args})}
    
    data_dict = plot_stacked_histogram_of_corpus_by_user_contributions(
        data_dict,
        title=title,
        xlabel=xlabel,
        yticklabels=yticklabels,
        relative=relative,
        cumulative=cumulative,
        count_unknown=count_unknown,
        path_to_output=path_to_output_plot
    )

    for name, data in data_dict.items():
        users_agg_to_excel(
            data["users_agg"],
            data["colors"],
            path_to_output_table=os.path.join(table_dir, f"{name}.xlsx")
        )
