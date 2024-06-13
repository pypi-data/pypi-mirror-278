from datetime import datetime
import pytz
import pandas as pd
from ntpath import basename
from process_twarc.util import get_output_path, save_to_parquet, concat_dataset, save_dict, get_files, load_dict
from itertools import chain
from typing import Tuple, Dict, Any, List
from tqdm import tqdm
import re
from unicodeblock.blocks import of as block
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

def get_file_type(file_path: str) -> str:
    """
    Get the file type of a file.

    Args:
        file_path (str): The path to the file.

    Returns:
        str: The file type of the file.
    """
    return file_path.split(".")[-1]

def convert_utc_to_jct(utc_datetime_string):
    """
    Convert a UTC datetime string to JCT (Japan Standard Time) and return the JCT datetime string.

    Args:
        utc_datetime_string (str): A string representing the input UTC datetime in the format 
            "%Y-%m-%dT%H:%M:%S.%fZ".

    Returns:
        str: A string representing the converted JCT datetime in the format "%Y-%m-%dT%H:%M:%S".
    """
    utc_datetime = datetime.strptime(utc_datetime_string, "%Y-%m-%dT%H:%M:%S.%fZ")
    
    # Set the timezone to UTC
    utc_timezone = pytz.timezone('UTC')
    utc_datetime = utc_timezone.localize(utc_datetime)
    
    # Convert the datetime to JCT (Japan Standard Time)
    jct_timezone = pytz.timezone('Asia/Tokyo')
    jct_datetime = utc_datetime.astimezone(jct_timezone)
    
    # Format the JCT datetime as a string without milliseconds
    jct_datetime_string = jct_datetime.strftime("%Y-%m-%dT%H:%M:%S")
    
    return jct_datetime_string

def build_rich_dataset(file_path: str, duplicate_ids: set, output_dir: str=""):
    """
    Given a path to a jsonl file packaged by Twarc, prepare a "rich" dataset with
    the following columns:
    
    Args:
        file_path: Path to the Twarc jsonl file.
        output_dir (str): Path to the directory where the file is to be saved.
    
    Returns:
        Paruqet table, with the columns:
            "tweet_id" (str): unique value assigned to the tweet by Twitter,
            "text" (str): text of the tweet, users and urls masked,
            "conversation_id" (str): tweet_id of the tweet at the root of the coversation tree,
            "place_id" (str): unique value assigned to place where the Tweet was made,
            "has_url" (bool): whether or not the Tweet has a url,
            "has_mention" (bool): whether or not the Tweet mentions another user,
            "has_hashtag" (bool): whether or not the Tweet has a hashtag
            "duplicate" (bool): whether or not the Tweet has been flagged as a duplicate
    """
    file_type = get_file_type(file_path)
    old_df = None
    if file_type == "json":
        old_df = pd.read_json(file_path, encoding="utf-8")
    elif file_type == "jsonl":
        old_df = pd.read_json(file_path, lines=True, encoding="utf-8")

    if old_df is not None:
        new_df = {
            "tweet_id": [],
            "text": [],
            "conversation_id": [],
            "user_id": [],
            "created_at": [],
            "place_id": [],
            "possibly_sensitive": [],
            "has_url": [],
            "has_mention": [],
            "has_hashtag": [],
            "duplicate": []
        }

        for i in range(len(old_df["data"])):
            for j in range(len(old_df["data"][i])):
                data = old_df["data"][i][j]

                tweet_id = str(data["id"])
                conversation_id = data["conversation_id"]
                user_id = data["author_id"]
                text = data["text"]
                created_at = convert_utc_to_jct(data["created_at"])
                if "geo" in data.keys() and "place_id" in data["geo"].keys():
                    place_id = data["geo"]["place_id"]
                else:
                    place_id = "None"
                possibly_sensitive = data["possibly_sensitive"]

                has_url, has_mention, has_hashtag = False, False, False
                if "entities" in data.keys():
                    entities = data["entities"]
                    if "urls" in entities.keys():
                        has_url = True
                    if "mentions" in entities.keys():
                        has_mention = True
                    if "hashtags" in entities.keys():
                        has_hashtag = True

                duplicate = tweet_id in duplicate_ids

                new_df["tweet_id"].append(tweet_id)
                new_df["text"].append(text)
                new_df["conversation_id"].append(conversation_id)
                new_df["user_id"].append(user_id)
                new_df["created_at"].append(created_at)
                new_df["place_id"].append(place_id)
                new_df["possibly_sensitive"].append(possibly_sensitive)
                new_df["has_url"].append(has_url)
                new_df["has_mention"].append(has_mention)
                new_df["has_hashtag"].append(has_hashtag)
                new_df["duplicate"].append(duplicate)

        new_df = pd.DataFrame(new_df)

        if output_dir:
            output_path = get_output_path(file_path, output_dir, file_type="parquet")
            save_to_parquet(new_df, output_path)
        return new_df

def flatten_users_and_places(file_path: str, output_dir: str) -> Tuple[Dict[str, Any], Dict[str, Any]]:
    """
    Extracts the "includes" column from a JSON or JSONL file, condenses batches into a simple dictionary,
    and returns separate dictionaries for users and places.

    Args:
        file_path (str): The path to the JSON or JSONL file.
        output_dir (str): The path to the directory where dictionaries are to be saved.


    Returns:
        Tuple[Dict[str, Any], Dict[str, Any]]: A tuple containing two dictionaries:
            - user_dict: A dictionary mapping user IDs to user data.
            - place_dict: A dictionary mapping place IDs to place data.
    """

    file_type = get_file_type(file_path)
    if file_type == "jsonl":
        df = pd.read_json(file_path, lines=True, encoding="utf-8")
    elif file_type == "json":
        df = pd.read_json(file_path, encoding="utf-8")
    else:
        df = None

    if isinstance(df, pd.DataFrame):
        includes = df["includes"]
        user_dict = {}
        place_dict = {}

        for i in range(len(includes)):
            users = includes[i]["users"]
            places = includes[i]["places"]

            user_id2data = {k: v for k, v in zip([user.pop("id") for user in users], users)}
            place_id2data = {k: v for k, v in zip([place.pop("id") for place in places], places)}
            user_dict.update(user_id2data)
            place_dict.update(place_id2data)
        
        user_output_path = get_output_path(file_path, f"{output_dir}/users")
        place_output_path = get_output_path(file_path, f"{output_dir}/places")

        save_dict(user_dict, user_output_path)
        save_dict(place_dict, place_output_path)
        return
    
def count_stats(rich_dir, output_dir):
    id_columns = ["user_id", "place_id"]

    target_columns = lambda id_column: [id_column, "possibly_sensitive", "has_url", "has_mention", "has_hashtag", "duplicate"]
    rich_files = get_files(rich_dir)

    for id_column in id_columns:
        if id_column == "user_id":
            path_to_output = f"{output_dir}/user-stats.json"
        elif id_column == "place_id":
            path_to_output = f"{output_dir}/place-stats.json"
        
        stats = {}
        rich_data = concat_dataset(rich_files, columns=target_columns(id_column))
        grouped_data = rich_data.groupby(id_column)
        for id_, group in tqdm(grouped_data, desc="Computing sample_tweet_count, possibly_sensitive_count, url_count, mention_count, hashtag_count, duplicate_count"):
            stats.update(
                {id_:{
                    "sample_tweet_count": len(group),
                    "possibly_sensitive_count": sum(group["possibly_sensitive"]),
                    "url_count": sum(group["has_url"]),
                    "mention_count": sum(group["has_mention"]),
                    "hashtag_count": sum(group["has_hashtag"]),
                    "duplicate_count": sum(group["duplicate"])
            }})
        
        save_dict(stats, path_to_output)
        print (f"Stats saved to {path_to_output}.")
    return

def filter_dictionary(dictionary:dict, target_keys:(str|List[str])):
    """"
    The inlcudes column contains more users and places then are present in the dataset.

    Given a dictionary, and a list of target keys, reduce the dictionary down to entries
    with the target keys.

    Args:
        dictionary(dict): Dictionary to be filtered.
        target_keys( (str|list(str)) ): List of keys to be included in the filtered dictionary.

    Returns:
        filtered_dictionary(dict): Dictionary with entries for only the target keys. 
    """
    return {key: dictionary[key] for key in target_keys if key in dictionary}

def tabulate_user_data(user_dict, user_ids, user_stats):
    user_dict = filter_dictionary(user_dict, user_ids)

    user_table = {
        "user_id": [],
        "username": [],
        "link": [],
        "created_at": [],
        "followers_count": [],
        "following_count": [],
        "listed_count": [],
        "lifetime_tweet_count": [],
        "lifetime_day_count": [],
        "lifetime_tweet_rate": [],
        "sample_tweet_count": [],
        "sample_day_count": [],  
        "sample_tweet_rate": [],
        "possibly_sensitive_count": [],
        "url_count": [],
        "mention_count": [],
        "hashtag_count": [],
        "duplicate_count": []
    }

    sample_end_date = datetime(2023, 5, 8)  # May 8, 2023
    sample_start_date = datetime(2022, 6, 12)  # June 12, 2022

    for user_id in user_dict.keys():
        data = user_dict[user_id]
        username = data["username"]
        link = f"https://twitter.com/{username}"

        created_at = data["created_at"]
        counts = user_stats[user_id]
        public_metrics = data["public_metrics"]

        followers_count = public_metrics["followers_count"]
        following_count = public_metrics["following_count"]
        lifetime_tweet_count = public_metrics["tweet_count"]
        sample_tweet_count = counts["sample_tweet_count"]
        listed_count = public_metrics["listed_count"]

        possibly_sensitive_count = counts["possibly_sensitive_count"]
        url_count = counts["url_count"]
        mention_count = counts["mention_count"]
        hashtag_count = counts["hashtag_count"]
        duplicate_count = counts["duplicate_count"]

        created_at_date = datetime.strptime(created_at, "%Y-%m-%dT%H:%M:%S.%fZ")
        lifetime_day_count = (sample_end_date - created_at_date).days

        if created_at_date >= sample_start_date:
            sample_day_count = (sample_end_date - created_at_date).days
        else:
            sample_day_count = (sample_end_date - sample_start_date).days

        lifetime_tweet_rate = int(lifetime_tweet_count / lifetime_day_count) if lifetime_day_count > 0 else 0
        sample_tweet_rate = int(sample_tweet_count / sample_day_count) if sample_day_count > 0 else 0

        user_table["user_id"].append(user_id)
        user_table["username"].append(username)
        user_table["link"].append(link)
        user_table["created_at"].append(created_at)
        user_table["followers_count"].append(followers_count)
        user_table["following_count"].append(following_count)
        user_table["listed_count"].append(listed_count)
        user_table["lifetime_tweet_count"].append(lifetime_tweet_count)
        user_table["lifetime_day_count"].append(lifetime_day_count)
        user_table["lifetime_tweet_rate"].append(lifetime_tweet_rate)
        user_table["sample_day_count"].append(sample_day_count)
        user_table["sample_tweet_count"].append(sample_tweet_count)
        user_table["sample_tweet_rate"].append(sample_tweet_rate)
        user_table["possibly_sensitive_count"].append(possibly_sensitive_count)
        user_table["url_count"].append(url_count)
        user_table["mention_count"].append(mention_count)
        user_table["hashtag_count"].append(hashtag_count)
        user_table["duplicate_count"].append(duplicate_count)
        
    user_table = pd.DataFrame(user_table)

    return user_table

def tabulate_place_data(place_dict, place_ids, place_stats):
    place_dict = filter_dictionary(place_dict, place_ids)

    place_table = {
        "place_id": [],
        "full_name": [],
        "name": [],
        "place_type": [],
        "sample_tweet_count": [],
        "possibly_sensitive_count": [],
        "url_count": [],
        "mention_count": [],
        "hashtag_count": [],
        "duplicate_count": [],
        "bbox": []
    }

    for place_id in place_dict.keys():
        counts = place_stats[place_id]

        data = place_dict[place_id]
        full_name = data["full_name"]
        name = data["name"]
        place_type = data["place_type"]

        sample_tweet_count = counts["sample_tweet_count"]
        possibly_sensitive_count = counts["possibly_sensitive_count"]
        url_count = counts["url_count"]
        mention_count = counts["mention_count"]
        hashtag_count = counts["hashtag_count"]
        duplicate_count = counts["duplicate_count"]

        bbox = str(data["geo"]["bbox"])

        place_table["place_id"].append(place_id)
        place_table["full_name"].append(full_name)
        place_table["name"].append(name)
        place_table["place_type"].append(place_type)
        place_table["sample_tweet_count"].append(sample_tweet_count)
        place_table["possibly_sensitive_count"].append(possibly_sensitive_count)
        place_table["url_count"].append(url_count)
        place_table["mention_count"].append(mention_count)
        place_table["hashtag_count"].append(hashtag_count)
        place_table["duplicate_count"].append(duplicate_count )
        place_table["bbox"].append(bbox)

    place_table = pd.DataFrame(place_table)
    return place_table


def tabulate_process(tabulate_method, data_dir: str, stats_dir: str, output_dir: str):

    if tabulate_method == tabulate_user_data:
        path_to_data = f"{data_dir}/users"
        stats = load_dict(f"{stats_dir}/user-stats.json")
        path_to_output = f"{output_dir}/users.csv"

    elif tabulate_method == tabulate_place_data:
        path_to_data = f"{data_dir}/places"
        stats = load_dict(f"{stats_dir}/place-stats.json")
        path_to_output = f"{output_dir}/places.csv"

    print("Compiling id list. . .")
    id_list = set(stats.keys())
    print("ID list compiled!")
    sort_by_basename = lambda file_paths: sorted(file_paths, key=lambda file_path: basename(file_path), reverse=True)
    file_paths = sort_by_basename(get_files(path_to_data))
    last_file = file_paths[-1]

    table = pd.DataFrame()
    for file_path in tqdm(file_paths, desc="Tabulating"):
        if id_list:
            data = load_dict(file_path)
            all_ids = list(data.keys())
            target_ids = [id_ for id_ in all_ids if id_ in id_list]
            id_list.difference_update(target_ids)
            chunk = tabulate_method(data, target_ids, stats)

            table = pd.concat([table, chunk])

            if (not id_list) or (file_path==last_file):
                table.to_csv(path_to_output, encoding="utf-8-sig", index=False)

            print(f"Searching for {len(id_list)} remaining IDs.")
        else:
            print("IDs Tabulated.")
            print(f"Total Places: {len(table)}")
            table.to_csv(path_to_output, encoding="utf-8-sig", index=False)
    return

def build_vocab_table(
    tokenizers: dict,
    path_to_simplified_blocks: str,
    token_counts_dir: str,
    path_to_upos_tags: str,
    path_to_jpos_tags: str,
    path_to_output: str,
    special_tokens: list = ["[UNK]", "[PAD]", "[CLS]", "[SEP]", "[MASK]", "[URL]", "[USER]", "\n"]
):
    def init_df():
        get_vocab = lambda tok: [token for token  in list(tok.get_vocab().keys()) if not re.fullmatch("\[unused\d+\]", token)]
        vocab = pd.DataFrame(columns=["token"])
        for name, tok in tokenizers.items():
            df = pd.DataFrame({
                "token": get_vocab(tok),
                name: True
            })
            vocab = vocab.merge(df, on="token", how="outer")
        vocab = vocab[vocab["token"].apply(lambda x: len(x) > 0)]
        vocab["token"] = vocab["token"].astype(str)
        vocab.fillna(False, inplace=True)
        return vocab
    
    def classify_char_family(vocab):

        def reverse_mapping(original_dict):
            reversed_dict = {}
            for key, value_list in original_dict.items():
                for value in value_list:
                    # Assign or append the original key to the reversed dictionary
                    if value in reversed_dict:
                        reversed_dict[value].append(key)
                    else:
                        reversed_dict[value] = [key]
            return reversed_dict

        simplified = load_dict(path_to_simplified_blocks)
        simplified = reverse_mapping(simplified)

        def get_blocks_fn(token):
            if token in special_tokens:
                return ["SPECIAL"]
            else:
                token = token.replace("##", "")
                blocks = [block(c) if block(c) else "PICTOGRAPHS" for c in token]
                blocks = list(set(chain.from_iterable([simplified.get(b, [b]) for b in blocks])))

                return blocks
        
        def classify_fn(blocks):
            if len(blocks) == 1:
                return blocks[0]
            priorty = ["KANJI", "HIRAGANA", "KATAKANA", "ROMAJI", "PICTOGRAPH", "SYMBOL"]
            for block in priorty:
                if block in blocks:
                    return block
        
        vocab["blocks"] = vocab["token"].apply(get_blocks_fn)
        vocab["char_family"] = vocab["blocks"].apply(classify_fn)
        vocab.drop(columns="blocks", inplace=True)
        return vocab
    
    def get_token_counts(vocab):
        files = get_files(token_counts_dir)
        for file in files:
            name = basename(file).split(".")[0]
            df = pd.read_csv(file)[["token", "count"]].dropna()
            df.rename(columns={"count": f"{name}_count"}, inplace=True)
            vocab = vocab.merge(df, on="token", how="left")
        vocab.fillna(0, inplace=True)
        for col in [col for col in vocab.columns if "count" in col]:
            vocab[col] = vocab[col].astype(int)
        return vocab
    
    def get_pos_tags(vocab):
        upos = pd.read_csv(path_to_upos_tags)[["token", "UPOS"]]
        jpos = pd.read_csv(path_to_jpos_tags)[["token", "JPOS"]]
        vocab = vocab.merge(upos, on="token", how="left")
        vocab = vocab.merge(jpos, on="token", how="left")
        vocab[["JPOS", "UPOS"]] = vocab[["JPOS", "UPOS"]].fillna("N/A")
        return vocab

    vocab = init_df()
    vocab = classify_char_family(vocab)
    vocab = get_token_counts(vocab)
    vocab = get_pos_tags(vocab)

    vocab.to_csv(path_to_output, index=False)
    return vocab

def df_to_excel_with_styling(df, filename="styled_table.xlsx"):
    # Create a workbook and a worksheet
    wb = openpyxl.Workbook()
    ws = wb.active

    num_columns = df.shape[1]
    a4_width_units = 8.27 * 7  # Approximate conversion to Excel's column width units
    column_width = a4_width_units / num_columns  # Divide by the number of columns

    # Convert DataFrame to rows in Excel
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)

            # Center alignment for all cells
            cell.alignment = Alignment(horizontal="center", vertical="center")

            # Header row styling
            if r_idx == 1:
                cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")  # Navy background
                cell.font = Font(color="FFFFFF", bold=True)  # Bold white text

            # Adjust column width
            ws.column_dimensions[openpyxl.utils.get_column_letter(c_idx)].width = column_width

    # Save the workbook
    wb.save(filename)
    print(f"Table saved to {filename}")
