# -*- coding:utf-8 -*-
"""
@Time : 2023/3/22
@Author : skyoceanchen
@TEL: 18916403796
@File : excel_operation.py
@PRODUCT_NAME : PyCharm
"""
import pandas as pd
from basic_type_operations.list_operation import ListOperation
from openpyxl import Workbook, load_workbook
from openpyxl.cell import MergedCell, ReadOnlyCell
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment  # 对齐方式
from openpyxl.utils import FORMULAE  # 公式，但是可以检查公式的名称
import openpyxl
import datetime as dt
import os
from django.conf import settings
from openpyxl.styles import Border  # 边框
from openpyxl.styles import Side
from openpyxl.styles import Font  # 字体
from openpyxl.utils import get_column_letter, column_index_from_string  # 获取列的头 get_column_letter(a)=A

"""
=HYPERLINK("https://www.baidu.com","百度收缩")# 插入超链接
"""


class ExportXlsx(object):  # 导出excel 居中并自适应列宽
    def create_data_img_xlsx(self, data_list, file_type=None, ):
        wb = Workbook()
        sheet = wb.active
        row = len(data_list)
        col = len(data_list[0])
        for j in range(1, col + 1):
            self.cell_alignment(sheet.cell(1, j, value=data_list[0][j - 1]))
        column_letter_width = 25
        for i in range(2, row + 1):
            for j in range(1, len(data_list[i - 1]) + 1):
                value = data_list[i - 1][j - 1]
                column_letter = openpyxl.utils.cell.get_column_letter(j)
                if isinstance(value, str):
                    if value.endswith('png') or value.endswith('jpg') or value.endswith("PNG") or value.endswith("JPG"):
                        if os.path.exists(value):
                            location = f"{column_letter}{i}"
                            img = Image(value)
                            img_times = int(img.height / 90) + 1
                            height = int(img.height / img_times)
                            width = int(img.width / img_times)
                            image_size = (width, height,)
                            img.width, img.height = (width, height,)  # 这两个属性分别是对应添加图片的宽高
                            sheet.row_dimensions[i].height = height
                            sheet.column_dimensions[column_letter].width = column_letter_width  # 修改列的列宽
                            sheet.add_image(img, f'{location}', )  # 向单元格内指定添加图片
                    else:
                        self.cell_alignment(sheet.cell(i, j, value=value))
                        sheet.column_dimensions[column_letter].width = len(value.encode('utf-8'))
                else:
                    self.cell_alignment(sheet.cell(i, j, value=value))
                    # sheet.column_dimensions[column_letter].width = len(str(value).encode('utf-8'))

        current_time = dt.datetime.now().strftime("%Y%m%d%H%M%S")
        filename = f"{current_time}.xlsx"
        thumb_path = os.path.join(settings.MEDIA_ROOT, file_type).replace('\\', '/')
        dest = os.path.join(thumb_path, filename).replace('\\', '/')
        pathRela = os.path.join(file_type, filename).replace('\\', '/')
        if not os.path.exists(thumb_path):
            os.makedirs(thumb_path)
        wb.save(f'{dest}')
        return filename, dest, pathRela

    def create_data_xlsx(self, data_list, file_type=None, ):
        wb = Workbook()
        sheet = wb.active
        row = len(data_list)
        col = len(data_list[0])
        for i in range(1, row + 1):
            for j in range(1, col + 1):
                self.cell_alignment(sheet.cell(i, j, value=data_list[i - 1][j - 1]))
        self.auto_width(sheet)
        current_time = dt.datetime.now().strftime("%Y%m%d%H%M%S")
        filename = f"{current_time}.xlsx"
        thumb_path = os.path.join(settings.MEDIA_ROOT, file_type).replace('\\', '/')
        dest = os.path.join(thumb_path, filename).replace('\\', '/')
        pathRela = os.path.join(file_type, filename).replace('\\', '/')
        if not os.path.exists(thumb_path):
            os.makedirs(thumb_path)
        wb.save(f'{dest}')
        return filename, dest, pathRela

    # <editor-fold desc="自适应列宽">
    def auto_width(self, ws):
        lks = []  # 英文变量太费劲，用汉语首字拼音代替
        for i in range(1, ws.max_column + 1):  # 每列循环
            lk = 1  # 定义初始列宽，并在每个行循环完成后重置
            for j in range(1, ws.max_row + 1):  # 每行循环
                sz = ws.cell(row=j, column=i).value  # 每个单元格内容
                if isinstance(sz, str):  # 中文占用多个字节，需要分开处理
                    lk1 = len(sz.encode('utf-8'))  # gbk解码一个中文两字节，utf-8一个中文三字节，gbk合适
                else:
                    lk1 = len(str(sz))
                if lk < lk1:
                    lk = lk1  # 借助每行循环将最大值存入lk中
            lks.append(lk)  # 将每列最大宽度加入列表。（犯了一个错，用lks = lks.append(lk)报错，append会修改列表变量，返回值none，而none不能继续用append方法）
        # 第二步：设置列宽
        for i in range(1, ws.max_column + 1):
            k = get_column_letter(i)  # 将数字转化为列名,26个字母以内也可以用[chr(i).upper() for i in range(97, 123)]，不用导入模块
            ws.column_dimensions[k].width = lks[i - 1] + 2  # 设置列宽，一般加两个字节宽度，可以根据实际情况灵活调整

    # 有汉字的自适应
    def column_dimensions_width(self, sheet, data):
        """
        data 表格所有数据/部分数据/二维
        :param ws:
        :param data:
        :return:
        """
        lis = ListOperation.dimensional_list_two(data)
        for index, li in enumerate(lis):
            # len_data = []
            # for m in li:
            #     if isinstance(m, float):
            #         m = round(m, 3)
            #     len_data.append(len(str(m)))
            # max_len = max(len_data)
            # ws.column_dimensions[get_column_letter(index + 1)].width = max_len * 2
            df = pd.DataFrame(li)
            # # ws自动设置列宽
            df_len = df.apply(lambda x: [(len(str(i).encode('utf-8')) - len(str(i))) / 2 + len(str(i)) for i in li],
                              axis=0)
            df_len_max = df_len.apply(lambda x: max(x), axis=0)
            column_letter = get_column_letter(index + 1)
            # 列的宽度
            columns_length = (len(str(index).encode('utf-8')) - len(str(index))) / 2 + len(str(index))
            data_max_length = df_len_max[0]
            column_width = [data_max_length if columns_length < data_max_length else columns_length][0]
            column_width = [column_width if column_width <= 50 else 50][0] + 3  # 列宽不能超过50
            # 更改列的宽度
            sheet.column_dimensions['{}'.format(column_letter)].width = column_width
        return sheet

    # </editor-fold>
    # <editor-fold desc="对齐方式-水平居中-垂直居中">
    def cell_alignment(self, cell):
        """
        horizontal_alignments = (
            "general:上", "left", "center", "right", "fill：满", "justify：两端对齐", "centerContinuous：中心连续",
            "distributed：分散",)
        vertical_aligments = ("top", "center", "bottom", "justify", "distributed",)
        """
        alignment = Alignment(
            # ("general", "left", "center", "right", "fill", "justify", "centerContinuous", "distributed",)
            horizontal='center',  # 水平居中
            # ( "top", "center", "bottom", "justify", "distributed",)
            vertical='center',  # 垂直居中
            textRotation=0,  # 文字旋转range(181) 0-180°
            wrapText=True,  # 自动换行 True/Flase
            shrinkToFit=True,  # 是否适应列宽 True/Flase
            indent=0,  # 缩进
            relativeIndent=0,  # 相对缩进
            justifyLastLine=None,  # 证明最后一行 True/Flase
            readingOrder=0,  # 阅读顺序
        )
        cell.alignment = alignment
        return cell

    # </editor-fold>
    def read_xlsx(self, path):
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        max_column = ws.max_column
        max_row = ws.max_row
        data_lis = []
        for index, i in enumerate(ws):
            # if index > 0:
            lis = [j.value for j in i]
            # lis[1] = dt.datetime.strptime(lis[1], "%Y-%m-%d %H:%M:%S")
            data_lis.append(lis)
        return data_lis


# openpyxl操作
class OpenpyxlCell(object):
    # <editor-fold desc="边框设置">
    def cell_border(self, cell):
        """
            style = NoneSet(values=('dashDot','dashDotDot', 'dashed','dotted',
                                'double','hair', 'medium', 'mediumDashDot', 'mediumDashDotDot',
                                'mediumDashed', 'slantDashDot', 'thick', 'thin')
                        )
        :param cell:
        :return:
        """
        border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000'),
            diagonal=Side(style='thin', color='000000'),
            diagonal_direction=0,  # 对角线方向
            vertical=Side(style='thin', color='000000'),
            horizontal=Side(style='thin', color='000000'),
            diagonalUp=False,  # True/False
            diagonalDown=False,  # True/False
            # outline=Side(style='medium', color='000000'),
            outline=True,  # True/False
            start=None,  # Side(style='thin', color='000000')
            end=None)  # Side(style='thin', color='000000')
        cell.border = border
        return cell

    # </editor-fold>
    # <editor-fold desc="文字设置">
    def cell_font(self, cell):
        """
        # 默认字体
        DEFAULT_FONT = Font(name="Calibri", sz=11, family=2, b=False, i=False,
                        color=Color(theme=1), scheme="minor")
        :param cell:
        :return:
        """
        font = Font(
            name='宋体',  # 字体名字/Calibri/宋体
            family=None,  # min=0, max=14
            size=30,  # 字体大小
            bold=True,  # 是否加粗
            italic=True,  # 斜体
            underline='none',  # 下划线 ('single：单线', 'double：双线', 'singleAccounting','doubleAccounting')
            color='000000',  # 颜色,可以用colors中的颜色
            charset=None,  # 字符集；字元集
            scheme=None,  # 计划 概形 方案  ("major":主要的, "minor":次要)
            strikethrough=None,  # 删除线，加删除线
            vertAlign=None,  # 纵向对齐方式('superscript:上标 标在上方的附加字', 'subscript:脚注', 'baseline:底线')
            outline=None,  # 轮廓 轮廓线 大纲
            shadow=None,  # 阴影 影子 影
            condense=None,  # 压缩 浓缩 紧缩
            extend=None  # 延伸 延长 扩展
        )
        cell.font = font
        return cell

    # </editor-fold>
    # <editor-fold desc="对齐方式-水平居中-垂直居中">
    def cell_alignment(self, cell):
        horizontal_alignments = (
            "general:上", "left", "center", "right", "fill：满", "justify：两端对齐", "centerContinuous：中心连续",
            "distributed：分散",)
        vertical_aligments = ("top", "center", "bottom", "justify", "distributed",)
        alignment = Alignment(
            # ("general", "left", "center", "right", "fill", "justify", "centerContinuous", "distributed",)
            horizontal='center',  # 水平居中
            # ( "top", "center", "bottom", "justify", "distributed",)
            vertical='center',  # 垂直居中
            textRotation=0,  # 文字旋转range(181) 0-180°
            wrapText=True,  # 自动换行 True/Flase
            shrinkToFit=True,  # 是否适应列宽 True/Flase
            indent=0,  # 缩进
            relativeIndent=0,  # 相对缩进
            justifyLastLine=None,  # 证明最后一行 True/Flase
            readingOrder=0,  # 阅读顺序
        )
        cell.alignment = alignment
        return cell

    # </editor-fold>


class openpyxlSheet(object):
    # <editor-fold desc="工作表保护 不可更改被保护的工作表">
    def password_sheet(self, sheet, enable=True, password=None):
        """

        :param ws:
        :param enable: True保护/Flase 不保护
        :param password:str 密码
        :return:
        """
        # 工作表保护 不可更改被保护的工作表
        # 也可以通过在openpyxl.worksheet.protection.SheetProtection对象上设置属性来锁定工作表的各个方面 。
        # 与工作簿保护不同，可以使用或不使用密码来启用工作表保护。
        # 使用openpxyl.worksheet.protection.SheetProtection.sheet属性或调用enable（）或disable（）可启用工作表保护 ：
        # 如果未指定密码，则用户无需指定密码即可禁用配置的工作表保护。
        # 否则，他们必须提供密码才能更改已配置的保护。
        # 使用openpxyl.worksheet.protection.SheetProtection.password()属性设置密码
        if enable:
            sheet.protection.password = password
            sheet.protection.sheet = True
            sheet.protection.enable()
        else:
            sheet.protection.password = password
            sheet.protection.sheet = True
            sheet.protection.disable()

    # </editor-fold>
    # <editor-fold desc="自适应列宽">
    def auto_width(self, sheet):
        lks = []  # 英文变量太费劲，用汉语首字拼音代替
        for i in range(1, sheet.max_column + 1):  # 每列循环
            lk = 1  # 定义初始列宽，并在每个行循环完成后重置
            for j in range(1, sheet.max_row + 1):  # 每行循环
                sz = sheet.cell(row=j, column=i).value  # 每个单元格内容
                if isinstance(sz, str):  # 中文占用多个字节，需要分开处理
                    lk1 = len(sz.encode('utf-8'))  # gbk解码一个中文两字节，utf-8一个中文三字节，gbk合适
                else:
                    lk1 = len(str(sz))
                if lk < lk1:
                    lk = lk1  # 借助每行循环将最大值存入lk中
            lks.append(lk)  # 将每列最大宽度加入列表。（犯了一个错，用lks = lks.append(lk)报错，append会修改列表变量，返回值none，而none不能继续用append方法）
        # 第二步：设置列宽
        for i in range(1, sheet.max_column + 1):
            k = get_column_letter(i)  # 将数字转化为列名,26个字母以内也可以用[chr(i).upper() for i in range(97, 123)]，不用导入模块
            sheet.column_dimensions[k].width = lks[i - 1] + 2  # 设置列宽，一般加两个字节宽度，可以根据实际情况灵活调整

    # </editor-fold>


class OpenpyxlOperation(openpyxlSheet, OpenpyxlCell):
    def __init__(self, path, read_only=False,
                 data_only=False, keep_links=True, rich_text=False, read=True, write=False, read_Write=None,
                 keep_vba=False,
                 sheet_title=None, tabColor=None):
        """
        :param filename:
        :param read_only:bool 优化阅读，内容不能编辑
        :param keep_vba:bool 控制是否保留任何Visual Basic元素（默认）。如果保留它们，则它们仍不可编辑
        :param data_only:bool 控制具有公式的单元格是否具有公式（默认值）或Excel上次读取工作表时存储的值。
        :param keep_links:bool 是否保留到外部工作簿的链接。默认值为True
        :param sheet_title:
        :param tabColor:
         """
        self.path = path
        self.read = read
        self.write = write
        self.read_only = read_only
        self.data_only = data_only
        self.keep_vba = keep_vba
        self.keep_links = keep_links
        self.rich_text = rich_text
        self.active()

    def get_sheet_title(self):
        return self.wb.sheetnames

    def active(self):
        # 阅读现有的工作簿
        if self.read:
            self.wb = load_workbook(self.path, read_only=False,
                                    data_only=False, keep_links=True, rich_text=False)

        elif self.write:
            self.wb = Workbook()
            # if sheet_title:
            #     ws.title = sheet_title
            # if tabColor:
            #     ws.sheet_properties.tabColor = tabColor
        sheet = self.wb.active
        return sheet

    # <editor-fold desc="选择一个工作表">
    def select_sheet(self, sheet_title):
        sheet = self.wb[sheet_title]
        return sheet

    # </editor-fold>
    # <editor-fold desc="创建新的工作表">
    def create_sheet(self, sheet_name, index=None):
        """
        :param sheet_name:
        :param index: None # 在末尾插入(默认) 0  # 第一个位置插入 -1 # 插入到倒数第二位置
        :return:
        """
        sheet = self.wb.create_sheet(title=sheet_name, index=index)
        return sheet

    # </editor-fold>
    # <editor-fold desc="单个工作簿中创建工作表的副本:复制工作表">
    def copy_worksheet(self, sheet):
        """
        仅复制单元格（包括值，样式，超链接和注释）和某些工作表属性（包括尺寸，格式和属性）。不复制所有其他工作簿/工作表属性-例如图像，图表。

您也不能在工作簿之间复制工作表。如果工作簿以只读或仅写 模式打开，则不能复制工作表。
        :param source:
        :return:
        """
        self.wb.copy_worksheet(sheet)

    # </editor-fold>
    # <editor-fold desc="获取所有的工作表名">
    def get_all_sheet(self, ):
        return self.wb.sheetnames

    # </editor-fold>
    # <editor-fold desc="取值:根据位置取值">
    def get_sheet_value(self, sheet, location, get_value=True):
        if get_value:
            if isinstance(location, int):
                # 返回第一行,对象
                return [i.value for i in sheet[location]]
            elif location[-1].isdigit():
                return sheet[location].value
            else:
                # 返回第一列 数据
                return [i.value for i in sheet[location]]
        else:
            if isinstance(location, int):
                # 返回第一行,对象
                return list(sheet[location])
            elif location[-1].isdigit():
                return sheet[location]
            else:
                # 返回第一列 数据
                return list(sheet[location])

    # </editor-fold>
    # <editor-fold desc="赋予值/修改值">
    def post_sheet_value(self, sheet, location, value):
        sheet[location] = value

    # </editor-fold>
    # <editor-fold desc="cell赋予值">
    def cell(self, sheet, value=None, top_distance=0, left_distance=0):
        for row in range(len(value)):
            for column in range(len(value[row])):
                sheet.cell(row=row + 1 + top_distance, column=column + 1 + left_distance, value=value[row][column])

    # </editor-fold>
    # <editor-fold desc=" col的表格列名：AA列 A列">
    def get_column_letter(self, col):
        return get_column_letter(col)

    # </editor-fold>
    # <editor-fold desc="合并/取消合并单元格">
    # 合并单元格
    def merge_cells(self, sheet, left_col_index=None, right_col_index=None, start_row=None,
                    start_column=None, end_row=None, end_column=None, range_string=None):
        """
        sheet.merge_cells('A2:D2')
        # or equivalently
        sheet.merge_cells(start_row=2, start_column=1, end_row=4, end_column=4)
        :return:
        """
        if left_col_index and left_col_index:
            sheet.merge_cells(range_string='%s:%s' % (left_col_index, right_col_index))
        elif start_row and start_column and end_row and end_column:
            sheet.merge_cells(start_row=start_row, start_column=start_column, end_row=end_row,
                              end_column=end_column)
        return sheet

    # 取消合并单元格
    def unmerge_cells(self, sheet, left_col_index=None, right_col_index=None, start_row=None,
                      start_column=None, end_row=None, end_column=None, range_string=None, ):
        """
        sheet.unmerge_cells('A2:D2')
        sheet.unmerge_cells(start_row=2, start_column=1, end_row=4, end_column=4)
        :return:
        """
        if left_col_index and left_col_index:
            sheet.unmerge_cells(range_string='%s:%s' % (left_col_index, right_col_index))
        elif start_row and start_column and end_row and end_column:
            sheet.unmerge_cells(start_row=start_row, start_column=start_column, end_row=end_row,
                                end_column=end_column)
        return sheet

    # </editor-fold>
    # <editor-fold desc="合并单元格并居中">
    def merge_cells_alignment(self, sheet, left_col_index=None, right_col_index=None, start_row=None,
                              start_column=None,
                              end_row=None, end_column=None):
        """
        range_string:'A2:D2'
        ws.merge_cells('A2:D2')
        # or equivalently
        ws.merge_cells(start_row=2, start_column=1, end_row=4, end_column=4)
        :return:
        """
        if left_col_index and left_col_index:
            sheet.merge_cells(range_string='%s:%s' % (left_col_index, right_col_index))
            cell = sheet[left_col_index]
        elif start_row and start_column and end_row and end_column:
            sheet.merge_cells(start_row=start_row, start_column=start_column, end_row=end_row,
                              end_column=end_column)
            cell = sheet.cell(start_row, start_column)
        else:
            cell = False
        self.cell_alignment(cell)
        return sheet

    # </editor-fold>
    # <editor-fold desc="添加图片">
    def add_image(self, sheet, img_path, location):
        """

        :param sheet:
        :param img_path:
        :param location:可选地为左上角的锚提供一个单元格
        :return:
        """
        img = Image(img_path)
        sheet.add_image(img=img, anchor=location)

    # </editor-fold>
    # <editor-fold desc="折叠（轮廓）行 列">
    # 列折叠
    def column_dimensions(self, sheet, left_col_index=None, right_col_index=None, outline_level=1, hidden=False):
        sheet.column_dimensions.group(left_col_index, right_col_index, outline_level=outline_level, hidden=hidden)

    # 行折叠
    def row_dimensions(self, sheet, left_row_index, right_row_index, outline_level=1, hidden=False):
        sheet.row_dimensions.group(left_row_index, right_row_index, outline_level=outline_level, hidden=hidden)

    # </editor-fold>
    # <editor-fold desc="求和">
    def sum(self, sheet, location, left_index, right_index):
        sheet[location] = "=SUM(%s,%s)" % (left_index, right_index)
        return sheet

    # </editor-fold>
    # <editor-fold desc="工作簿保护">
    def sheet_name_password(self, workbookPassword=None, lockStructure=True):
        # 为防止其他用户查看隐藏的工作表，添加，移动，删除或隐藏工作表以及重命名工作表，
        # 可以使用密码保护工作簿的结构。可以使用openpyxl.workbook.protection.WorkbookProtection.workbookPassword()属性设置密码
        self.wb.security.workbookPassword = workbookPassword
        self.wb.security.lockStructure = lockStructure
        # 工作簿保护

        # # 同样，可以通过设置另一个密码来防止从共享工作簿中删除更改跟踪和更改历史记录。
        # # 可以使用openpyxl.workbook.protection.WorkbookProtection.revisionsPassword()属性设置此密码
        # wb.security.revisionsPassword = '123'
        # # openpyxl.workbook.protection.WorkbookProtection对象上的其他属性精确地控制了什么限制，
        # # 但是只有设置了适当的密码后，这些属性才会强制实施。
        # # 如果需要在不使用默认哈希算法的情况下设置原始密码值，则提供特定的设置器功能-例如
        # hashed_password = 123
        # wb.security.set_workbook_password(hashed_password, already_hashed=True)

    # </editor-fold>


class OpenpyxlReadOperation(OpenpyxlOperation):

    def read_xlsx(self, sheet):
        max_column = sheet.max_column
        max_row = sheet.max_row
        data_lis = []
        for index, i in enumerate(sheet):
            # if index > 0:
            lis = [j.value for j in i]
            # lis[1] = dt.datetime.strptime(lis[1], "%Y-%m-%d %H:%M:%S")
            data_lis.append(lis)
        return data_lis

    # 取消合并行列并每各分开值相同
    def get_excel(self, sheet_name, to_path):
        sheet = self.select_sheet(sheet_name)

        # 获取单元格的值，若为合并单元格，则返回合并区域左上角的值
        def get_cell_val(i, j):
            cell = sheet.cell(i, j)
            # 判断单元格类型是否为合并单元格
            if isinstance(cell, MergedCell):
                for merged_range in sheet.merged_cells.ranges:
                    if cell.coordinate in merged_range:
                        cell = sheet.cell(row=merged_range.min_row, column=merged_range.min_col)
            return cell.value

        rows = sheet.max_row
        cols = sheet.max_column
        pd_lis = []
        for row in range(1, rows + 1):
            lis = []
            for col in range(1, cols + 1):
                lis.append(get_cell_val(row, col))
            pd_lis.append(lis)
        pf = pd.DataFrame(pd_lis)
        pf.to_excel(to_path, index=False, header=False)


class OpenpyxlWriteOperation(OpenpyxlOperation):

    def write(self, sheet, lis):
        for row in lis:
            sheet.append(row)

    def save(self, path):
        self.wb.save(path)
