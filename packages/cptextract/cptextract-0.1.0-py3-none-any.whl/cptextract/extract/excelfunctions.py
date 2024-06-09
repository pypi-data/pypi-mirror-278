def extract_xls_cpt_data(filepath, text_search):
    """Opens CPT xls Excel files and finds the starting row based on the 
    text_seed (e.g. 'H [m]')

    Parameters
    ----------
    filepath : str
        Location of .csv file
    text_search : str
        The keyword to find to signal the start of the data to extract

    Returns
    -------
    list
        a list containing 
        -[0] a dataframe containing all the values extracted from the keyword and beyond
        -[1] warning message
        -[3] the location of the keyword in Excel index
    """
    import xlrd
    import os 
    import pandas as pd

    msg = ""
    book = xlrd.open_workbook(filepath, logfile=open(os.devnull, 'w'))
    count = 0
    for sheet in book.sheets():
        for rowidx in range(sheet.nrows):
            row = sheet.row(rowidx)
            for colidx, cell in enumerate(row):
                if cell.value == text_search :
                    sheet_key = sheet.name
                    skip = rowidx
                    count += 1
                    depth_col = colidx

    if count > 1.0:
        msg = f"more than one depth header may have been found. Please check file!"  
    
    df = pd.read_excel(filepath, sheet_name=sheet_key, skiprows=skip)

    return [df, msg, depth_col]


def extract_xlsx_cpt_data(filepath, text_search):
    """Opens CPT xlsx Excel files and finds the starting row based on the 
    text_seed (e.g. 'H [m]')

    Parameters
    ----------
    filepath : str
        Location of .csv file
    text_search : str
        The keyword to find to signal the start of the data to extract

    Returns
    -------
    list
        a list containing 
        -[0] a dataframe containing all the values extracted from the keyword and beyond
        -[1] warning message
        -[3] the location of the keyword in Excel index
    """
    #load openpyxl library for 'newer xlsx files only
    import openpyxl
    from openpyxl import load_workbook
    import pandas as pd

    count = 0
    msg = " "
    wb = openpyxl.load_workbook(filepath)
    for i in range(len(wb.worksheets)):
        ws = wb.worksheets[i]
       
        for tupleOfCells in ws.iter_rows():
            for cell in tupleOfCells:
                if (cell.value == text_search):
                    skip = cell.row - 1
                    depth_col = cell.column
                    sheet_key = wb.sheetnames[i]
                    count += 1

    if count > 1.0:
        msg = f"more than one depth header may have been found. Please check file!"  

    df = pd.read_excel(filepath, sheet_name=sheet_key, skiprows=skip)
        
    return [df, msg, depth_col-1]